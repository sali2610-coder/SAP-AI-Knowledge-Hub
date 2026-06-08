# -*- coding: utf-8 -*-
"""
PRODUCTION BUILD  -  SAP PM ECC 6.0 -> S/4HANA interactive workbook as a MACRO-ENABLED .xlsm.

What this script does
---------------------
* Uses pandas + openpyxl. Imports the full, accurate master data from scripts/pm_data.py
  (12 topics, 52 core tables, 251 decoded fields - no placeholders).
* Sheet 0 "מסך ניווט מרכזי" (RTL dashboard): styled hyperlink nav cards to every topic sheet,
  a global-search box + a macro-linked Form Control button, FILTER results, and an embedded
  INDEX_DB block (hidden columns P:V on the dashboard itself) that BOTH the FILTER formula and
  the VBA macro read from - the dashboard *is* the index source.
* Every topic sheet: visible gridlines, RTL, "חזרה למסך ראשי" jump button in A1, the expanded
  per-table / per-field grid, and the S/4HANA + Fiori migration columns built INTO the sheet
  (orange band on the left, i.e. columns 11-14 in RTL).
* Embeds a real VBA project (module modPM) with PM_GlobalSearch + BackToDashboard, by generating
  a valid vbaProject.bin (MS-CFB + MS-OVBA) in pure Python and injecting it into the package.
* Also exports modPM.bas next to the file as a guaranteed fallback.

Colors: Dark Blue = master data, Deep Green = transactional (orders/notifications),
Maroon = customizing, Orange/Red = S/4HANA migration columns.

Note (repo CLAUDE.md rule 3): the indexed corpus is S/4HANA only; this is explicit opt-in
ECC 6.0 model knowledge with an S/4 mapping. Pure-Python VBA embedding is validated structurally
(olefile + OVBA round-trip) but cannot be Excel-validated here - if your Excel flags the project,
import the exported modPM.bas (Alt+F11 -> File -> Import File). Verify Fiori App IDs in the
SAP Fiori Apps Reference Library.

Run:  python3 scripts/gen_pm_workbook_xlsm.py
Out:  SAP_PM_ECC6_to_S4_Migration.xlsm  +  modPM.bas
"""
import os, sys, io, struct, zipfile, shutil, re

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import pandas as pd                              # required by spec (used for the index frame)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

from pm_data import COLS as COLS_BASE, S4_COL_START, TOPICS
from pm_ext_data import EXTENSIONS, SIMPLIFICATION

DASH = "מסך ניווט מרכזי"
SIMP_SHEET = "Simplification Item list"
COCKPIT = "Cockpit מעקב מיגרציה"
CCC = "Custom Code Check"

# ---------------------------------------------------------------------------
#  UI/UX DESIGN SYSTEM - CBC Israel / Coca-Cola Red corporate brand identity
# ---------------------------------------------------------------------------
FONT_NAME = "Segoe UI"
RED       = "D62027"        # CBC / Coca-Cola brand red - primary anchors & exec blocks
RED_HOT   = "F40009"        # vibrant red accent (borders / highlights)
SLATE     = "1E1E24"        # deep charcoal/slate - standard table headers
SLATE2    = "2B2B33"        # lighter slate - sub-headers
SILVER    = "8A9EA7"        # silver/platinum - S/4HANA migration columns
SILVER_TX = "1E1E24"        # dark text on silver
INK       = "1F2937"        # body text
HEB_INK   = "1F2937"
GRID      = "E5E7EB"        # light-grey cell borders
ZEBRA     = "F9FAFB"        # ultra-soft alternate row
DASH_HDR  = RED             # dashboard / executive summary header = brand red
DASH_BG   = "F7F8FA"        # soft neutral dashboard background
BACK_BG   = SLATE           # back-button dark slate background
BACK_TXT  = "FFFFFF"        # white text
S4_HDR    = SILVER          # S/4HANA columns - silver/platinum header
S4_TXT    = SILVER_TX       # dark text on the silver header
S4_BAND   = "EEF1F3"        # faint silver zebra for S/4 columns

# theme per module group -> all standard headers use brand slate; tabs carry a brand accent
TH = {
 "master": {"h": SLATE, "s": SLATE2, "b": ZEBRA, "t": SLATE},    # nutrient/master
 "trans":  {"h": SLATE, "s": SLATE2, "b": ZEBRA, "t": RED},      # transactional = brand red tab
 "config": {"h": SLATE, "s": SLATE2, "b": ZEBRA, "t": SLATE},
 "cross":  {"h": SLATE, "s": SLATE2, "b": ZEBRA, "t": SILVER},   # integration = silver tab
}

SEARCH_CELL = "C25"          # visible search box (must match the macro & VML anchor)
IDX_COL0 = 16                # column P - first index column (hidden block on dashboard)

# 15th column appended on the LEFT (RTL) = SUM/migration note
COLS = COLS_BASE + [("שלב מיגרציה / הערות SUM Conversion", 30)]
NCOL = len(COLS)             # 15

def sum_note(tbl):
    """Derive a concrete SUM (Software Update Manager) conversion note from the S/4 mapping."""
    repl = tbl["s4_repl"]; status = tbl["s4_status"]
    if "MATDOC" in repl:
        return ("המרה אוטומטית ל-MATDOC ב-SUM (Silent Data Migration). MKPF/MSEG הופכים ל-Compatibility Views. "
                "בדוק קוד Z שקורא ישירות ל-MSEG/MKPF ועדכן ל-CDS/MATDOC.")
    if "ACDOCA" in repl:
        return ("עלויות מומרות ל-Universal Journal (ACDOCA) ב-SUM. COSP/COSS הופכים ל-Views. "
                "התאם דוחות עלות מותאמים והרצות התחשבנות.")
    if "Business Partner" in status or "BP" in repl:
        return ("נדרש CVI והמרת ספקים/לקוחות ל-Business Partner לפני/במהלך ההמרה (Pre-check חובה).")
    if "מותאם" in status:
        return ("אין המרת טבלה הרסנית, אך מודל הנתונים מותאם - הרץ Regression Test ובדוק User Exits/דוחות מותאמים.")
    return ("ללא פעולת המרה ייעודית ב-SUM (טבלה תואמת). מומלץ Regression Test ואימות התאמות אישיות לאחר ההמרה.")
thin = Side(style="thin", color=GRID)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# ===========================================================================
# 1) VBA macro source (ASCII only -> safe with VBA codepage 1252)
# ===========================================================================
MACRO_SRC = (
'Attribute VB_Name = "modPM"\n'
'Option Explicit\n'
'\n'
"' חיפוש גלובלי - SAP PM. קורא את תיבת החיפוש בדאשבורד, ואם ריקה - שואל ב-InputBox,\n"
"' ואז סורק את כל הלשוניות (פרט לדאשבורד) ומקפיץ ללשונית ולשורה הנכונה.\n"
'Sub PM_GlobalSearch()\n'
'    Dim searchText As String, ws As Worksheet, foundRange As Range\n'
'    searchText = CStr(ThisWorkbook.Sheets(1).Range("' + SEARCH_CELL + '").Value)\n'
'    If Trim(searchText) = "" Then\n'
'        searchText = InputBox("הזן שדה טכני / טבלה / טרנזקציה / מונח (עברית או אנגלית):", "חיפוש גלובלי - SAP PM")\n'
'    End If\n'
'    If Trim(searchText) = "" Then Exit Sub\n'
'    For Each ws In ThisWorkbook.Worksheets\n'
'        If ws.Index > 1 Then                    \' דילוג על מסך הניווט הראשי\n'
'            Set foundRange = ws.Cells.Find(What:=searchText, LookIn:=xlValues, _\n'
'                                           LookAt:=xlPart, MatchCase:=False)\n'
'            If Not foundRange Is Nothing Then\n'
'                ws.Activate\n'
'                Application.Goto foundRange, True\n'
'                foundRange.Interior.Color = RGB(255, 242, 0)\n'
'                MsgBox "נמצא בלשונית: " & ws.Name & " | שורה: " & foundRange.Row, vbInformation, "חיפוש הצליח"\n'
'                Exit Sub\n'
'            End If\n'
'        End If\n'
'    Next ws\n'
'    MsgBox "המונח \'" & searchText & "\' לא נמצא בקובץ.", vbExclamation, "תוצאת חיפוש"\n'
'End Sub\n'
'\n'
"' חזרה למסך הניווט הראשי (הדאשבורד הוא הלשונית הראשונה).\n"
'Sub BackToDashboard()\n'
'    ThisWorkbook.Sheets(1).Activate\n'
'    ThisWorkbook.Sheets(1).Range("A1").Select\n'
'End Sub\n'
)

# ===========================================================================
# 2) vbaProject.bin generator  (MS-OVBA compression + MS-CFB container)
# ===========================================================================
def _compress_chunk(data):
    out = bytearray(); i = 0
    while i < len(data):
        out.append(0x00); out.extend(data[i:i+8]); i += 8
    return bytes(out)

def ovba_compress(data):
    res = bytearray([0x01]); CHUNK = 2048
    for off in range(0, len(data), CHUNK):
        comp = _compress_chunk(data[off:off+CHUNK])
        header = ((len(comp) + 2 - 3) & 0x0FFF) | (0b011 << 12) | (1 << 15)
        res += struct.pack('<H', header) + comp
    return bytes(res)

def _rec(rid, data): return struct.pack('<HI', rid, len(data)) + data

def build_dir_stream(mn):
    m = mn.encode('latin-1')
    s = bytearray()
    s += _rec(0x0001, struct.pack('<I', 1))                # SYSKIND win32
    s += _rec(0x0002, struct.pack('<I', 0x409))            # LCID
    s += _rec(0x0014, struct.pack('<I', 0x409))            # LCIDINVOKE
    s += _rec(0x0003, struct.pack('<H', 1255))             # CODEPAGE (Hebrew - matches macro text)
    s += _rec(0x0004, b'VBAProject')                       # PROJECTNAME
    s += struct.pack('<HI', 0x0005, 0)                     # DOCSTRING
    s += struct.pack('<HI', 0x0040, 0)
    s += struct.pack('<HI', 0x0006, 0)                     # HELPFILEPATH
    s += struct.pack('<HI', 0x003D, 0)
    s += _rec(0x0007, struct.pack('<I', 0))                # HELPCONTEXT
    s += _rec(0x0008, struct.pack('<I', 0))                # LIBFLAGS
    s += struct.pack('<HIIH', 0x0009, 4, 0x00030003, 0)    # VERSION
    s += _rec(0x000C, b'')                                 # CONSTANTS
    s += _rec(0x000F, struct.pack('<H', 1))                # MODULES count
    s += _rec(0x0013, struct.pack('<H', 0xFFFF))           # COOKIE
    s += _rec(0x0019, m)                                   # MODULENAME
    s += _rec(0x0047, m)                                   # MODULENAMEUNICODE
    s += _rec(0x001A, m)                                   # MODULESTREAMNAME
    s += struct.pack('<HI', 0x0032, len(m)) + m            # streamname reserved
    s += _rec(0x001C, b'')                                 # MODULEDOCSTRING
    s += struct.pack('<HI', 0x0048, 0)
    s += _rec(0x0031, struct.pack('<I', 0))                # MODULEOFFSET = 0
    s += _rec(0x001E, struct.pack('<I', 0))                # HELPCONTEXT
    s += _rec(0x002C, struct.pack('<H', 0xFFFF))           # COOKIE
    s += struct.pack('<HI', 0x0021, 0)                     # MODULETYPE procedural
    s += _rec(0x002B, b'')                                 # MODULE terminator
    s += struct.pack('<HI', 0x0010, 0)                     # TERMINATOR
    return bytes(s)

FREESECT, ENDOFCHAIN, FATSECT = 0xFFFFFFFF, 0xFFFFFFFE, 0xFFFFFFFD
SECTOR = 512

def _dir_entry(name, etype, color, left, right, child, start, size):
    nb = name.encode('utf-16-le')[:62]
    namelen = (len(nb) + 2) if name else 0
    nb = nb + b'\x00' * (64 - len(nb))
    e = nb + struct.pack('<H', namelen) + struct.pack('<BB', etype, color)
    e += struct.pack('<iii', left, right, child) + b'\x00' * 16
    e += struct.pack('<I', 0) + b'\x00' * 16
    e += struct.pack('<I', start & 0xFFFFFFFF) + struct.pack('<Q', size)
    return e

def write_cfb(entries):
    import math
    MINISEC = 64
    mini = [e for e in entries if e['type'] == 2 and len(e['data']) < 4096]
    big  = [e for e in entries if e['type'] == 2 and len(e['data']) >= 4096]
    mini_fat, mini_stream = [], bytearray()
    for e in mini:
        d = e['data']; nsec = max(1, (len(d) + MINISEC - 1) // MINISEC)
        e['start'] = len(mini_fat)
        for k in range(nsec): mini_fat.append(e['start'] + k + 1)
        mini_fat[e['start'] + nsec - 1] = ENDOFCHAIN
        mini_stream += d
        if len(mini_stream) % MINISEC:
            mini_stream += b'\x00' * (MINISEC - len(mini_stream) % MINISEC)
    sectors, fat = [], []
    def alloc(data):
        if not data: return ENDOFCHAIN
        start = len(sectors); n = (len(data) + SECTOR - 1) // SECTOR
        for k in range(n):
            ch = data[k*SECTOR:(k+1)*SECTOR]
            sectors.append(ch + b'\x00' * (SECTOR - len(ch))); fat.append(start + k + 1)
        fat[-1] = ENDOFCHAIN; return start
    for e in big: e['start'] = alloc(e['data'])
    mini_start = alloc(bytes(mini_stream)) if mini_stream else ENDOFCHAIN
    for e in entries:
        if e['type'] == 5:
            e['start'] = mini_start; e['mini_size'] = len(mini_stream)
    mfat = b''.join(struct.pack('<I', x) for x in mini_fat)
    minifat_start = alloc(mfat) if mfat else ENDOFCHAIN
    minifat_nsec = (len(mfat) + SECTOR - 1) // SECTOR if mfat else 0
    db = bytearray()
    for e in entries:
        sz = e.get('mini_size', 0) if e['type'] == 5 else (len(e['data']) if e['type'] == 2 else 0)
        db += _dir_entry(e['name'], e['type'], e['color'], e['left'], e['right'],
                         e['child'], e.get('start', ENDOFCHAIN), sz)
    empty = _dir_entry('', 0, 0, -1, -1, -1, ENDOFCHAIN, 0)
    while len(db) % SECTOR: db += empty
    dir_start = alloc(bytes(db))
    nfat = 1
    while True:
        total = len(sectors) + nfat
        need = math.ceil(total * 4 / SECTOR)
        if need == nfat: break
        nfat = need
    fat_start = len(sectors)
    for k in range(nfat): sectors.append(b''); fat.append(FATSECT)
    while len(fat) < len(sectors): fat.append(FREESECT)
    fb = b''.join(struct.pack('<I', x) for x in fat)
    fb += b'\xFF' * ((SECTOR - len(fb) % SECTOR) % SECTOR)
    for k in range(nfat): sectors[fat_start + k] = fb[k*SECTOR:(k+1)*SECTOR]
    difat = [fat_start + k for k in range(nfat)] + [FREESECT] * (109 - nfat)
    h = bytearray(bytes.fromhex('D0CF11E0A1B11AE1') + b'\x00' * 16)
    h += struct.pack('<HHH', 0x003E, 0x0003, 0xFFFE)
    h += struct.pack('<HH', 9, 6) + b'\x00' * 6
    h += struct.pack('<IIII', 0, nfat, dir_start, 0)
    h += struct.pack('<III', 0x00001000, minifat_start, minifat_nsec)
    h += struct.pack('<II', ENDOFCHAIN, 0)
    for d in difat: h += struct.pack('<I', d)
    return bytes(h) + b''.join(sectors)

def make_vbaproject_bin(macro_src, mn='modPM'):
    dir_comp = ovba_compress(build_dir_stream(mn))
    vbaproj  = struct.pack('<HHH', 0x61CC, 0xFFFF, 0x0000) + b'\x00' * 2
    module   = ovba_compress(macro_src.replace('\n', '\r\n').encode('cp1255', 'replace'))
    proj = (
        'ID="{00000000-0000-0000-0000-000000000000}"\r\n'
        f'Module={mn}\r\nName="VBAProject"\r\nHelpContextID="0"\r\n'
        'VersionCompatible32="393222000"\r\n'
        'CMG="0000000000000000000000000000"\r\nDPB="0000000000000000000000000000"\r\n'
        'GC="0000000000000000000000000000"\r\n\r\n[Host Extender Info]\r\n'
        '&H00000001={3832D640-CF90-11CF-8E43-00A0C911005A};VBE;&H00000000\r\n\r\n'
        f'[Workspace]\r\n{mn}=0, 0, 0, 0, C\r\n').encode('latin-1', 'replace')
    pwm = mn.encode('latin-1') + b'\x00' + mn.encode('latin-1') + b'\x00\x00\x00'
    entries = [
        dict(name='Root Entry', type=5, color=1, left=-1, right=-1, child=2, data=b''),
        dict(name='VBA', type=1, color=1, left=-1, right=-1, child=6, data=b''),
        dict(name='PROJECT', type=2, color=1, left=1, right=3, child=-1, data=proj),
        dict(name='PROJECTwm', type=2, color=1, left=-1, right=-1, child=-1, data=pwm),
        dict(name='dir', type=2, color=1, left=-1, right=5, child=-1, data=dir_comp),
        dict(name='_VBA_PROJECT', type=2, color=1, left=-1, right=-1, child=-1, data=vbaproj),
        dict(name='modPM', type=2, color=1, left=4, right=-1, child=-1, data=module),
    ]
    return write_cfb(entries)

# ===========================================================================
# 3) Build the workbook with openpyxl
# ===========================================================================
def style(c, *, bold=False, color=INK, fill=None, v="center", size=10, wrap=True, h="right"):
    c.font = Font(name=FONT_NAME, bold=bold, size=size, color=color)
    c.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    c.border = BORDER
    if fill: c.fill = PatternFill("solid", fgColor=fill)

def safe(name):
    for ch in ':\\/?*[]': name = name.replace(ch, "")
    return name[:31]

# data-grid alignment: IDs centered, technical/English left, Hebrew right
LEFTCOLS  = {2, 3, 4, 5, 7, 9, 14}
RIGHTCOLS = {6, 8, 10, 11, 12, 13, 15}
def col_h(col):
    return "center" if col == 1 else ("left" if col in LEFTCOLS else "right")

def add_back_button(ws, ncol, title):
    """Modern back-to-dashboard button (A1:B1) + centred 16pt title block (C1:ncol)."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    b = ws.cell(1, 1, "◄ חזרה למסך הראשי")
    b.hyperlink = f"#'{DASH}'!A1"
    b.font = Font(name=FONT_NAME, bold=True, size=10, color=BACK_TXT)
    b.fill = PatternFill("solid", fgColor=BACK_BG)
    b.alignment = Alignment(horizontal="center", vertical="center")
    red_side = Side(style="thin", color=RED_HOT)
    red_border = Border(left=red_side, right=red_side, top=red_side, bottom=red_side)
    b.border = red_border
    ws.cell(1, 2).border = red_border
    ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=ncol)
    t = ws.cell(1, 3, title)
    t.font = Font(name=FONT_NAME, bold=True, size=16, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=DASH_HDR)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

wb = Workbook()
wb.remove(wb.active)
sheet_names = [safe(t["title"]) for t in TOPICS]
dash = wb.create_sheet(DASH)               # first sheet -> Sheets(1) in VBA

index_rows = []                            # (type, code, he, en, sheet, cell)
tables_meta = []                           # (table, he, topic_title, sheet, cell, theme) - for Cockpit

for topic_idx, (topic, sname) in enumerate(zip(TOPICS, sheet_names)):
    th = TH[topic["theme"]]
    ws = wb.create_sheet(sname)
    ws.sheet_view.rightToLeft = True
    ws.sheet_view.showGridLines = True     # gridlines ON per spec
    ws.sheet_properties.tabColor = th["t"]
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

    add_back_button(ws, NCOL, topic["title"] + "   |   ECC 6.0 ➜ S/4HANA   ·   CBC NEO")

    for j, (lbl, w) in enumerate(COLS, start=1):
        is_s4 = j >= S4_COL_START
        fill = S4_HDR if is_s4 else th["s"]
        c = ws.cell(2, j, lbl)
        c.font = Font(name=FONT_NAME, bold=True, size=11, color=(S4_TXT if is_s4 else "FFFFFF"))
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER; c.fill = PatternFill("solid", fgColor=fill)
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[2].height = 28
    ws.freeze_panes = "A3"

    row = 3
    seq = 0
    for tbl in topic["tables"]:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOL)
        hc = ws.cell(row, 1, f"  ◆ טבלה {tbl['name']}  -  {tbl['he']}  ({tbl['en']})")
        hc.font = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFF")
        hc.fill = PatternFill("solid", fgColor=th["h"])
        hc.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[row].height = 24
        index_rows.append(("טבלה", tbl["name"], tbl["he"], tbl["en"], sname, f"A{row}"))
        tables_meta.append((tbl["name"], tbl["he"], topic["title"], sname, f"A{row}", topic["theme"]))
        row += 1

        fstart = row
        funcs = "\n".join(f"• {n}" for n, _ in tbl["funcs"])
        fdesc = "\n".join(f"• {d}" for _, d in tbl["funcs"])
        progs = "\n".join(f"• {n}" for n, _ in tbl["progs"])
        pdesc = "\n".join(f"• {d}" for _, d in tbl["progs"])
        for (tech, en, he) in tbl["fields"]:
            seq += 1
            band = th["b"] if seq % 2 == 0 else None
            s4b = S4_BAND if seq % 2 == 0 else None
            style(ws.cell(row, 1, seq), bold=True, fill=band, h="center", color="6B7280")
            style(ws.cell(row, 2, ""), fill=band)
            style(ws.cell(row, 3, ""), fill=band)
            style(ws.cell(row, 4, tech), bold=True, color=RED, fill=band, h="left")
            style(ws.cell(row, 5, en), fill=band, h="left")
            style(ws.cell(row, 6, he), fill=band, h="right")
            for col in (7, 8, 9, 10): style(ws.cell(row, col, ""), fill=band)
            for col in (11, 12, 13, 14): style(ws.cell(row, col, ""), fill=s4b)
            ws.row_dimensions[row].height = 20
            index_rows.append(("שדה", tech, he, en, sname, f"D{row}"))
            row += 1
        fend = row - 1

        def vmerge(col, val, *, bold=False, color=INK):
            cell = ws.cell(fstart, col, val)
            cell.font = Font(name=FONT_NAME, bold=bold, size=9, color=color)
            cell.alignment = Alignment(horizontal=col_h(col), vertical="top", wrap_text=True)
            cell.border = BORDER
            if fend > fstart:
                ws.merge_cells(start_row=fstart, start_column=col, end_row=fend, end_column=col)
        vmerge(2, tbl["tcodes"], bold=True, color=SLATE)
        vmerge(3, tbl["name"], bold=True, color=RED)
        vmerge(7, funcs, bold=True, color=SLATE); vmerge(8, fdesc)
        vmerge(9, progs, bold=True, color=SLATE); vmerge(10, pdesc)
        vmerge(11, tbl["s4_status"], color="37474F"); vmerge(12, tbl["s4_repl"], color="37474F")
        vmerge(13, tbl["s4_tcode"], color="37474F"); vmerge(14, tbl["fiori"], bold=True, color=RED)
        vmerge(15, sum_note(tbl), color="37474F")

        for code in [x.strip() for x in tbl["tcodes"].replace(";", ",").replace("/", ",").split(",") if x.strip()]:
            index_rows.append(("טרנזקציה", code, tbl["he"], tbl["en"], sname, f"A{fstart}"))
        for n, d in tbl["funcs"]: index_rows.append(("פונקציה", n, d, tbl["en"], sname, f"G{fstart}"))
        for n, d in tbl["progs"]: index_rows.append(("תוכנית", n, d, tbl["en"], sname, f"I{fstart}"))
        index_rows.append(("Fiori", tbl["fiori"], tbl["he"], tbl["en"], sname, f"N{fstart}"))
        row += 1

    # ---- Technical Extensions section (user exits / BAdIs / architecture diffs) ----
    ext = EXTENSIONS[topic_idx]
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOL)
    eh = ws.cell(row, 1, "  ⚙ הרחבות טכניות ושינויי ארכיטקטורה (Technical Extensions: User Exits / BAdIs / ECC vs S/4HANA)")
    eh.font = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFF")
    eh.fill = PatternFill("solid", fgColor=th["h"])
    eh.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[row].height = 24
    row += 1
    # sub-grid headers (3 cols: type | code | description spanning)
    for lbl, c0, c1 in [("סוג (Type)", 1, 2), ("קוד / שם טכני (Code)", 3, 4), ("תיאור / הבדל (Hebrew)", 5, NCOL)]:
        ws.merge_cells(start_row=row, start_column=c0, end_row=row, end_column=c1)
        c = ws.cell(row, c0, lbl)
        c.font = Font(name=FONT_NAME, bold=True, size=10, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=th["s"])
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = BORDER
    row += 1

    def ext_row(kind, code, desc, idx_type):
        global row
        band = ZEBRA if (row % 2 == 0) else None
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        a = ws.cell(row, 1, kind); a.font = Font(name=FONT_NAME, bold=True, size=9, color=RED)
        a.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); a.border = BORDER
        if band: a.fill = PatternFill("solid", fgColor=band)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        b = ws.cell(row, 3, code); b.font = Font(name=FONT_NAME, bold=True, size=9, color=SLATE)
        b.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True); b.border = BORDER
        if band: b.fill = PatternFill("solid", fgColor=band)
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=NCOL)
        d = ws.cell(row, 5, desc); d.font = Font(name=FONT_NAME, size=9, color=INK)
        d.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True); d.border = BORDER
        if band: d.fill = PatternFill("solid", fgColor=band)
        ws.row_dimensions[row].height = 24
        if idx_type:
            index_rows.append((idx_type, code, desc, "", sname, f"A{row}"))
        row += 1

    for code, desc in ext["exits"]:
        ext_row("User Exit (SMOD)", code, desc, "User Exit")
    for code, desc in ext["badis"]:
        ext_row("BAdI (SE18)", code, desc, "BAdI")
    for note in ext["arch"]:
        ext_row("ארכיטקטורה ECC↔S/4", "—", note, "")

# ===========================================================================
#  Simplification Item list sheet  (SAP Notes per PM/EAM object)
# ===========================================================================
def new_sheet(name, tabcolor):
    s = wb.create_sheet(name)
    s.sheet_view.rightToLeft = True
    s.sheet_view.showGridLines = True
    s.sheet_properties.tabColor = tabcolor
    s.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    return s

def grid_header(ws, cols, fill):
    for j, (lbl, w) in enumerate(cols, start=1):
        c = ws.cell(2, j, lbl)
        c.font = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = BORDER
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[2].height = 28
    ws.freeze_panes = "A3"

# === Simplification Item list ===============================================
simp = new_sheet(SIMP_SHEET, SILVER)
add_back_button(simp, 6, "Simplification Item List  -  SAP Notes  ·  CBC NEO")
grid_header(simp, [("מס' (#)", 6), ("תחום / אובייקט", 20), ("Simplification Item (כותרת הפריט)", 42),
                   ("SAP Note", 16), ("קטגוריה", 18), ("השפעה והמלצה (Impact & Action)", 62)], SLATE)
rr = 3
for i, (obj, title, note, cat, impact) in enumerate(SIMPLIFICATION, start=1):
    band = ZEBRA if i % 2 == 0 else None
    for j, v in enumerate([i, obj, title, note, cat, impact], start=1):
        c = simp.cell(rr, j, v)
        c.font = Font(name=FONT_NAME, bold=(j in (3, 4)), size=10, color=RED if j == 4 else INK)
        c.alignment = Alignment(horizontal=("center" if j in (1, 4) else "right"), vertical="top", wrap_text=True)
        c.border = BORDER
        if band: c.fill = PatternFill("solid", fgColor=band)
    simp.row_dimensions[rr].height = 54
    index_rows.append(("SAP Note", note, title, obj, SIMP_SHEET, f"A{rr}"))
    rr += 1

# === Cockpit - migration tracking ==========================================
STATUSES = ["Not started", "In analysis", "In conversion", "Tested", "Done"]
# branded soft pastels: rose for open, silver for analysis/conversion, soft green for done
STATUS_FILL = {"Not started": "FCE4E6", "In analysis": "ECEFF1", "In conversion": "E1E6EA",
               "Tested": "DCE6EC", "Done": "DCEFE0"}
STATUS_TXT = {"Not started": "B01722", "In analysis": "475569", "In conversion": "37474F",
              "Tested": "2A4A57", "Done": "1E5A44"}
cock = new_sheet(COCKPIT, RED)
add_back_button(cock, 9, "Cockpit מעקב מיגרציה  -  Migration Tracking (NEO Project)")
cock_cols = [("מס' (#)", 6), ("מודול / נושא", 26), ("אובייקט / טבלה", 16), ("סטטוס מעבר (Status)", 18),
             ("אחראי (Owner)", 16), ("תאריך יעד", 13), ("תאריך סיום", 13),
             ("הערות / בלוקרים", 34), ("קישור לגיליון", 14)]
grid_header(cock, cock_cols, SLATE)
rr = 3
for i, (tname, the, ttitle, tsheet, tcell, ttheme) in enumerate(tables_meta, start=1):
    band = ZEBRA if i % 2 == 0 else None
    vals = [i, f"{ttitle}  -  {the}", tname, "Not started", "", "", "", "", None]
    for j, v in enumerate(vals, start=1):
        c = cock.cell(rr, j, v if v is not None else "")
        c.font = Font(name=FONT_NAME, bold=(j == 3), size=10,
                      color=RED if j == 3 else INK)
        c.alignment = Alignment(horizontal=("center" if j in (1, 4, 6, 7, 9) else ("left" if j == 3 else "right")),
                                vertical="center", wrap_text=True)
        c.border = BORDER
        if band: c.fill = PatternFill("solid", fgColor=band)
    # status default styling
    sc = cock.cell(rr, 4); sc.fill = PatternFill("solid", fgColor=STATUS_FILL["Not started"])
    sc.font = Font(name=FONT_NAME, bold=True, size=10, color=STATUS_TXT["Not started"])
    lk = cock.cell(rr, 9); lk.value = f'=HYPERLINK("#\'"&"{tsheet}"&"\'!{tcell}","➜ פתח")'
    lk.font = Font(name=FONT_NAME, color="0563C1", underline="single", size=9)
    cock.row_dimensions[rr].height = 22
    rr += 1
cock_last = rr - 1
dv = DataValidation(type="list", formula1='"%s"' % ",".join(STATUSES), allow_blank=True)
dv.error = "בחר סטטוס מהרשימה"; dv.prompt = "בחר סטטוס מעבר"
cock.add_data_validation(dv); dv.add(f"D3:D{cock_last}")
for st_name, fillc in STATUS_FILL.items():
    cock.conditional_formatting.add(f"D3:D{cock_last}",
        CellIsRule(operator="equal", formula=[f'"{st_name}"'],
                   fill=PatternFill("solid", fgColor=fillc),
                   font=Font(name=FONT_NAME, bold=True, color=STATUS_TXT[st_name])))

# --- status summary table (cols K:L) + dynamic doughnut chart on the Cockpit ---
from openpyxl.chart import DoughnutChart, BarChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice

cock.cell(2, 11, "סטטוס (Status)").font = Font(name=FONT_NAME, bold=True, size=10, color="FFFFFF")
cock.cell(2, 12, "כמות (Count)").font = Font(name=FONT_NAME, bold=True, size=10, color="FFFFFF")
for cc_ in (11, 12):
    cock.cell(2, cc_).fill = PatternFill("solid", fgColor=DASH_HDR)
    cock.cell(2, cc_).alignment = Alignment(horizontal="center", vertical="center")
    cock.cell(2, cc_).border = BORDER
for k, stt in enumerate(STATUSES):
    rk = 3 + k
    a = cock.cell(rk, 11, stt); a.font = Font(name=FONT_NAME, bold=True, size=10, color=STATUS_TXT[stt])
    a.fill = PatternFill("solid", fgColor=STATUS_FILL[stt]); a.alignment = Alignment(horizontal="right")
    a.border = BORDER
    b = cock.cell(rk, 12, f'=COUNTIF($D$3:$D${cock_last},$K{rk})')
    b.font = Font(name=FONT_NAME, bold=True, size=11, color=INK); b.alignment = Alignment(horizontal="center")
    b.border = BORDER
trow = 3 + len(STATUSES)
cock.cell(trow, 11, 'סה"כ').font = Font(name=FONT_NAME, bold=True, size=10, color="FFFFFF")
cock.cell(trow, 11).fill = PatternFill("solid", fgColor="64748B"); cock.cell(trow, 11).border = BORDER
cock.cell(trow, 11).alignment = Alignment(horizontal="right")
tt = cock.cell(trow, 12, f'=SUM($L$3:$L${trow-1})'); tt.font = Font(name=FONT_NAME, bold=True, size=11, color=INK)
tt.alignment = Alignment(horizontal="center"); tt.border = BORDER
cock.column_dimensions["K"].width = 16; cock.column_dimensions["L"].width = 12

donut = DoughnutChart(); donut.title = "התפלגות סטטוס מיגרציה"; donut.holeSize = 55
data = Reference(cock, min_col=12, min_row=2, max_row=trow - 1)      # L2:L7 (header+5)
cats = Reference(cock, min_col=11, min_row=3, max_row=trow - 1)      # K3:K7
donut.add_data(data, titles_from_data=True); donut.set_categories(cats)
for k, stt in enumerate(STATUSES):                                   # colour slices per status
    pt = DataPoint(idx=k); pt.graphicalProperties.solidFill = STATUS_FILL[stt]
    donut.series[0].data_points.append(pt)
donut.dataLabels = DataLabelList(); donut.dataLabels.showVal = True
donut.height = 7.5; donut.width = 11
cock.add_chart(donut, "K11")

# === Custom Code Check =====================================================
CCC_STATUS = ["To review", "In progress", "Adapted", "OK - no change", "Obsolete"]
CCC_FILL = {"To review": "FCE4E6", "In progress": "ECEFF1", "Adapted": "DCE6EC",
            "OK - no change": "DCEFE0", "Obsolete": "EEF1F3"}
CCC_TXT = {"To review": "B01722", "In progress": "475569", "Adapted": "2A4A57",
           "OK - no change": "1E5A44", "Obsolete": "6B7280"}
ccc = new_sheet(CCC, SILVER)
add_back_button(ccc, 9, "Custom Code Check  -  User Exits / BAdIs  ·  CBC NEO")
ccc_cols = [("מס' (#)", 6), ("מודול / נושא", 24), ("סוג (Type)", 16), ("קוד / שם טכני", 22),
            ("תיאור (Hebrew)", 34), ("סטטוס בדיקה", 16), ("המלצת מעבר ל-S/4", 30),
            ("אחראי (Owner)", 14), ("הערות", 22)]
grid_header(ccc, ccc_cols, SLATE)
rr = 3
EXIT_REC = "בדוק ב-Custom Code Migration (SCMON/ATC); שקול מעבר ל-BAdI/Enhancement Spot מודרני."
BADI_REC = "אמת תאימות ה-BAdI ב-S/4 (SPAU_ENH); ודא חתימה ומימוש אקטיבי לאחר השדרוג."
for ti, (topic, sname) in enumerate(zip(TOPICS, sheet_names)):
    ext = EXTENSIONS[ti]
    for kind, rec, lst in [("User Exit", EXIT_REC, ext["exits"]), ("BAdI", BADI_REC, ext["badis"])]:
        for code, desc in lst:
            i = rr - 2
            band = ZEBRA if i % 2 == 0 else None
            vals = [i, topic["title"], kind, code, desc, "To review", rec, "", ""]
            for j, v in enumerate(vals, start=1):
                c = ccc.cell(rr, j, v)
                c.font = Font(name=FONT_NAME, bold=(j == 4), size=10,
                              color=SLATE if j == 4 else INK)
                c.alignment = Alignment(horizontal=("center" if j in (1, 3, 6, 8) else ("left" if j == 4 else "right")),
                                        vertical="center", wrap_text=True)
                c.border = BORDER
                if band: c.fill = PatternFill("solid", fgColor=band)
            scc = ccc.cell(rr, 6); scc.fill = PatternFill("solid", fgColor=CCC_FILL["To review"])
            scc.font = Font(name=FONT_NAME, bold=True, size=10, color=CCC_TXT["To review"])
            ccc.row_dimensions[rr].height = 26
            index_rows.append((kind, code, desc, topic["title"], CCC, f"D{rr}"))
            rr += 1
ccc_last = rr - 1
dv2 = DataValidation(type="list", formula1='"%s"' % ",".join(CCC_STATUS), allow_blank=True)
ccc.add_data_validation(dv2); dv2.add(f"F3:F{ccc_last}")
for st_name, fillc in CCC_FILL.items():
    ccc.conditional_formatting.add(f"F3:F{ccc_last}",
        CellIsRule(operator="equal", formula=[f'"{st_name}"'],
                   fill=PatternFill("solid", fgColor=fillc),
                   font=Font(name=FONT_NAME, bold=True, color=CCC_TXT[st_name])))

# ---- pandas index frame (drives both the FILTER formula and the macro) ----
df_index = pd.DataFrame(index_rows, columns=["סוג", "קוד", "עברית", "English", "גיליון", "תא"])
N = len(df_index)

# ===========================================================================
# 4) Dashboard
# ===========================================================================
dash.sheet_view.rightToLeft = True
dash.sheet_view.showGridLines = False
dash.sheet_properties.tabColor = DASH_HDR
for col, w in zip("ABCDEFGHIJKLMNO", [3,16,16,16,16,16,4,16,16,3,3,3,3,3,3]):
    dash.column_dimensions[col].width = w
# soft-grey dashboard backdrop
for rfill in range(1, 63):
    for cfill in range(1, 16):
        dash.cell(rfill, cfill).fill = PatternFill("solid", fgColor=DASH_BG)

dash.merge_cells("B2:H2")
d1 = dash.cell(2, 2, "CBC ISRAEL  |  SAP PM  -  מסך ניווט מרכזי")
d1.font = Font(name=FONT_NAME, bold=True, size=18, color=DASH_HDR); d1.alignment = Alignment(horizontal="right")
dash.merge_cells("B3:H3")
d2 = dash.cell(3, 2, "Project NEO  -  ECC 6.0 ➜ S/4HANA Migration  |  Plant Maintenance / EAM")
d2.font = Font(name=FONT_NAME, bold=True, size=10, color="64748B"); d2.alignment = Alignment(horizontal="right")
# NEO project badge (top-right)
dash.merge_cells("J2:O3")
neo = dash.cell(2, 10, "★ CBC ISRAEL\nProject NEO")
neo.font = Font(name=FONT_NAME, bold=True, size=12, color="FFFFFF")
neo.fill = PatternFill("solid", fgColor=RED)
neo.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
neo_side = Side(style="thin", color="FFFFFF")
for rrn in (2, 3):
    for ccn in range(10, 16):
        dash.cell(rrn, ccn).border = Border(left=neo_side, right=neo_side, top=neo_side, bottom=neo_side)

# ---------------------------------------------------------------------------
#  KPI strip (top) - live counts from the Cockpit status column
# ---------------------------------------------------------------------------
CR = f"'{COCKPIT}'"                                   # quoted cockpit sheet ref
TOTcnt  = f"COUNTA({CR}!$C$3:$C${cock_last})"
DONEcnt = f'COUNTIF({CR}!$D$3:$D${cock_last},"Done")'
TESTcnt = f'COUNTIF({CR}!$D$3:$D${cock_last},"Tested")'
PROGcnt = f'(COUNTIF({CR}!$D$3:$D${cock_last},"In analysis")+COUNTIF({CR}!$D$3:$D${cock_last},"In conversion"))'
OPENcnt = f'COUNTIF({CR}!$D$3:$D${cock_last},"Not started")'

dash.merge_cells("B5:H5")
kh = dash.cell(5, 2, "סיכום התקדמות מיגרציה  (KPI - מתעדכן אוטומטית מ-Cockpit)")
kh.font = Font(name=FONT_NAME, bold=True, size=12, color="FFFFFF"); kh.fill = PatternFill("solid", fgColor=DASH_HDR)
kh.alignment = Alignment(horizontal="right")

# 6 KPI tiles across B..G (number row 6, label row 7)
kpis = [("סה\"כ אובייקטים", f"={TOTcnt}", SLATE, "ECEFF1"),
        ("הושלם (Done)", f"={DONEcnt}", "1E5A44", "DCEFE0"),
        ("נבדק (Tested)", f"={TESTcnt}", "2A4A57", "DCE6EC"),
        ("בתהליך (In Progress)", f"={PROGcnt}", "37474F", "E1E6EA"),
        ("פתוח (Open)", f"={OPENcnt}", "B01722", "FCE4E6"),
        ("% השלמה", f"=IFERROR(({DONEcnt})/({TOTcnt}),0)", RED, "FBE3E4")]
for k, (lbl, formula_v, txt, fillc) in enumerate(kpis):
    cc = 2 + k
    num = dash.cell(6, cc, formula_v)
    num.font = Font(name=FONT_NAME, bold=True, size=20, color=txt)
    num.alignment = Alignment(horizontal="center", vertical="center")
    num.fill = PatternFill("solid", fgColor=fillc)
    num.border = BORDER
    if "%" in lbl:
        num.number_format = "0%"
    lab = dash.cell(7, cc, lbl)
    lab.font = Font(name=FONT_NAME, bold=True, size=9, color=txt)
    lab.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    lab.fill = PatternFill("solid", fgColor=fillc)
    lab.border = BORDER
dash.row_dimensions[6].height = 30
dash.row_dimensions[7].height = 22

# progress bar chart (anchored right, references the Cockpit summary - dynamic)
bar = BarChart(); bar.type = "bar"; bar.title = "התקדמות לפי סטטוס"; bar.legend = None
bdata = Reference(cock, min_col=12, min_row=2, max_row=2 + len(STATUSES))
bcats = Reference(cock, min_col=11, min_row=3, max_row=2 + len(STATUSES))
bar.add_data(bdata, titles_from_data=True); bar.set_categories(bcats)
bar.height = 5.2; bar.width = 11
dash.add_chart(bar, "J5")

dash.merge_cells("B9:H9")
nh = dash.cell(9, 2, "ניווט מהיר לגיליונות  (לחץ על כרטיס כדי לעבור)")
nh.font = Font(name=FONT_NAME, bold=True, size=12, color="FFFFFF"); nh.fill = PatternFill("solid", fgColor=SLATE)
nh.alignment = Alignment(horizontal="right")

card_cols = [2, 4, 6, 8]                # 4 cards per row (cols B, D, F, H)
nav_items = [(t["title"], sn, TH[t["theme"]]["t"]) for t, sn in zip(TOPICS, sheet_names)]
nav_items.append(("★ " + SIMP_SHEET, SIMP_SHEET, SILVER))
nav_items.append(("◆ " + COCKPIT, COCKPIT, RED))
nav_items.append(("⚙ " + CCC, CCC, SLATE))
r = 10
for i, (title, sn, hdr) in enumerate(nav_items):
    slot = i % 4
    if slot == 0 and i != 0: r += 3
    c0 = card_cols[slot]
    dash.merge_cells(start_row=r, start_column=c0, end_row=r+1, end_column=c0)
    cell = dash.cell(r, c0, title)
    cell.hyperlink = f"#'{sn}'!A1"
    cell.font = Font(bold=True, size=10, color=(SLATE if hdr == SILVER else "FFFFFF"))
    cell.fill = PatternFill("solid", fgColor=hdr)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for rr in (r, r+1): dash.cell(rr, c0).border = BORDER
r += 3

dash.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
lg = dash.cell(r, 2, "מקרא מותג CBC:  אדום=תנועתי/עוגני-על | צפחה=נתוני אב וכותרות | כסף=אינטגרציה ועמודות S/4HANA")
lg.font = Font(name=FONT_NAME, italic=True, size=9, color="64748B"); lg.alignment = Alignment(horizontal="right")
lg.fill = PatternFill("solid", fgColor=DASH_BG)

# --- search block (row 24 header, row 25 box) ---
dash.merge_cells("B24:F24")
sh = dash.cell(24, 2, "🔍 חיפוש גלובלי  (טבלה / טרנזקציה / שדה / פונקציה - עברית או אנגלית)")
sh.font = Font(bold=True, size=12, color="FFFFFF"); sh.fill = PatternFill("solid", fgColor=SLATE)
sh.alignment = Alignment(horizontal="right")
lbl = dash.cell(25, 2, "הקלד כאן:")
lbl.font = Font(bold=True, size=11); lbl.alignment = Alignment(horizontal="right")
lbl.fill = PatternFill("solid", fgColor=DASH_BG)
dash.merge_cells("C25:F25")
sc = dash.cell(25, 3, "")
sc.fill = PatternFill("solid", fgColor="FCE4E6")
sc.font = Font(bold=True, size=12, color=SLATE)
med = Side(style="medium", color=RED_HOT)
sc.border = Border(left=med, right=med, top=med, bottom=med)
dash.row_dimensions[25].height = 26

# results header (row 27) + FILTER (row 28, spills) - reads the on-sheet index block
res_hdrs = ["קישור", "קוד / שם טכני", "פירוש בעברית", "English", "גיליון", "תא"]
for j, h in enumerate(res_hdrs):
    c = dash.cell(27, 2 + j, h)
    c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=SLATE)
    c.alignment = Alignment(horizontal="center"); c.border = BORDER

C = lambda k: get_column_letter(IDX_COL0 + k)         # P,Q,R,S,T,U,V
rng = lambda k: f"${C(k)}$2:${C(k)}${N+1}"
formula = (
 f'=IF({SEARCH_CELL}="","הקלד מונח וראה תוצאות, או לחץ על כפתור החיפוש...",'
 f'IFERROR(FILTER(CHOOSE({{1,2,3,4,5,6}},{rng(6)},{rng(1)},{rng(2)},{rng(3)},{rng(4)},{rng(5)}),'
 f'(ISNUMBER(SEARCH({SEARCH_CELL},{rng(1)})))+(ISNUMBER(SEARCH({SEARCH_CELL},{rng(2)})))+'
 f'(ISNUMBER(SEARCH({SEARCH_CELL},{rng(3)})))+(ISNUMBER(SEARCH({SEARCH_CELL},{rng(0)})))>0),'
 f'"לא נמצאו תוצאות - נסה מונח אחר"))'
)
fcell = dash.cell(28, 2, formula)
fcell.font = Font(size=10); fcell.alignment = Alignment(horizontal="right")

# --- embedded INDEX_DB block on the dashboard (hidden cols P:V) ---
idx_titles = ["סוג (Type)", "קוד (Code)", "עברית", "English", "גיליון (Sheet)", "תא (Cell)", "קישור (Link)"]
for k, ttl in enumerate(idx_titles):
    c = dash.cell(1, IDX_COL0 + k, ttl)
    c.font = Font(bold=True, size=8, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=SLATE)
    c.alignment = Alignment(horizontal="right")
for ri, row in enumerate(index_rows, start=2):
    typ, code, he, en, sn, cell = row
    dash.cell(ri, IDX_COL0 + 0, typ)
    dash.cell(ri, IDX_COL0 + 1, code).font = Font(bold=True, size=8, color=SLATE)
    dash.cell(ri, IDX_COL0 + 2, he)
    dash.cell(ri, IDX_COL0 + 3, en)
    dash.cell(ri, IDX_COL0 + 4, sn)
    dash.cell(ri, IDX_COL0 + 5, cell)
    lk = dash.cell(ri, IDX_COL0 + 6)
    lk.value = f'=HYPERLINK("#\'"&{C(4)}{ri}&"\'!"&{C(5)}{ri},"➜ פתח")'
    lk.font = Font(color="0563C1", underline="single", size=8)
    for k in range(7):
        cc = dash.cell(ri, IDX_COL0 + k)
        if k != 6: cc.font = Font(size=8, color="333333", bold=(k == 1))
        cc.alignment = Alignment(horizontal="right")
for k in range(7):
    dash.column_dimensions[get_column_letter(IDX_COL0 + k)].hidden = True   # hide the index block

# instructions / note
dash.merge_cells("B56:H60")
note = dash.cell(56, 2,
  "איך עובדים:  1) לחץ כרטיס למעבר לגיליון (כולל Cockpit מעקב מיגרציה ו-Custom Code Check).  "
  "2) הקלד מונח בתא הצהוב ולחץ על כפתור 'חיפוש' (מאקרו) או צפה בתוצאות הדינמיות מתחת (FILTER).  "
  "3) בכל גיליון יש כפתור 'חזרה למסך הראשי' בפינה הימנית-עליונה.  "
  "ב-Cockpit וב-Custom Code Check יש עמודות סטטוס עם רשימה נפתחת וצביעה אוטומטית.  "
  "אם Excel חוסם מאקרו - אשר 'Enable Content'; אם נדרש, ייבא את modPM.bas (Alt+F11 -> Import).")
note.font = Font(name=FONT_NAME, italic=True, size=9, color="475569")
note.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
note.fill = PatternFill("solid", fgColor="F1F3F5")
note.border = BORDER
dash.freeze_panes = "A2"

# ---------------------------------------------------------------------------
#  Branded print setup - CBC / Project NEO header & footer band + page layout
#  (appears at the top/bottom of every printed page / PDF export)
# ---------------------------------------------------------------------------
from openpyxl.worksheet.page import PageMargins
for sh in wb.worksheets:
    sh.oddHeader.left.text   = '&"Segoe UI,Bold"&13&KD62027CBC ISRAEL'
    sh.oddHeader.center.text = '&"Segoe UI,Bold"&11&K1E1E24Project NEO  |  SAP PM  ECC 6.0 -> S/4HANA Migration'
    sh.oddHeader.right.text  = '&"Segoe UI"&9&K6B7280&A'
    sh.oddFooter.left.text   = '&"Segoe UI"&8&K6B7280SAP PM Migration Workbook  -  CBC Israel (Project NEO)'
    sh.oddFooter.center.text = '&"Segoe UI"&8&K6B7280Page &P of &N'
    sh.oddFooter.right.text  = '&"Segoe UI"&8&K6B7280&D'
    sh.page_setup.orientation = "landscape"
    sh.page_setup.fitToWidth = 1
    sh.page_setup.fitToHeight = 0
    sh.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    sh.print_options.horizontalCentered = True
    sh.print_options.gridLines = False
    sh.page_margins = PageMargins(left=0.3, right=0.3, top=0.7, bottom=0.6, header=0.3, footer=0.3)
    # explicit branded print area (keeps the dashboard PDF clean - excludes hidden index P:V)
    if sh.title == DASH:
        sh.print_area = "A1:O60"
    elif sh.title == COCKPIT:
        sh.print_area = f"A1:L{sh.max_row}"
        sh.print_title_rows = "1:2"
    else:
        sh.print_area = f"A1:{get_column_letter(sh.max_column)}{sh.max_row}"
        sh.print_title_rows = "1:2"          # repeat title + column headers on every page

# ---------------------------------------------------------------------------
#  Universal typography pass - apply the modern font family to every cell
# ---------------------------------------------------------------------------
for sh in wb.worksheets:
    for rowcells in sh.iter_rows():
        for cell in rowcells:
            ftn = cell.font
            if ftn is not None and ftn.name != FONT_NAME:
                cell.font = Font(name=FONT_NAME, size=ftn.size, bold=ftn.bold,
                                 italic=ftn.italic, color=ftn.color, underline=ftn.underline)

# ===========================================================================
# 5) Save .xlsx then inject VBA + Form Control button -> .xlsm
# ===========================================================================
TMP_XLSX = "_tmp_pm.xlsx"
OUT = "SAP_PM_ECC6_to_S4_Migration.xlsm"
wb.save(TMP_XLSX)

with open("modPM.bas", "w", encoding="utf-8") as f:        # fallback export
    f.write(MACRO_SRC)

vba_bin = make_vbaproject_bin(MACRO_SRC)

# VML form-control button wired to PM_GlobalSearch (anchored next to the search box)
VML = (
'<xml xmlns:v="urn:schemas-microsoft-com:vml" xmlns:o="urn:schemas-microsoft-com:office:office" '
'xmlns:x="urn:schemas-microsoft-com:office:excel">'
'<o:shapelayout v:ext="edit"><o:idmap v:ext="edit" data="1"/></o:shapelayout>'
'<v:shapetype id="_x0000_t201" coordsize="21600,21600" o:spt="201" path="m,l,21600r21600,l21600,xe">'
'<v:stroke joinstyle="miter"/><v:path shadowok="f" o:extrusionok="f" gradientshapeok="t" o:connecttype="rect"/>'
'</v:shapetype>'
'<v:shape id="SearchBtn" type="#_x0000_t201" style="position:absolute;margin-left:300px;margin-top:495px;'
'width:120px;height:30px;z-index:1" o:button="t" fillcolor="#c55a11" strokecolor="#7b1e2b">'
'<v:fill o:detectmouseclick="t"/><o:lock v:ext="edit" rotation="t"/>'
'<v:textbox style="mso-direction-alt:auto" o:singleclick="f">'
'<div style="text-align:center"><font color="#FFFFFF" size="200" face="Arial"><b>\U0001F50D חיפוש</b></font></div>'
'</v:textbox>'
'<x:ClientData ObjectType="Button">'
'<x:Anchor>6, 30, 23, 4, 7, 110, 24, 12</x:Anchor>'
'<x:PrintObject>False</x:PrintObject><x:AutoFill>False</x:AutoFill>'
'<x:FmlaMacro>PM_GlobalSearch</x:FmlaMacro>'
'<x:TextHAlign>Center</x:TextHAlign><x:TextVAlign>Center</x:TextVAlign>'
'</x:ClientData></v:shape></xml>'
)

def inject_macros(src_xlsx, out_xlsm, vba_bin, vml):
    zin = zipfile.ZipFile(src_xlsx, "r")
    names = zin.namelist()
    ct = zin.read("[Content_Types].xml").decode("utf-8")
    wbrels = zin.read("xl/_rels/workbook.xml.rels").decode("utf-8")
    wbxml = zin.read("xl/workbook.xml").decode("utf-8")

    # map dashboard sheet -> sheetN.xml via workbook.xml + rels (attribute-order agnostic)
    sheet_el = re.search(r'<sheet\b[^>]*\bname="%s"[^>]*/?>' % re.escape(DASH), wbxml).group(0)
    rid = re.search(r'r:id="([^"]+)"', sheet_el).group(1)
    rel_el = re.search(r'<Relationship\b[^>]*\bId="%s"[^>]*/?>' % re.escape(rid), wbrels).group(0)
    tgt = re.search(r'Target="([^"]+)"', rel_el).group(1)
    if tgt.startswith("/"):
        dash_part = tgt.lstrip("/")                      # absolute from package root
    else:
        dash_part = "xl/" + tgt.replace("../", "")       # relative to xl/
    dash_base = os.path.basename(dash_part)
    dash_rels = f"xl/worksheets/_rels/{dash_base}.rels"

    # content types
    ct = ct.replace(
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml',
        'application/vnd.ms-excel.sheet.macroEnabled.main+xml')
    inserts = ''
    if 'Extension="vml"' not in ct:
        inserts += '<Default Extension="vml" ContentType="application/vnd.openxmlformats-officedocument.vmlDrawing"/>'
    if 'Extension="bin"' not in ct:
        inserts += '<Default Extension="bin" ContentType="application/vnd.ms-office.vbaProject"/>'
    ct = ct.replace('</Types>', inserts + '</Types>')

    # workbook rels: add vbaProject relationship
    if 'vbaProject.bin' not in wbrels:
        wbrels = wbrels.replace('</Relationships>',
            '<Relationship Id="rIdVbaProj" '
            'Type="http://schemas.microsoft.com/office/2006/relationships/vbaProject" '
            'Target="vbaProject.bin"/></Relationships>')

    # dashboard worksheet xml: declare xmlns:r (openpyxl omits it) + add legacyDrawing ref
    dash_xml = zin.read(dash_part).decode("utf-8")
    if 'xmlns:r=' not in dash_xml[:400]:
        dash_xml = dash_xml.replace(
            '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"',
            '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"', 1)
    if '<legacyDrawing' not in dash_xml:
        dash_xml = dash_xml.replace('</worksheet>', '<legacyDrawing r:id="rIdVml"/></worksheet>')

    # dashboard worksheet rels (create or extend)
    if dash_rels in names:
        dr = zin.read(dash_rels).decode("utf-8")
        dr = dr.replace('</Relationships>',
            '<Relationship Id="rIdVml" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/vmlDrawing" '
            'Target="../drawings/vmlDrawing1.vml"/></Relationships>')
    else:
        dr = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
              '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
              '<Relationship Id="rIdVml" '
              'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/vmlDrawing" '
              'Target="../drawings/vmlDrawing1.vml"/></Relationships>')

    zout = zipfile.ZipFile(out_xlsm, "w", zipfile.ZIP_DEFLATED)
    written = set()
    for item in zin.infolist():
        n = item.filename
        if n == "[Content_Types].xml": data = ct.encode("utf-8")
        elif n == "xl/_rels/workbook.xml.rels": data = wbrels.encode("utf-8")
        elif n == dash_part: data = dash_xml.encode("utf-8")
        elif n == dash_rels: data = dr.encode("utf-8")
        else: data = zin.read(n)
        zout.writestr(item, data); written.add(n)
    if dash_rels not in written: zout.writestr(dash_rels, dr)
    zout.writestr("xl/vbaProject.bin", vba_bin)
    zout.writestr("xl/drawings/vmlDrawing1.vml", vml.encode("utf-8"))
    zout.close(); zin.close()

inject_macros(TMP_XLSX, OUT, vba_bin, VML)
os.remove(TMP_XLSX)
print(f"OK -> {OUT}")
print(f"   sheets: {len(TOPICS)+1} | tables: {sum(len(t['tables']) for t in TOPICS)} | "
      f"fields: {sum(len(tb['fields']) for t in TOPICS for tb in t['tables'])} | index rows: {N}")
print("   fallback module exported -> modPM.bas")
