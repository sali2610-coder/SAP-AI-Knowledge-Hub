# -*- coding: utf-8 -*-
"""
Generator: SAP PM (Plant Maintenance) ECC 6.0 - Master Blueprint & Learning Syllabus
Builds a stylized, RTL, Hebrew-first Excel workbook (12 sheets) for a Junior PM Implementer.

Scope note (per repo CLAUDE.md rule 3): the indexed corpus is S/4HANA only. This workbook is
explicit ECC 6.0 model knowledge requested by the user (opt-in). Verify version-specific items
(table/field availability, screen-config transactions) against SAP Help Portal before production use.

Run:  python3 scripts/gen_pm_workbook.py
Out:  SAP_PM_ECC6_Blueprint.xlsx  (repo root)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------------------
# Theme palette (header fill + font color). ASCII '-' only per project rules.
# ----------------------------------------------------------------------------
THEMES = {
    "master":   {"hdr": "1F3864", "font": "FFFFFF", "band": "DCE3F2", "tab": "1F3864"},  # Dark Blue  - Master Data
    "trans":    {"hdr": "1E5E34", "font": "FFFFFF", "band": "DCEFE0", "tab": "1E5E34"},  # Deep Green - Transactional
    "config":   {"hdr": "7B1E2B", "font": "FFFFFF", "band": "F0DCE0", "tab": "7B1E2B"},  # Maroon     - Customizing
    "cross":    {"hdr": "C55A11", "font": "FFFFFF", "band": "FBE6D5", "tab": "C55A11"},  # Orange     - Cross-Module
}

COLUMNS = [
    ("מספר סידורי (ID)", 9),
    ("טרנזקציות מרכזיות (Main T-codes)", 24),
    ("טרנזקציות משניות (Secondary T-codes)", 26),
    ("טבלאות בסיס נתונים מרכזיות (Core Tables)", 26),
    ("שדות בולטים וחשובים (Key Fields)", 40),
    ("פונקציות מרכזיות / BAPIs (Key Functions)", 34),
    ("תיאור פונקציונלי של הפונקציות", 50),
    ("תוכניות ודוחות רקע (Background Programs)", 26),
    ("תיאור ומטרת התוכנית", 50),
]
# 0-based indexes of columns whose text must be bold (T-codes, tables, fields).
BOLD_COLS = {1, 2, 3, 4}

# ----------------------------------------------------------------------------
# Content. Each sheet: (title, theme, [rows]). Each row = 9 cells matching COLUMNS.
# ----------------------------------------------------------------------------
SHEETS = []

# === Sheet 1 - Org Structure & Technical Objects Base ========================
SHEETS.append(("1. מבנה ארגוני ותשתית", "master", [
    ["1",
     "IL01 / IL02 / IL03",
     "IL04 (יצירה מתבנית), IL05, IH01 (מבנה היררכי)",
     "IFLOT, IFLOTX, ILOA",
     "TPLNR (מיקום פונקציונלי / Functional Location), STRNO (מספר מבנה), TPLKZ (מחוון מבנה / StrIndicator), TPLMA (מיקום אב / Superior FL), OBJNR",
     "BAPI_FUNCLOC_CREATE, BAPI_FUNCLOC_CHANGE, BAPI_FUNCLOC_GETDETAIL",
     "יצירה / שינוי / שליפת פרטי מיקום פונקציונלי דרך ממשק תקני; שומר על תקפות מבנה היררכי וירושת נתוני מיקום/חשבון.",
     "RIIFLO20",
     "תוכנית הרשימה מאחורי IH06 - שליפת מיקומים פונקציונליים לפי מבנה, סוג ושימוש; בסיס לדוחות מצאי טכני."],
    ["2",
     "IL03 (תצוגת מבנה)",
     "OIFN (קטגוריות), OIMR (מחוון מבנה), OIA1 (מסיכות עריכה)",
     "T370F, T370FT, T370S, T370U",
     "FLTYP (קטגוריית מיקום / FL Category), STRIN (מחוון מבנה), EDITM (מסיכת עריכה / Edit Mask), HIERA (רמת היררכיה)",
     "FILABEL_CHECK, FUNCT_LOCATION_NUMBER_CHECK",
     "אימות תקינות תווית/מבנה המיקום מול מסיכת העריכה והמחוון; אכיפת מבנה קוד היררכי (לדוגמה AAA-BBB-CCC).",
     "RIFLET00",
     "כלי תחזוקת קטגוריות ומסיכות עריכה; משמש בעת הגדרת מחוון מבנה חדש ב-Customizing."],
    ["3",
     "IR01 / IR02 / IR03",
     "IR04 (העתקה), CR05 (היררכיית מרכזים), IRC1 (קיבולות)",
     "CRHD, CRTX, CRCA, CRCO, KAKO",
     "ARBPL (מרכז עבודה / Work Center), OBJID, OBJTY ('A' = Work Center), WERKS (מפעל), VERWE (שימוש / Usage), VERAN (אחראי)",
     "BAPI_WORKCENTER_CREATE (RFC), CR_WORKCENTER_READ",
     "יצירה ושליפה של מרכז עבודה כולל שיוך מרכז עלות, קיבולת ושימוש בפק\"ע; מחבר בין משאב ביצוע לעלות.",
     "RCRACO00",
     "דוח מרכזי עבודה וקיבולות; בדיקת עומסים והקצאת קיבולת מול דרישות פק\"ע."],
    ["4",
     "(Customizing) SPRO",
     "OIBR (אישורי עבודה / Permits), OIBP (פרופיל אישורים)",
     "T357, T357G, T357G_T",
     "PMGEW (קבוצת אישורים), GEWRK, BEGRU (קבוצת הרשאה / Auth Group), GEWRK (קוד אישור עבודה)",
     "PERMIT_CHECK, OBJECT_PERMIT_READ",
     "ניהול אישורי עבודה (Permits) ושיוכם לאובייקטים טכניים; חוסם שחרור פק\"ע עד קבלת אישור בטיחות.",
     "RIPERMIT_LIST",
     "סקירת אישורי עבודה פעילים/חסרים על פני אובייקטים טכניים לפני שחרור פקודות."],
    ["5",
     "IL02 (שיוך נתוני מיקום)",
     "IB11 (עץ מוצר מיקום), IK01 (נקודות מדידה)",
     "ILOA",
     "KOSTL (מרכז עלות / Cost Center), SWERK (מפעל אחזקה / Maint. Plant), INGRP (קבוצת תכנון / Planner Group), GEWRK, BEBER (אזור מפעל)",
     "ILOA_READ, ILOA_UPDATE",
     "טבלת ILOA מרכזת נתוני מיקום וחיוב (מרכז עלות, מפעל אחזקה, קבוצת תכנון) המשותפים למיקומים, ציוד ופק\"ע.",
     "RILOAA00",
     "התאמת נתוני מיקום/חיוב באובייקטים טכניים; אכיפת עקביות בין ILOA לבין רשומות הציוד/המיקום."],
]))

# === Sheet 2 - Equipment Master Data =========================================
SHEETS.append(("2. ציוד ונתוני מאסטר", "master", [
    ["1",
     "IE01 / IE02 / IE03",
     "IE05 (שינוי רשימתי), IE07 (רשימה רב-רמתית), IE10 (יצירה מרובה)",
     "EQUI, EQKT",
     "EQUNR (מספר ציוד / Equipment), EQTYP (סוג ציוד / Category), EQART (סוג טכני), OBJNR, HERST (יצרן), TYPBZ (דגם)",
     "BAPI_EQUI_CREATE, BAPI_EQUI_CHANGE, BAPI_EQUI_GETDETAIL, BAPI_EQMT_CREATE",
     "יצירה / שינוי / שליפת ציוד דרך ממשק תקני כולל טקסטים, נתוני מיקום וסיווג; הליבה לאוטומציית מאסטר ציוד.",
     "RIEQUI20",
     "תוכנית הרשימה מאחורי IH08 - שליפת ציוד לפי מפעל, סוג, מיקום וקבוצת תכנון; בסיס לדוחות מצאי."],
    ["2",
     "IE02 (פלחי זמן)",
     "IE03 (תצוגת היסטוריה), IH08 (רשימת ציוד)",
     "EQUZ, ILOA",
     "DATBI (תקף עד / Valid To), HEQUI (ציוד עליון / Superior Eq.), HEQNR, TPLNR (מיקום משויך), EQLFN (מונה פלח)",
     "EQUIPMENT_READ_SINGLE, EQUI_TIMESEGMENT_READ",
     "EQUZ מנהל את פלחי הזמן של הציוד (Time Segment) - שינויי התקנה/פירוק ושיוך היררכי לאורך זמן.",
     "RIEQUI00",
     "ניתוח שינויי ההתקנה וההיסטוריה של הציוד; מעקב אחר התקנות/פירוקים בין מיקומים וציוד אב."],
    ["3",
     "IQ01 / IQ02 / IQ03",
     "IQ04 (יצירה מרובה), IQ08 (רשימת סריאלים), IQ09",
     "OBJK, SER01, SER03, EQUI",
     "SERNR (מספר סידורי / Serial No.), MATNR (חומר), SERNP (פרופיל סריאליזציה), OBKNR, LFBNR (מסמך תנועה)",
     "BAPI_OBJCL_CREATE, SERIAL_NUMBER_CREATE, SERNR_ADD_TO_DOCUMENT",
     "ניהול מספרים סידוריים (Serialization) ושיוכם לתעודות מלאי ולרשומת ציוד; מקשר חומר פיזי לאובייקט טכני.",
     "RISERNR0",
     "דוח מספרים סידוריים מול תנועות מלאי; איתור פריטים סריאליים בתעודות רכש/מסירה."],
    ["4",
     "MM01 / MM02 (תצוגות PM)",
     "IE4N (התקנה/פירוק), IE36, IE37",
     "MARA, MARC, MAPL",
     "MATNR, WERKS, SERLV (רמת סריאליזציה), SERAIL, INSMK (מחוון אחזקה)",
     "BAPI_MATERIAL_GET_DETAIL, MATERIAL_READ_PLANT",
     "תצוגות ה-PM בחומר (Material Master) קובעות אם חומר ניתן לסריאליזציה ולשיוך כציוד; בסיס להתקנה/פירוק.",
     "RMMMBIM0",
     "ניתוח נתוני חומר רלוונטיים לאחזקה (סריאליזציה, מפעלי אחזקה) לאיתור חוסרים בהגדרה."],
    ["5",
     "IE02 (סעיף מיקום/חיוב)",
     "IL02 (מקור הירושה)",
     "ILOA, EQUZ",
     "KOSTL (מרכז עלות יונק/לא יונק), SWERK (מפעל אחזקה), INGRP (קבוצת תכנון), ANLNR (נכס קבוע), GEWRK",
     "ILOA_INHERIT_FROM_FUNCLOC, EQUI_ILOA_UPDATE",
     "לוגיקת \"יונק / לא יונק\": שדות מרכז עלות, מפעל אחזקה וקבוצת תכנון יורשים אוטומטית מהמיקום הפונקציונלי בעת ההתקנה, אלא אם דרוסו ידנית בציוד.",
     "RIEQUI20",
     "אימות עקביות ירושת נתוני מיקום/חיוב בין מיקום פונקציונלי לציוד המותקן עליו."],
]))

# === Sheet 3 - PM Bills of Material ==========================================
SHEETS.append(("3. עצי מוצר של אחזקה", "master", [
    ["1",
     "IB01 / IB02 / IB03",
     "IB05 (סוג מבנה / Construction), IB18 (רשימה)",
     "EQST, STKO, STPO",
     "EQUNR (ציוד), STLNR (מספר עץ מוצר / BOM No.), STLTY (קטגוריית BOM = 'E'), STLAL (חלופה), POSNR (פריט), IDNRK (רכיב)",
     "CSAP_MAT_BOM_MAINTAIN, CS_BOM_EXPL_EQU_RC1",
     "יצירה ותחזוקה של עץ מוצר לציוד (Equipment BOM) ופיצוצו (Explosion) לצורך תכנון רכיבים בפק\"ע.",
     "RCSBI010",
     "פיצוץ עץ מוצר רב-רמתי; הצגת מבנה הרכיבים לתכנון חומרים והזמנות."],
    ["2",
     "IB11 / IB12 / IB13",
     "IB15 (רשימה), IB08 (היכן בשימוש)",
     "TPST, STKO, STPO",
     "TPLNR (מיקום פונקציונלי), STLNR, STLTY (קטגוריה = 'T'), STLKN (צומת), MENGE (כמות), MEINS",
     "CSAP_MAT_BOM_READ, CS_BOM_EXPL_FLO_RC1",
     "עץ מוצר למיקום פונקציונלי (Functional Location BOM); מאפשר תכנון רכיבים ברמת המיקום ולא רק הציוד.",
     "RCSBI020",
     "פיצוץ והצגת עצי מוצר של מיקומים פונקציונליים; ניתוח רכיבים נדרשים לתחזוקה."],
    ["3",
     "CS01 / CS02 / CS03",
     "CS11 (רב-רמתי), CS12, CS15 (היכן בשימוש)",
     "MAST, STKO, STPO, STAS",
     "MATNR (חומר), WERKS, STLNR, STLTY ('M'), STLAL, STLKN, IDNRK, POSTP (סוג פריט)",
     "CSAP_MAT_BOM_CREATE, CSAP_MAT_BOM_MAINTAIN, BAPI_MATERIAL_BOM_GROUP_CREATE",
     "יצירה ותחזוקה של עץ מוצר לחומר (Material BOM); הבסיס המשותף לפיצוץ רכיבים בכל סוגי ה-BOM.",
     "RCSBI010 / CS_BT",
     "פיצוץ עץ מוצר לחומר; דוח 'היכן בשימוש' (Where-Used) לאיתור השפעות שינוי רכיב."],
    ["4",
     "(Customizing) OISD",
     "IB11 (סוג מבנה), IB01",
     "MAST, MARA",
     "STLAN (שימוש עץ מוצר / BOM Usage = 4 אחזקה), CSTMAT (חומר סוג מבנה / Construction Type), STLST (סטטוס)",
     "CONSTRUCTION_TYPE_READ, BOM_USAGE_CHECK",
     "סוג מבנה (Construction Type) מקשר חומר תבנית לאובייקט טכני, ומאפשר שכפול עץ מוצר סטנדרטי בין יחידות זהות.",
     "RCSBI040",
     "ניתוח שימושי עץ מוצר (BOM Usage) ושיוך סוגי מבנה; אחידות תצורה בין אובייקטים זהים."],
]))

# === Sheet 4 - Measuring Points & Counters ===================================
SHEETS.append(("4. נקודות מדידה ומונים", "master", [
    ["1",
     "IK01 / IK02 / IK03",
     "IK07 (רשימת נקודות), IK08 (שינוי רשימתי)",
     "IMPTT",
     "POINT (נקודת מדידה / Measuring Point), MPOBJ (אובייקט מדידה), MPTYP (סוג נקודה), ATINN (מאפיין), MPDAU (מונה / Counter), INDCT (קוד כיוון)",
     "BAPI_MEASUREMENTPOINT_CREATE, BAPI_MEASUREMENTPOINT_GETLIST, MEASUREMENT_POINT_READ",
     "יצירה ושליפה של נקודות מדידה ומונים על אובייקטים טכניים; מגדיר מאפיין נמדד ויחידת מדידה.",
     "RIMPOINT",
     "רשימת נקודות מדידה ומונים לפי אובייקט; סקירת הגדרות מדידה לפני קליטת קריאות."],
    ["2",
     "IK11 / IK12 / IK13",
     "IK21 / IK22 (קליטה מרוכזת), IK41 (רשימת מסמכים)",
     "IMRG, IMPTT",
     "MDOCM (מסמך מדידה / Measurement Doc.), READG (קריאה / Reading), RECDV (ערך נמדד), CNTRR (קריאת מונה), CDIFF (הפרש מונה), IDATE (תאריך)",
     "BAPI_MEASUREMENTDOCUM_CREATE, BAPI_MEASUREMENTDOCUM_CREATEM, MEASUREM_DOCUM_RFC_SINGLE_001",
     "יצירת מסמכי מדידה (Measurement Documents) - קליטת קריאות מונה/ערך; בסיס לתחזוקה מבוססת ביצועים.",
     "RIMRGRDT (IK21)",
     "קליטה מרוכזת של קריאות מונה לאובייקטים רבים; הזנת קריאות תקופתיות לטריגר אחזקה מונעת."],
    ["3",
     "IK13 (היסטוריית קריאות)",
     "IK17 (רשימת מסמכי מדידה), IP10 (תזמון)",
     "IMRG, IMRC",
     "CNTRC (קריאה כוללת מצטברת), CANCL (מחוון ביטול), OVERFLOW (גלישת מונה), READGINT (קריאה פנימית), TOTRG (ערך כולל)",
     "MEASUREMENT_COUNTER_OVERFLOW, MEASUREM_TRANSFER_TO_PLAN",
     "לוגיקת גלישת מונה (Counter Overflow) ומעבר קריאות: כאשר המונה מתאפס, המערכת מחשבת קריאה מצטברת רציפה ומעבירה אותה לתכנית האחזקה.",
     "RISTRA20",
     "ניטור מועדים (Deadline Monitoring) מבוסס מונה; חישוב מועדי אחזקה הבאים לפי קצב צריכה."],
    ["4",
     "(Customizing) OIH1",
     "IK04 (יצירה מתבנית)",
     "T370P, IMPTT",
     "PSORT (מיון נקודות), DECIM (ספרות עשרוניות), MSEHI (יחידת מידה / UoM), ATWRT (ערך מאפיין)",
     "CHARACT_MEASURING_POINT_READ",
     "שיוך מאפיין סיווג (Characteristic) לנקודת מדידה והגדרת יחידות/דיוק; קובע את טווח הערכים החוקי לקריאה.",
     "RIMPCHAR",
     "בדיקת התאמת מאפייני סיווג לנקודות מדידה; אכיפת יחידות מידה ודיוק אחיד."],
]))

# === Sheet 5 - Catalogs, Codes & Profiles ====================================
SHEETS.append(("5. קטלוגים, קודים ופרופילים", "config", [
    ["1",
     "QS41 (סוגי קטלוג)",
     "QS42 (סוגי קוד), QS43 (קודים)",
     "TQ15, TQ15T",
     "KATALOGART (סוג קטלוג / Catalog Type), KATART, KZLAR (סוג קוד), KATALOGART_TXT",
     "CATALOG_TYPE_READ, QPK1_CATALOG_READ",
     "ניהול סוגי קטלוג (Catalog Types): B=סיבות, C=פעולות, D=ליקויים, 5=אובייקט חלקי, 2=פעולות הודעה.",
     "RQEEAW01",
     "סקירת מבנה הקטלוגים והסוגים; בסיס להגדרת קודי ליקוי/סיבה/פעולה תקניים."],
    ["2",
     "QS51 (קבוצות קוד)",
     "QS61 (תצוגת קודים), QS31",
     "QPGR, QPCD, QPGT, QPCT",
     "CODEGRUPPE (קבוצת קוד / Code Group), CODE (קוד), KURZTEXT (טקסט קצר), KATALOGART, STATUS (סטטוס קבוצה)",
     "QPK1_CODEGROUP_READ, QPK1_CODE_TEXT_READ",
     "תחזוקת קבוצות קוד וקודים (Code Groups & Codes) בתוך כל סוג קטלוג; היחידות הבסיסיות לסיווג ליקויים/סיבות.",
     "RQEEAW02",
     "דוח קבוצות קוד וקודים לפי קטלוג; אימות שלמות הקידוד לפני שיוכו לפרופיל."],
    ["3",
     "(Customizing) OIM1",
     "QS41 (קטלוגים), OIN1",
     "T352, T352T",
     "TXTCD (פרופיל קטלוג / Catalog Profile), QMART (סוג הודעה), MATKL, KATALOGART, AUSWAHLMGE (קבוצת בחירה / Selection Set)",
     "CATALOG_PROFILE_READ, T352_READ",
     "פרופיל קטלוג (Catalog Profile) מגדיר אילו קבוצות קוד מותרות לכל אובייקט חלקי, ליקוי, סיבה, פעולה ומשימה בהודעה.",
     "RIPROFCAT",
     "ניתוח פרופילי קטלוג והשיוך שלהם לסוגי הודעה ולקטגוריות ציוד; בקרת אחידות קידוד."],
    ["4",
     "(Customizing) OIN4",
     "IE02 (שיוך פרופיל לציוד)",
     "T352B, T352B_T",
     "RBNR (פרופיל קטלוג בציוד), AUSWMENGEN, CODEGRUPPE, MERKMAL (מאפיין), EQART",
     "REPORT_PROFILE_DETERMINE",
     "T352B מקשרת בין פרופיל קטלוג לקבוצות הבחירה; קובעת אילו קודים זמינים לפי קטגוריית ציוד/סוג הודעה (קישור Profile->Object).",
     "RIQPROFILE",
     "מיפוי שיוך פרופילי קטלוג לקטגוריות ציוד ולסוגי הודעה; איתור אובייקטים ללא פרופיל."],
]))

# === Sheet 6 - Maintenance Notifications =====================================
SHEETS.append(("6. הודעות אחזקה", "trans", [
    ["1",
     "IW21 / IW22 / IW23",
     "IW24 (תקלה M2), IW25 (פעילות M3), IW26 (M1)",
     "QMEL",
     "QMNUM (מספר הודעה / Notification), QMART (סוג הודעה M1/M2/M3), QMTXT (תיאור), OBJNR, EQUNR, TPLNR, QMKAT/QMGRP (קוד ליקוי)",
     "BAPI_ALM_NOTIF_CREATE, BAPI_ALM_NOTIF_SAVE, BAPI_ALM_NOTIF_GET_DETAIL, BAPI_ALM_NOTIF_CLOSE",
     "יצירה / שמירה / סגירה ושליפת הודעת אחזקה דרך ממשק תקני; הליבה לקליטת תקלות ובקשות אוטומטית.",
     "RIQMEL20",
     "תוכנית הרשימה מאחורי IW28/IW29 - שליפת הודעות לפי סטטוס, סוג, אובייקט ותקופה; בסיס לדוחות PMIS."],
    ["2",
     "IW22 (פריטים)",
     "IW64 (היסטורית הודעות)",
     "QMFE, QMSM",
     "FENUM (מספר פריט / Item), FEGRP/FECOD (קוד ליקוי - קבוצה/קוד), OTGRP/OTEIL (אובייקט חלקי / Object Part), MANUM (משימה / Task)",
     "BAPI_ALM_NOTIF_DATA_ADD, NOTIF_ITEM_READ",
     "QMFE מנהל את פריטי ההודעה (Items) - תיאור הליקוי ומיקומו על האובייקט החלקי; QMSM מנהל משימות לתיקון.",
     "RIQMEL00",
     "ניתוח פריטי ליקוי לפי אובייקט וקוד; זיהוי דפוסי כשל חוזרים (Failure Analysis)."],
    ["3",
     "IW22 (סיבות/פעולות)",
     "IW66 / IW67 (ניתוח פריטים)",
     "QMUR, QMMA",
     "URNUM (מספר סיבה), URGRP/URCOD (קוד סיבה / Cause), MNGRP/MNCOD (קוד פעילות / Activity), PETXT (טקסט)",
     "NOTIF_CAUSE_READ, NOTIF_ACTIVITY_READ",
     "QMUR מנהל סיבות לליקוי (Causes) ו-QMMA מנהל פעילויות שבוצעו (Activities); הבסיס לניתוח שורש (RCA) ולתיעוד פעולות.",
     "RIQMUR00",
     "דוח סיבות ופעילויות לפי אובייקט/קוד; תמיכה בניתוח אמינות (Reliability) ו-MTBF."],
    ["4",
     "(Customizing) OIAL",
     "OIM1 (פרופיל קטלוג), QCC0",
     "TQ80, T356 (עדיפויות), TQ8T",
     "QMART, RBNR (פרופיל קטלוג), PRIOK (עדיפות / Priority), ARTPR (סוג עדיפות), AUART (סוג פק\"ע ברירת מחדל)",
     "PM_NOTIFICATION_TYPE_READ, PRIORITY_DETERMINE",
     "OIAL קובע את פריסת המסך (Screen Layout) לפי סוג הודעה; שיוך פרופיל קטלוג, סוג עדיפות וסוג פק\"ע ברירת מחדל.",
     "RISCREEN_NOTIF",
     "בקרת פריסות מסך ושדות חובה לפי סוג הודעה; אכיפת מדיניות קליטה אחידה."],
    ["5",
     "IW21 (סטטוס מערכת)",
     "BS22 (סטטוסי מערכת), IW23",
     "JEST, JCDS",
     "OBJNR (מפתח אובייקט), STAT (סטטוס OSNO/OSTS), INACT (פעיל/לא פעיל), CHGNR (מספר שינוי), USNAM",
     "STATUS_READ, STATUS_CHANGE_INTERN",
     "אינטגרציה עם ניהול סטטוסים (Status Management): פתיחה (OSNO), הושלם (NOCO/MSPT) ומחיקה - שולטים בפעולות המותרות על ההודעה.",
     "RIQMEL20",
     "סינון הודעות לפי סטטוס מערכת/משתמש; מעקב אחר זרימת הטיפול (Workflow) של תקלות."],
]))

# === Sheet 7 - Maintenance Orders ============================================
SHEETS.append(("7. פקודות עבודה (פק\"ע)", "trans", [
    ["1",
     "IW31 / IW32 / IW33",
     "IW34 (מהודעה), IW36 (תת-פק\"ע), IW38/IW39 (רשימה)",
     "AUFK, AFKO",
     "AUFNR (מספר פק\"ע / Order), AUART (סוג פק\"ע PM01/PM02/PM03), OBJNR, AUFPL (מס' תוכנית פעולות), KTEXT (תיאור), GSTRP (תאריך התחלה בסיסי)",
     "BAPI_ALM_ORDER_MAINTAIN, BAPI_ALM_ORDER_GET_DETAIL, BAPI_ALM_ORDERHEAD_GET_LIST",
     "יצירה / שינוי / שליפה של פק\"ע דרך ממשק תקני אחד; מטפל בכותרת, פעולות, רכיבים וכלי עבודה במחזור חיים שלם.",
     "RIAUFK20",
     "תוכנית הרשימה מאחורי IW38/IW39 - שליפת פק\"ע לפי סטטוס, סוג, אובייקט ותקופה; בסיס לדוחות תפעול."],
    ["2",
     "IW32 (פעולות)",
     "IW37N (פעולות+רכיבים), CM01 (קיבולת)",
     "AFVC, AFVV, AFFL",
     "VORNR (מספר פעולה / Operation), ARBID (מרכז עבודה), STEUS (מפתח בקרה / Control Key), ARBEI (עבודה מתוכננת), DAUNO (משך), LARNT (סוג פעילות)",
     "BAPI_ALM_ORDER_MAINTAIN (METHOD=OPERATION), CO_ZF_OPERATIONS_READ",
     "AFVC/AFVV מנהלים את פעולות הפק\"ע (Operations) - מרכז עבודה, מפתח בקרה (PM01=פנימי, PM02=חיצוני), זמני עבודה וקיבולת.",
     "RIAFVC00",
     "ניתוח פעולות ועומסי קיבולת על פני פקודות; תכנון כוח אדם ומשאבים (Capacity Planning)."],
    ["3",
     "IW32 (רכיבים)",
     "MB1A (פליטה למלאי), ME57 (הקצאת PR)",
     "RESB, AFPO",
     "RSNUM (מספר הזמנה / Reservation), RSPOS, MATNR (רכיב), BDMNG (כמות נדרשת), POSNR (פריט פק\"ע), BESKZ (סוג רכש)",
     "BAPI_ALM_ORDER_MAINTAIN (METHOD=COMPONENT), RESERVATION_READ",
     "RESB מנהל הזמנות רכיבים (Component Reservations); רכיבי מלאי מייצרים הזמנת מלאי, רכיבים שאינם במלאי מייצרים דרישת רכש (PR) אוטומטית.",
     "RIAFPO00",
     "ניתוח רכיבים והזמנות לפי פק\"ע; בקרת זמינות חומרים (Availability Check) לפני שחרור."],
    ["4",
     "IW32 (תמחיר/חיוב)",
     "IW31 (DP90 חיוב), KO88 (התחשבנות)",
     "AUFK, COEP, AFRU",
     "AUFNR, OBJNR, KSTAR (אלמנט עלות), WRTTP (סוג ערך - מתוכנן/בפועל), VKORG (ארגון מכירות לחיוב), KDAUF (הזמנת לקוח)",
     "BAPI_ALM_ORDER_MAINTAIN (COSTING), K_ORDER_COSTING_CALCULATE",
     "יכולות אינטגרציה של הפק\"ע: יצירת PR/PO לשירותים חיצוניים, הזמנת רכיבים, חיוב/מזכר חיוב (Debit Memo) ללקוח, ותכנון קיבולת.",
     "RKAUF000",
     "סקירת עלויות מתוכננות מול בפועל ברמת הפק\"ע; בסיס לתמחיר ולמזכרי חיוב."],
    ["5",
     "IW31 (שחרור REL)",
     "IW3K (אישור), IW41/IW42 (דיווח)",
     "AFRU, AFWI, JEST",
     "RUECK (מספר דיווח / Confirmation), RMZHL, ISMNW (עבודה בפועל), AUERU (דיווח סופי), OBJNR, STAT (REL/TECO/CLSD)",
     "BAPI_ALM_CONF_CREATE, BAPI_ALM_CONF_GETLIST, BAPI_ALM_CONF_CANCEL",
     "AFRU מנהל דיווחי ביצוע (Confirmations) על פעולות; AFWI מקשר תנועות מלאי שנרשמו בדיווח. שחרור הפק\"ע (REL) מאפשר רישום זמן וחומרים.",
     "RIAFRU00",
     "ניתוח דיווחי ביצוע מול תכנון; מעקב אחר זמני עבודה בפועל וסגירת פעולות."],
]))

# === Sheet 8 - Status Management =============================================
SHEETS.append(("8. ניהול סטטוסים", "config", [
    ["1",
     "(תצוגה) IW33 / IE03",
     "BS22 (סטטוסי מערכת), BS23 (תצוגה)",
     "JEST, JSTO",
     "OBJNR (מפתח אובייקט), STAT (סטטוס מערכת), INACT (לא פעיל), OBTYP (סוג אובייקט), STSMA (פרופיל סטטוס)",
     "STATUS_READ, STATUS_OBJECT_INFORMATION, STATUS_CHANGE_INTERN",
     "JEST שומר כל סטטוס פעיל/לא-פעיל לכל אובייקט (Object Status). JSTO מחזיק את כותרת אובייקט הסטטוס ואת פרופיל הסטטוס המשויך.",
     "RISTAT00",
     "ניתוח סטטוסים פעילים לאובייקטים (פק\"ע/הודעה/ציוד); איתור אובייקטים תקועים בסטטוס מסוים."],
    ["2",
     "(מערכת) CRTD/REL/TECO/CLSD",
     "IW32 (שינוי סטטוס), IW33",
     "TJ02, TJ02T",
     "ISTAT (סטטוס מערכת פנימי), TXT04 (CRTD/REL/TECO/CLSD), TXT30 (תיאור ארוך), SPRAS",
     "STATUS_TEXT_READ, SYSTEM_STATUS_CHECK",
     "סטטוסי מערכת (System Status) קבועים מראש: CRTD (נוצר), REL (שוחרר), TECO (הושלם טכנית), CLSD (סגור עסקית). שולטים בפעולות המותרות.",
     "RITJ0200",
     "טבלת מיפוי קודי סטטוס מערכת לתיאורים; עזר בפענוח שדה OBJNR/STAT בדוחות."],
    ["3",
     "(משתמש) שינוי ב-IW32",
     "BS02 (פרופיל סטטוס), BS03",
     "TJ30, TJ30T, TJ20",
     "STSMA (פרופיל סטטוס / Status Profile), ESTAT (סטטוס משתמש), TXT04, VORNR (סדר), BERSL (תהליך עסקי / Business Transaction)",
     "STATUS_CHANGE_EXTERN, USER_STATUS_CHECK",
     "סטטוס משתמש (User Status) מוגדר בפרופיל (Status Profile) ב-BS02; מאפשר/חוסם תהליכים עסקיים ומחייב סדר מעבר בין סטטוסים.",
     "RIBS0200",
     "סקירת פרופילי סטטוס משתמש והתהליכים שהם חוסמים/מתירים; בקרת זרימת אישורים."],
    ["4",
     "(Customizing) OIBS",
     "BS02 (תחזוקת פרופיל), OIBS",
     "TJ30, TJ02T, TJ04",
     "STSMA, ESTAT, OKVOR (תהליך מותר / Set), VRGNG (קוד תהליך), INACT, ANWND (יישום = PM)",
     "STATUS_CHANGE_INTERN, OBJECT_STATUS_RAISE",
     "OIBS משייך פרופיל סטטוס לסוגי פק\"ע/הודעה ב-PM; כל סטטוס מגדיר אילו תהליכים (כגון פליטת מלאי) מותרים/אסורים - לדוגמה חסימת GI לפני שחרור.",
     "RISTSMA0",
     "מיפוי שיוך פרופילי סטטוס לסוגי אובייקט; אכיפת חוקי חסימה עסקיים אחידים."],
]))

# === Sheet 9 - PM-MM Integration =============================================
SHEETS.append(("9. אינטגרציית מלאי ורכש (PM-MM)", "cross", [
    ["1",
     "MB1A (תנועה 261)",
     "MIGO (פליטת מלאי), MB21 (הזמנה ידנית)",
     "RESB, MSEG, MKPF",
     "RSNUM (הזמנה), RSPOS, MATNR, BDMNG (נדרש), ENMNG (נפלט), BWART (סוג תנועה 261/262), AUFNR (פק\"ע)",
     "BAPI_GOODSMVT_CREATE, RESERVATION_CREATE, BAPI_RESERVATION_CREATE1",
     "פליטת חומר לפק\"ע (Goods Issue, תנועה 261); RESB מקור ההזמנה, MSEG/MKPF מתעדים את התנועה בפועל מול הפק\"ע.",
     "RM07RESL",
     "דוח הזמנות מלאי פתוחות (Reservation List) מול פק\"ע; ניתוח חוסרי פליטה וזמינות חומרים."],
    ["2",
     "ME51N (דרישת רכש)",
     "ME52N / ME53N, ME57 (הקצאה)",
     "EBAN, EBKN",
     "BANFN (מספר דרישת רכש / PR), BNFPO (פריט), MATNR, KNTTP (סוג חיוב = F פק\"ע), AUFNR, MENGE, PSTYP (סוג פריט)",
     "BAPI_REQUISITION_CREATE, BAPI_PR_CREATE, BAPI_REQUISITION_GETDETAIL",
     "דרישת רכש (PR) הנוצרת אוטומטית מפעולת שירות חיצוני (PM02) או מרכיב שאינו במלאי; EBKN מקשר את ה-PR לחיוב הפק\"ע.",
     "RM06BANF",
     "דוח דרישות רכש פתוחות לפי פק\"ע/מרכז עלות; מעקב אחר רכש שירותים וחומרים חיצוניים."],
    ["3",
     "ME21N (הזמנת רכש)",
     "ME22N / ME23N, ME2N (רשימה)",
     "EKKO, EKPO, EKKN",
     "EBELN (מספר הזמנת רכש / PO), EBELP, LIFNR (ספק), AUFNR (חיוב פק\"ע), KNTTP, NETPR (מחיר), PACKNO (שירותים)",
     "BAPI_PO_CREATE1, BAPI_PO_GETDETAIL, BAPI_PO_CHANGE",
     "הזמנת רכש (PO) לחומרים/שירותים חיצוניים בפק\"ע; EKKN מקשר את החיוב לפק\"ע, ומאפשר רישום עלות בפועל בעת קליטת הסחורה/השירות.",
     "RM06EM00",
     "דוח הזמנות רכש פתוחות מול פק\"ע; בקרת התחייבויות (Commitments) מול תקציב."],
    ["4",
     "MIGO (קליטה 101)",
     "MB31, MB1C (תנועה 561)",
     "MSEG, MKPF, MARD",
     "MBLNR (מסמך חומר / Material Doc.), ZEILE, BWART (101/102/261/262), DMBTR (סכום), MENGE, LGORT (מחסן), AUFNR",
     "BAPI_GOODSMVT_CREATE, BAPI_GOODSMVT_GETITEMS, MB_CREATE_GOODS_MOVEMENT",
     "MSEG/MKPF הם ליבת תנועות המלאי - קליטת סחורה (101) לשירות חיצוני בפק\"ע ופליטה (261) לרכיבי מלאי; כל תנועה מעדכנת עלות בפועל.",
     "RM07DOCS",
     "דוח מסמכי חומר (Material Document List) לפי פק\"ע/תנועה; ביקורת תנועות מלאי לאחזקה."],
]))

# === Sheet 10 - PM-CO Financial Integration ==================================
SHEETS.append(("10. עלויות, תקציבים והתחשבנות (PM-CO)", "cross", [
    ["1",
     "KO88 (התחשבנות יחידה)",
     "KO8G (התחשבנות מרוכזת), IW32 (כרטיס עלות)",
     "COSP, COSS",
     "OBJNR (אובייקט CO), KSTAR (אלמנט עלות / Cost Element), WRTTP (סוג ערך 01 מתוכנן/04 בפועל), WTG001..016 (ערך לפי תקופה), GJAHR (שנה)",
     "K_COSTS_READ, BAPI_INTERNALORDER_GETDETAIL, K_ORDER_SETTLEMENT",
     "COSP (עלויות חיצוניות / Primary) ו-COSS (עלויות פנימיות / Secondary) מצברים עלויות מתוכננות ובפועל לפי תקופה; בסיס להתחשבנות הפק\"ע.",
     "RKO7KO88",
     "תוכנית ההתחשבנות מאחורי KO88 - העברת עלויות בפועל מהפק\"ע ליעד ההתחשבנות (מרכז עלות / נכס / WBS)."],
    ["2",
     "IW32 (חוק התחשבנות)",
     "KO02 (תחזוקת חוק), KOSRLIST",
     "COBRA, COBRB",
     "OBJNR (מקור), BUREG (חוק התחשבנות / Settlement Rule), PERBZ (סוג תקופה), KONTY (סוג מקבל), EMPGE (מקבל), PROZS (אחוז / %), AVORG (תהליך)",
     "K_SETTLEMENT_RULE_READ, BAPI_INTERNALORDER_CREATE (with rule)",
     "COBRA (כותרת חוק התחשבנות) ו-COBRB (פריטי החוק / Distribution Rules) מגדירים למי ובאיזה אחוז מועברות עלויות הפק\"ע (מרכז עלות / נכס / WBS).",
     "RKACOR40",
     "דוח חוקי התחשבנות חסרים/שגויים לפני סגירת תקופה; אכיפת קיום חוק לכל פק\"ע."],
    ["3",
     "KSB1 (פריטי עלות מרכז)",
     "KOB1 (פריטי פק\"ע), S_ALR_87013611",
     "COEP, COBK",
     "BELNR (מסמך CO), KSTAR, OBJNR, WOGBTR (ערך אובייקט), BUKRS, REFBN (מסמך מקור), VRGNG (תהליך עסקי)",
     "BAPI_ACC_DOCUMENT_POST (CO), K_LINE_ITEMS_READ",
     "COEP/COBK מנהלים את פריטי העלות הבודדים (Actual Line Items) של הפק\"ע; מקור לניתוח עלות מפורט לפי אלמנט ותהליך.",
     "RKAEP000",
     "דוח פריטי עלות בפועל לפי פק\"ע/מרכז עלות; פירוק עלויות חומרים, עבודה ושירותים חיצוניים."],
    ["4",
     "CJ20N (אלמנט WBS)",
     "IW32 (שיוך WBS), CJI3",
     "PRPS, COSP, DKOBR",
     "PSPNR (אלמנט WBS / WBS Element), POSID, OBJNR, PBUKR (חברה), PRART (סוג פרויקט), KOSTL (מרכז עלות יעד)",
     "BAPI_PS_PRECOMMIT, K_WBS_ELEMENT_READ, PROJECT_COSTS_READ",
     "שיוך פק\"ע לאלמנט WBS לפרויקטי השבתה/השקעה; DKOBR מתעד נתוני התחשבנות ברמת מקבל, ומאפשר התחשבנות לפרויקט (CO/PS).",
     "RCJ_DIST",
     "דוח עלויות פרויקט/WBS מאוחד עם פקודות אחזקה משויכות; בקרת תקציב פרויקטי אחזקה."],
]))

# === Sheet 11 - Preventive Maintenance & Task Lists ==========================
SHEETS.append(("11. אחזקה מונעת ותוכניות", "master", [
    ["1",
     "IA05 / IA06 (כללי)",
     "IA01/IA02 (ציוד), IA11/IA12 (מיקום), IA08",
     "PLKO, PLPO, PLAS",
     "PLNNR (קבוצת רשימת פעולות / Task List Group), PLNAL (חלופה), PLNTY (סוג רשימה A/E/T), VORNR (פעולה), ARBID (מרכז עבודה), STEUS (מפתח בקרה)",
     "BAPI_TASKLIST_CREATE (RFC), CARO_TASKLIST_READ, CP_DI_TASKLIST",
     "רשימות פעולות אחזקה (Task Lists): כללית (PLNTY=A), לציוד (E), למיקום (T). PLKO=כותרת, PLPO=פעולות, PLAS=שיוך.",
     "RIPLKO00",
     "דוח רשימות פעולות לפי קבוצה/שימוש; סטנדרטיזציה של תהליכי אחזקה חוזרים."],
    ["2",
     "IP01 / IP02 / IP03",
     "IP04 (פריט), IP10 (תזמון), IP19 (לוח גרפי)",
     "MPLA, MPOS",
     "WARPL (תכנית אחזקה / Maintenance Plan), WPTXT (תיאור), MPTYP (סוג תכנית), ZYKL1 (מחזור), STRAT (אסטרטגיה), WAPOS, EQUNR/TPLNR",
     "BAPI_MAINTENANCEPLAN_CREATE (RFC), IMPP_MAINTENANCE_PLAN_READ",
     "תכנית אחזקה (Maintenance Plan): MPLA=כותרת התכנית והתזמון, MPOS=פריט התכנית המגדיר את האובייקט, סוג הפק\"ע/ההודעה ורשימת הפעולות.",
     "RIMPLA00",
     "דוח תכניות אחזקה ופריטיהן; סקירת כיסוי אחזקה מונעת על פני אובייקטים טכניים."],
    ["3",
     "IP41 (מחזור יחיד)",
     "IP42 (אסטרטגיה), IP43 (מרובת מונים)",
     "MPLA, MMPT, T351",
     "ZEITL (יחידת מחזור זמן), STRAT (אסטרטגיה / Strategy), PAKET (חבילת אחזקה / Package), POINT (נקודת מדידה למונה), HORIZ (אופק שחרור), ABRHO (אופק שחרור)",
     "STRATEGY_PACKAGE_READ, MAINTENANCE_CYCLE_CALCULATE",
     "תכנית מבוססת זמן (IP41 מחזור יחיד) או אסטרטגיה (IP42 חבילות אחזקה), ומבוססת ביצועים (IP43, לפי מונה/נקודת מדידה).",
     "RISTRA10",
     "ניתוח אסטרטגיות וחבילות אחזקה; אופטימיזציה של תדירויות תחזוקה מונעת."],
    ["4",
     "IP10 (תזמון תכנית)",
     "IP30 (ניטור מועדים רקע), IP24",
     "MHIO, MHIS",
     "WARPL, ABNUM (מספר קריאה / Call No.), NPLDA (תאריך מתוכנן הבא), NTERM (מועד), HORIZ (אופק), ABRUF (סטטוס קריאה), AUFNR (פק\"ע שנוצרה)",
     "MAINTENANCE_PLAN_SCHEDULE, ISCHED_CALL_GENERATE, IMPP_SCHEDULE_PLAN",
     "MHIO=אובייקטי קריאת התכנית (Call Objects), MHIS=היסטוריית התזמון. תזמון מחשב מועד הבא ויוצר פק\"ע/הודעה כשמגיע אופק השחרור.",
     "RISTRA20",
     "תוכנית ניטור המועדים מאחורי IP30 - הרצת רקע יומית המשחררת קריאות אחזקה מונעת אוטומטית בכל המערכת."],
    ["5",
     "(Customizing) OIPM",
     "IP11/IP11Z (אסטרטגיה), OIP1",
     "T399W, T399A, T351P",
     "IWERK (מפעל מתכנן), AUART (סוג פק\"ע), TERMA (סוג תזמון), HORIZ (אופק שחרור), TERMT (פרמטר בקרה), CYCLE_MODIFICATION (פקטור התאמה)",
     "PLAN_CONTROL_PARAM_READ, SCHEDULING_PARAM_DETERMINE",
     "T399W/T399A מחזיקים פרמטרי בקרת תזמון ברמת מפעל (אופק שחרור, פקטור התאמת מחזור, סוג השלמה); קובעים את התנהגות ה-Scheduling.",
     "RIT399W0",
     "סקירת פרמטרי בקרת תזמון לפי מפעל; כיול אופקי שחרור והתאמות מחזור."],
]))

# === Sheet 12 - Maintenance History & Archiving ==============================
SHEETS.append(("12. היסטוריית אחזקה וארכיון", "trans", [
    ["1",
     "IW64 (היסטורית הודעות)",
     "IW65, IW66/IW67 (ניתוח פריטים)",
     "QMEL, RIHQMEL",
     "QMNUM (הודעה), QMART, AUSVN/AUSBS (תקופת השבתה / Breakdown), MAUEH (יחידת זמן השבתה), OBJNR, RIHQMEL (מבנה לוגי)",
     "BAPI_ALM_NOTIF_LIST_FILTER, NOTIF_HISTORY_READ",
     "RIHQMEL הוא מבנה מסד הנתונים הלוגי (Logical DB) לשליפת הודעות היסטוריות עם זמני השבתה לניתוח אמינות (MTBF/MTTR).",
     "RIQMEL20",
     "דוח היסטוריית הודעות לפי אובייקט/תקופה; חישוב זמינות (Availability) וזמני השבתה מצטברים."],
    ["2",
     "IW39 (פק\"ע גמורות)",
     "IW3L (היסטוריה), IW13 (לפי חומר)",
     "AUFK, RIHAUFK, ILOA",
     "AUFNR (פק\"ע), AUART, OBJNR, IDAT2 (תאריך TECO), ERDAT (נוצר), RIHAUFK (מבנה לוגי לפק\"ע), GSTRP/GLTRP (תאריכי תכנון)",
     "BAPI_ALM_ORDERHEAD_GET_LIST, ORDER_HISTORY_READ",
     "RIHAUFK הוא מבנה מסד הנתונים הלוגי לשליפת פקודות עבודה היסטוריות; בסיס לניתוח עלויות ותדירות תיקונים לאורך זמן.",
     "RIAUFK20",
     "דוח פק\"ע סגורות לפי אובייקט/עלות; ניתוח מגמות תחזוקה ועלות מחזור חיים (LCC)."],
    ["3",
     "SARA (ניהול ארכיון)",
     "AOBJ (אובייקטי ארכוב), IW64",
     "ADMI_RUN, ADMI_FILES",
     "ARCHIV_KEY, OBJECT (PM_ORDER / QMEL_DEL), DOCUMENT_NUMBER, ARCHIV_OFFS, RUNID, GENER_DATE",
     "ARCHIVE_OPEN_FOR_WRITE, ARCHIVE_GET_NEXT_OBJECT, ARCHIVE_DELETE_FROM_DB",
     "ניהול ארכוב (Archiving) דרך SARA לאובייקטים PM_ORDER (פקודות) ו-PM_QMEL (הודעות); סורק, כותב לארכיון ומוחק מה-DB החי.",
     "RIARCPM1",
     "תוכנית הכתיבה/מחיקה של ארכוב פקודות אחזקה; הקטנת נפח DB תוך שמירת היסטוריה נגישה."],
    ["4",
     "IHVORG (תצוגת הוראות)",
     "IW64 (היסטוריה), DB15 (תלויות)",
     "IHVORG, IHSG, IHGNS",
     "VORGN (קוד תהליך עסקי / Business Transaction), OBJNR, ACTVT, IHVORG (קישור תהליך-סטטוס), VRGNG, BERSL",
     "TRANSACTION_HISTORY_READ, OBJECT_VORGANG_CHECK",
     "IHVORG מקשר תהליכים עסקיים (Business Transactions) לבקרת סטטוס; קובע אילו פעולות מתועדות בהיסטוריית האובייקט הטכני.",
     "RIHIST00",
     "ניתוח יומן התהליכים העסקיים על אובייקטים טכניים; ביקורת (Audit Trail) של שינויי סטטוס ופעולות."],
]))

# ----------------------------------------------------------------------------
# Build workbook
# ----------------------------------------------------------------------------
thin = Side(style="thin", color="BBBBBB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()
wb.remove(wb.active)

# ---- Cover / index sheet ----
cover = wb.create_sheet("מפתח (Index)")
cover.sheet_view.rightToLeft = True
cover.sheet_properties.tabColor = "404040"
cover["A1"] = "SAP PM (Plant Maintenance) ECC 6.0 - תבנית אב ותכנית לימוד למיישם זוטר"
cover["A1"].font = Font(bold=True, size=16, color="1F3864")
cover["A2"] = "Master Blueprint & Learning Syllabus - Junior PM Implementer"
cover["A2"].font = Font(bold=True, size=11, color="7F7F7F")
cover["A4"] = ("הערת היקף: הקורפוס המאונדקס הוא S/4HANA בלבד. חוברת זו מבוססת ידע מודל מפורש על ECC 6.0 (לפי בקשת המשתמש). "
               "יש לאמת פריטים תלויי גרסה (טבלאות/שדות, טרנזקציות קונפיגורציה) מול SAP Help Portal לפני שימוש בייצור.")
cover["A4"].font = Font(italic=True, size=10, color="7B1E2B")
cover["A4"].alignment = Alignment(wrap_text=True, horizontal="right", vertical="top")
cover.merge_cells("A4:E6")
cover.column_dimensions["A"].width = 40
for c in "BCDE":
    cover.column_dimensions[c].width = 18

cover["A8"] = "גיליון (Sheet)"
cover["B8"] = "קטגוריה (Category)"
for col in ("A8", "B8"):
    cover[col].font = Font(bold=True, color="FFFFFF")
    cover[col].fill = PatternFill("solid", fgColor="404040")
    cover[col].border = border
    cover[col].alignment = Alignment(horizontal="right")

CAT_LABEL = {"master": "נתוני אב (Master Data)", "trans": "נתונים תנועתיים (Transactional)",
             "config": "קונפיגורציה (Customizing)", "cross": "אינטגרציה חוצת-מודולים (Cross-Module)"}
r = 9
for title, theme, _ in SHEETS:
    cover.cell(r, 1, title).alignment = Alignment(horizontal="right")
    cover.cell(r, 1).border = border
    lab = cover.cell(r, 2, CAT_LABEL[theme])
    lab.alignment = Alignment(horizontal="right")
    lab.border = border
    lab.fill = PatternFill("solid", fgColor=THEMES[theme]["band"])
    r += 1

# ---- Data sheets ----
def safe_tab(name):
    # Excel tab name: max 31 chars, no : \ / ? * [ ]
    bad = ':\\/?*[]'
    n = "".join(ch for ch in name if ch not in bad)
    return n[:31]

for title, theme, rows in SHEETS:
    th = THEMES[theme]
    ws = wb.create_sheet(safe_tab(title))
    ws.sheet_view.rightToLeft = True
    ws.sheet_properties.tabColor = th["tab"]

    n_cols = len(COLUMNS)

    # Title band (row 1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    tcell = ws.cell(1, 1, title)
    tcell.font = Font(bold=True, size=14, color="FFFFFF")
    tcell.fill = PatternFill("solid", fgColor=th["hdr"])
    tcell.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[1].height = 28

    # Header (row 2)
    for j, (label, width) in enumerate(COLUMNS, start=1):
        c = ws.cell(2, j, label)
        c.font = Font(bold=True, size=11, color=th["font"])
        c.fill = PatternFill("solid", fgColor=th["hdr"])
        c.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        c.border = border
        ws.column_dimensions[get_column_letter(j)].width = width
    ws.row_dimensions[2].height = 46

    # Data rows
    for i, row in enumerate(rows):
        excel_row = i + 3
        band = (i % 2 == 1)
        for j, val in enumerate(row):
            c = ws.cell(excel_row, j + 1, val)
            is_bold = j in BOLD_COLS
            c.font = Font(bold=is_bold, size=10,
                          color="1F3864" if is_bold else "222222")
            c.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
            c.border = border
            if band:
                c.fill = PatternFill("solid", fgColor=th["band"])
        ws.row_dimensions[excel_row].height = 96

    # Freeze header + ID column (RTL -> freeze keeps row 2 and col A visible)
    ws.freeze_panes = "B3"
    # Print setup
    ws.sheet_view.showGridLines = False
    ws.print_options.horizontalCentered = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

out = "SAP_PM_ECC6_Blueprint.xlsx"
wb.save(out)
print(f"OK -> {out}  ({len(SHEETS)} content sheets + index)")
