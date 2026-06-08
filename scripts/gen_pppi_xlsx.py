# -*- coding: utf-8 -*-
"""
SAP PP-PI ECC 6.0 -> S/4HANA migration Data Dictionary - clean macro-free, chart-free .xlsx.
CBC Coca-Cola Red / Dark Slate branding. Field-level grid (Field / Data Type / Length / Key /
English / Hebrew) + table-level JOIN ON (SQL/CDS), Hebrew functional guide, S/4 gap, SAP Help link.
Includes Dashboard (live KPIs + global FILTER search), Migration Cockpit (COUNTIF + status dropdown),
and ER - Join Map. VBA exported as plain text (pppi_modPM.bas).

Run:  python3 scripts/gen_pppi_xlsx.py   ->  SAP_PPPI_ECC6_to_S4_Migration.xlsx + pppi_modPM.bas
"""
import os, sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from pppi_data import TOPICS, OPS

# ---- CBC brand palette ----
FONT = "Segoe UI"; RED = "D62027"; REDHOT = "F40009"; SLATE = "1E1E24"; SLATE2 = "2B2B33"
SILVER = "8A9EA7"; SILVER_TX = "1E1E24"; INK = "1F2937"; GRID = "E5E7EB"; ZEBRA = "F9FAFB"
DASH_HDR = RED; DASH_BG = "F7F8FA"; S4_BAND = "FBF1E6"
TH = {"master": SLATE, "trans": SLATE, "config": SLATE, "cross": SLATE}
TAB = {"master": SLATE, "trans": RED, "config": SLATE, "cross": SILVER}
DASH = "מסך ניווט מרכזי"; COCKPIT = "Cockpit מעקב מיגרציה"; ER = "ER - Join Map"
SEARCH_CELL = "C25"; IDX0 = 14
thin = Side(style="thin", color=GRID); BORDER = Border(thin, thin, thin, thin)

COLS = [("מס' (#)", 5), ("טבלה (Table)", 12), ("שדה טכני (Field)", 16), ("סוג נתונים (Type)", 12),
        ("אורך (Len)", 8), ("מפתח (Key)", 9), ("English Term", 24), ("תרגום עברי (Hebrew)", 26),
        ("JOIN ON (SQL / CDS)", 40), ("הסבר פונקציונלי בעברית", 46), ("פער / הערת S/4HANA", 30),
        ("SAP Help", 16)]
NCOL = len(COLS); S4_COL = 11
KEYFILL = {"PK": ("FCE4E6", "B01722"), "FK": ("ECEFF1", "37474F"), "PK/FK": ("FBE3E4", "D62027"), "-": ("F3F4F6", "9AA3AF")}

MACRO_SRC = (
'Attribute VB_Name = "modPPPI"\nOption Explicit\n\n'
"' SAP PP-PI - global search across all sheets (skips the dashboard).\n"
'Sub PPPI_GlobalSearch()\n'
'    Dim q As String, ws As Worksheet, f As Range\n'
'    q = CStr(ThisWorkbook.Sheets(1).Range("' + SEARCH_CELL + '").Value)\n'
'    If Trim(q) = "" Then q = InputBox("הזן שדה / טבלה / טרנזקציה / מונח:", "PP-PI Global Search")\n'
'    If Trim(q) = "" Then Exit Sub\n'
'    For Each ws In ThisWorkbook.Worksheets\n'
'        If ws.Index > 1 Then\n'
'            Set f = ws.Cells.Find(What:=q, LookIn:=xlValues, LookAt:=xlPart, MatchCase:=False)\n'
'            If Not f Is Nothing Then\n'
'                ws.Activate: Application.Goto f, True\n'
'                f.Interior.Color = RGB(255, 242, 0)\n'
'                MsgBox "נמצא בלשונית: " & ws.Name & " | שורה: " & f.Row, vbInformation, "חיפוש"\n'
'                Exit Sub\n'
'            End If\n'
'        End If\n'
'    Next ws\n'
'    MsgBox "לא נמצא: " & q, vbExclamation\n'
'End Sub\n\n'
'Sub BackToDashboard()\n'
'    ThisWorkbook.Sheets(1).Activate\n'
'    ThisWorkbook.Sheets(1).Range("A1").Select\n'
'End Sub\n'
)

def fiori_url(fid):
    return f"https://fioriappslibrary.hana.ondemand.com/sap/fix/externalViewer/#/detail/Apps('{fid}')" if (fid and fid[0] == "F") else None

def merged_row(ws, row, segs, h=18):
    # segs: list of (c0,c1,val,bold,align,color,mono,link)
    for (c0, c1, val, bold, al, color, mono, link) in segs:
        ws.merge_cells(start_row=row, start_column=c0, end_row=row, end_column=c1)
        cell = ws.cell(row, c0, val)
        cell.font = Font(name=("Consolas" if mono else FONT), size=9, bold=bold,
                         color=("0563C1" if link else color), underline=("single" if link else None))
        cell.alignment = Alignment(horizontal=al, vertical="center", wrap_text=True); cell.border = BORDER
        if link: cell.hyperlink = link
    ws.row_dimensions[row].height = h

def render_ops(ws, ops, row, sname):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOL)
    c = ws.cell(row, 1, "  🔧 שכבה תפעולית (Operational Layer)  -  T-codes ◄► Fiori  |  ממשקים BAPI/IDoc/RFC  |  תוכניות רקע ודוחות")
    c.font = f(11, True, "FFFFFF"); c.fill = PatternFill("solid", fgColor=RED); c.alignment = Alignment("right", "center")
    ws.row_dimensions[row].height = 22; row += 1
    def subhdr(text):
        nonlocal row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOL)
        cc = ws.cell(row, 1, "    " + text); cc.font = f(10, True, "FFFFFF"); cc.fill = PatternFill("solid", fgColor=SLATE2)
        cc.alignment = Alignment("right", "center"); ws.row_dimensions[row].height = 18; row += 1
    # ---- T-code <-> Fiori mapping ----
    subhdr("מיפוי T-code ◄► Fiori App (ECC ◄► S/4HANA)")
    merged_row(ws, row, [(1, 3, "ECC T-code", True, "center", "FFFFFF", False, None),
                         (4, 8, "S/4HANA Fiori App", True, "center", "FFFFFF", False, None),
                         (9, 12, "Fiori ID", True, "center", "FFFFFF", False, None)])
    for s in (ws.cell(row, 1), ws.cell(row, 4), ws.cell(row, 9)): s.fill = PatternFill("solid", fgColor=SILVER)
    for s in (ws.cell(row, 1), ws.cell(row, 4), ws.cell(row, 9)): s.font = f(9, True, SILVER_TX)
    row += 1
    for ecc, app, fid in ops["tcodes"]:
        lk = fiori_url(fid)
        merged_row(ws, row, [(1, 3, ecc, True, "left", RED, True, None),
                             (4, 8, app, False, "right", INK, False, None),
                             (9, 12, fid, True, "center", SLATE, False, lk)])
        index_rows.append(("טרנזקציה", ecc.split()[0], app, fid, sname, f"A{row}"))
        if fid and fid != "-": index_rows.append(("Fiori", fid, app, "", sname, f"A{row}"))
        row += 1
    # ---- Interfaces ----
    subhdr("ממשקים ובלוקים פונקציונליים: BAPIs / IDocs / RFCs (Extraction & Messaging)")
    merged_row(ws, row, [(1, 2, "Type", True, "center", "FFFFFF", False, None),
                         (3, 6, "Name", True, "center", "FFFFFF", False, None),
                         (7, 12, "תיאור / Description", True, "center", "FFFFFF", False, None)])
    for cc in (1, 3, 7): ws.cell(row, cc).fill = PatternFill("solid", fgColor=SILVER); ws.cell(row, cc).font = f(9, True, SILVER_TX)
    row += 1
    for kind, name, desc in ops["interfaces"]:
        merged_row(ws, row, [(1, 2, kind, True, "center", RED, False, None),
                             (3, 6, name, True, "left", SLATE, True, None),
                             (7, 12, desc, False, "right", INK, False, None)])
        index_rows.append((kind, name, desc, "", sname, f"A{row}")); row += 1
    # ---- Background programs ----
    subhdr("תוכניות רקע ודוחות סטנדרטיים (Background Programs / Reports)")
    merged_row(ws, row, [(1, 4, "Program / Report", True, "center", "FFFFFF", False, None),
                         (5, 12, "תיאור ומטרה / Usage & Trigger", True, "center", "FFFFFF", False, None)])
    for cc in (1, 5): ws.cell(row, cc).fill = PatternFill("solid", fgColor=SILVER); ws.cell(row, cc).font = f(9, True, SILVER_TX)
    row += 1
    for name, desc in ops["programs"]:
        merged_row(ws, row, [(1, 4, name, True, "left", SLATE, True, None),
                             (5, 12, desc, False, "right", INK, False, None)])
        index_rows.append(("תוכנית", name, desc, "", sname, f"A{row}")); row += 1
    return row + 1

wb = Workbook(); wb.remove(wb.active)
def safe(n):
    for ch in ':\\/?*[]': n = n.replace(ch, "")
    return n[:31]
sheet_names = [safe(t["title"]) for t in TOPICS]
dash = wb.create_sheet(DASH)
index_rows = []; tables_meta = []; joins = []

def f(sz=10, b=False, c=INK, u=None): return Font(name=FONT, size=sz, bold=b, color=c, underline=u)
def back_btn(ws, ncol, title):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    bb = ws.cell(1, 1, "◄ חזרה למסך הראשי"); bb.hyperlink = f"#'{DASH}'!A1"
    bb.font = f(10, True, "FFFFFF"); bb.fill = PatternFill("solid", fgColor=SLATE)
    bb.alignment = Alignment(horizontal="center", vertical="center")
    rs = Side(style="thin", color=REDHOT); rb = Border(rs, rs, rs, rs); bb.border = rb; ws.cell(1, 2).border = rb
    ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=ncol)
    t = ws.cell(1, 3, title); t.font = f(15, True, "FFFFFF"); t.fill = PatternFill("solid", fgColor=RED)
    t.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[1].height = 30

# ---- data sheets ----
for topic, sname, ops in zip(TOPICS, sheet_names, OPS):
    ws = wb.create_sheet(sname); ws.sheet_view.rightToLeft = True; ws.sheet_view.showGridLines = True
    ws.sheet_properties.tabColor = TAB[topic["theme"]]
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    back_btn(ws, NCOL, topic["title"] + "   |   ECC 6.0 ➜ S/4HANA  ·  CBC NEO")
    for j, (lbl, w) in enumerate(COLS, start=1):
        is_s4 = j >= S4_COL
        c = ws.cell(2, j, lbl); c.font = f(11, True, SILVER_TX if is_s4 else "FFFFFF")
        c.fill = PatternFill("solid", fgColor=SILVER if is_s4 else SLATE2)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = BORDER
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[2].height = 28; ws.freeze_panes = "A3"
    row = render_ops(ws, ops, 3, sname); seq = 0
    for tb in topic["tables"]:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOL)
        hc = ws.cell(row, 1, f"  ◆ טבלה {tb['name']}  -  {tb['he']}  ({tb['en']})   |   T-codes: {tb['tcodes']}")
        hc.font = f(11, True, "FFFFFF"); hc.fill = PatternFill("solid", fgColor=SLATE)
        hc.alignment = Alignment(horizontal="right", vertical="center"); ws.row_dimensions[row].height = 22
        index_rows.append(("טבלה", tb["name"], tb["he"], tb["en"], sname, f"A{row}"))
        tables_meta.append((tb["name"], tb["he"], topic["title"], sname, f"A{row}"))
        if tb["join"].startswith("FROM"):
            joins.append((tb["name"], tb["join"], tb["he"], sname, f"A{row}"))
        row += 1
        fstart = row
        for (tech, en, he, dt, ln, key) in tb["fields"]:
            seq += 1; band = ZEBRA if seq % 2 == 0 else None
            def sc(col, val, *, bold=False, color=INK, h="right", fill=band):
                c = ws.cell(row, col, val); c.font = f(10, bold, color)
                c.alignment = Alignment(horizontal=h, vertical="center", wrap_text=True); c.border = BORDER
                if fill: c.fill = PatternFill("solid", fgColor=fill)
                return c
            sc(1, seq, h="center", color="6B7280")
            sc(2, "")
            sc(3, tech, bold=True, color=RED, h="left")
            sc(4, dt, h="center"); sc(5, ln, h="center")
            kf = KEYFILL[key]; kc = sc(6, key, bold=(key != "-"), color=kf[1], h="center")
            kc.fill = PatternFill("solid", fgColor=kf[0])
            sc(7, en, h="left"); sc(8, he, h="right")
            for col in (9, 10, 11, 12): sc(col, "", fill=(S4_BAND if (col >= S4_COL and band) else band))
            ws.row_dimensions[row].height = 26
            index_rows.append(("שדה", tech, he, en, sname, f"C{row}"))
            row += 1
        fend = row - 1
        def vm(col, val, *, bold=False, color=INK, mono=False, link=None, h="right"):
            c = ws.cell(fstart, col, val)
            c.font = Font(name=("Consolas" if mono else FONT), size=9, bold=bold, color=color,
                          underline=("single" if link else None))
            c.alignment = Alignment(horizontal=h, vertical="top", wrap_text=True); c.border = BORDER
            if link: c.hyperlink = link
            if fend > fstart: ws.merge_cells(start_row=fstart, start_column=col, end_row=fend, end_column=col)
        vm(2, tb["name"], bold=True, color=RED, h="center")
        vm(9, tb["join"], mono=True, color="0B5394", h="left")
        vm(10, tb["guide"])
        vm(11, tb["s4"], color="9A5A23")
        vm(12, tb["help"][1], color="0563C1", link=tb["help"][0], h="center")
        for code in [x.strip() for x in tb["tcodes"].replace(";", ",").replace("/", ",").split(",") if x.strip()]:
            index_rows.append(("טרנזקציה", code, tb["he"], tb["en"], sname, f"A{fstart}"))
        row += 1
    # print branding
    ws.oddHeader.left.text = '&"Segoe UI,Bold"&13&KD62027CBC ISRAEL'
    ws.oddHeader.center.text = '&"Segoe UI,Bold"&11&K1E1E24Project NEO  |  SAP PP-PI  ECC 6.0 -> S/4HANA'
    ws.oddFooter.center.text = '&"Segoe UI"&8&K6B7280Page &P of &N'
    ws.page_setup.orientation = "landscape"; ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
    ws.print_options.gridLines = False; ws.print_title_rows = "1:2"
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.7, bottom=0.6, header=0.3, footer=0.3)

# ---- Cockpit ----
STAT = ["Not started", "In analysis", "In conversion", "Tested", "Done"]
SF = {"Not started": "FCE4E6", "In analysis": "ECEFF1", "In conversion": "E1E6EA", "Tested": "DCE6EC", "Done": "DCEFE0"}
STX = {"Not started": "B01722", "In analysis": "475569", "In conversion": "37474F", "Tested": "2A4A57", "Done": "1E5A44"}
DS = 8
co = wb.create_sheet(COCKPIT); co.sheet_view.rightToLeft = True; co.sheet_properties.tabColor = RED
back_btn(co, 6, "Cockpit מעקב מיגרציה  -  PP-PI Migration Tracking (Project NEO)")
cc = [("מס' (#)", 6), ("מודול / נושא", 30), ("טבלה", 12), ("סטטוס מעבר (Status)", 18), ("קישור", 12), ("תיאור", 34)]
for j, (lbl, w) in enumerate(cc, start=1):
    co.column_dimensions[get_column_letter(j)].width = w
co_last = DS + len(tables_meta) - 1; DR = f"$D${DS}:$D${co_last}"; CRr = f"$C${DS}:$C${co_last}"
co.merge_cells("B2:F2")
kb = co.cell(2, 2, "KPI & Summary  -  סיכום חי (COUNTIF; בסיס מוכן לגרף נייטיב ב-2 קליקים)")
kb.font = f(12, True, "FFFFFF"); kb.fill = PatternFill("solid", fgColor=RED); kb.alignment = Alignment(horizontal="right")
tiles = [("סה\"כ", f"=COUNTA({CRr})", SLATE, "ECEFF1"), ("Done", f'=COUNTIF({DR},"Done")', "1E5A44", "DCEFE0"),
         ("Tested", f'=COUNTIF({DR},"Tested")', "2A4A57", "DCE6EC"),
         ("In Progress", f'=COUNTIF({DR},"In analysis")+COUNTIF({DR},"In conversion")', "37474F", "E1E6EA"),
         ("Open", f'=COUNTIF({DR},"Not started")', "B01722", "FCE4E6"),
         ("% Done", f'=IFERROR(COUNTIF({DR},"Done")/COUNTA({CRr}),0)', RED, "FBE3E4")]
for k, (lbl, fm, tx, fl) in enumerate(tiles):
    n = co.cell(3, 2 + k, fm); n.font = f(18, True, tx); n.alignment = Alignment("center", "center")
    n.fill = PatternFill("solid", fgColor=fl); n.border = BORDER
    if "%" in lbl: n.number_format = "0%"
    l = co.cell(4, 2 + k, lbl); l.font = f(9, True, tx); l.alignment = Alignment("center", "center")
    l.fill = PatternFill("solid", fgColor=fl); l.border = BORDER
co.row_dimensions[3].height = 30
# status/count source for native charts (K:L)
co.cell(2, 11, "Status").font = f(10, True, "FFFFFF"); co.cell(2, 12, "Count").font = f(10, True, "FFFFFF")
for cn in (11, 12): co.cell(2, cn).fill = PatternFill("solid", fgColor=SLATE); co.cell(2, cn).border = BORDER
for k, s in enumerate(STAT):
    a = co.cell(3 + k, 11, s); a.font = f(10, True, STX[s]); a.fill = PatternFill("solid", fgColor=SF[s]); a.border = BORDER
    b = co.cell(3 + k, 12, f'=COUNTIF({DR},$K{3+k})'); b.font = f(11, True, INK); b.alignment = Alignment("center"); b.border = BORDER
co.column_dimensions["K"].width = 15; co.column_dimensions["L"].width = 10
for j, (lbl, w) in enumerate(cc, start=1):
    c = co.cell(7, j, lbl); c.font = f(11, True, "FFFFFF"); c.fill = PatternFill("solid", fgColor=SLATE)
    c.alignment = Alignment("center", "center", wrap_text=True); c.border = BORDER
co.row_dimensions[7].height = 26; co.freeze_panes = "A8"
rr = DS
for i, (tn, the, ttitle, tsheet, tcell) in enumerate(tables_meta, start=1):
    band = ZEBRA if i % 2 == 0 else None
    vals = [i, ttitle, tn, "Not started", None, the]
    for j, v in enumerate(vals, start=1):
        c = co.cell(rr, j, v if v is not None else "")
        c.font = f(10, j == 3, RED if j == 3 else INK)
        c.alignment = Alignment(("center" if j in (1, 4, 5) else ("left" if j == 3 else "right")), "center", wrap_text=True)
        c.border = BORDER
        if band: c.fill = PatternFill("solid", fgColor=band)
    co.cell(rr, 4).fill = PatternFill("solid", fgColor=SF["Not started"]); co.cell(rr, 4).font = f(10, True, STX["Not started"])
    lk = co.cell(rr, 5); lk.value = f'=HYPERLINK("#\'"&"{tsheet}"&"\'!{tcell}","➜ פתח")'; lk.font = f(9, False, "0563C1", u="single")
    rr += 1
dv = DataValidation(type="list", formula1='"%s"' % ",".join(STAT), allow_blank=True)
co.add_data_validation(dv); dv.add(f"D{DS}:D{co_last}")
for s, fl in SF.items():
    co.conditional_formatting.add(f"D{DS}:D{co_last}",
        CellIsRule(operator="equal", formula=[f'"{s}"'], fill=PatternFill("solid", fgColor=fl), font=f(10, True, STX[s])))

# ---- ER - Join Map ----
er = wb.create_sheet(ER); er.sheet_view.rightToLeft = True; er.sheet_properties.tabColor = RED
back_btn(er, 4, "ER - Join Map  -  JOIN ON (SQL / CDS Views)  ·  CBC NEO")
ec = [("מס' (#)", 5), ("טבלה (Child)", 14), ("JOIN ON (SQL / CDS Views Syntax)", 70), ("תיאור", 40)]
for j, (lbl, w) in enumerate(ec, start=1):
    c = er.cell(2, j, lbl); c.font = f(11, True, "FFFFFF"); c.fill = PatternFill("solid", fgColor=SLATE)
    c.alignment = Alignment("center", "center", wrap_text=True); c.border = BORDER
    er.column_dimensions[get_column_letter(j)].width = w
er.row_dimensions[2].height = 26; er.freeze_panes = "A3"
rr = 3
for i, (tn, jn, he, sn, cell) in enumerate(joins, start=1):
    band = ZEBRA if i % 2 == 0 else None
    for j, (v, mono, col, h) in enumerate([(i, False, "6B7280", "center"), (tn, False, RED, "center"),
                                           (jn, True, "0B5394", "left"), (he, False, INK, "right")], start=1):
        c = er.cell(rr, j, v); c.font = Font(name=("Consolas" if mono else FONT), size=10, bold=(j in (2, 3)), color=col)
        c.alignment = Alignment(h, "center", wrap_text=True); c.border = BORDER
        if band: c.fill = PatternFill("solid", fgColor=band)
    er.row_dimensions[rr].height = 24
    index_rows.append(("Join", jn, he, "", ER, f"A{rr}")); rr += 1

# ---- Dashboard ----
df = pd.DataFrame(index_rows, columns=["t", "code", "he", "en", "sheet", "cell"]); N = len(df)
dash.sheet_view.rightToLeft = True; dash.sheet_view.showGridLines = False; dash.sheet_properties.tabColor = RED
for col, w in zip("ABCDEFGHIJKLM", [3, 16, 16, 16, 16, 16, 4, 16, 16, 3, 3, 3, 3]): dash.column_dimensions[col].width = w
for r0 in range(1, 60):
    for c0 in range(1, 14): dash.cell(r0, c0).fill = PatternFill("solid", fgColor=DASH_BG)
dash.merge_cells("B2:H2"); d1 = dash.cell(2, 2, "CBC ISRAEL  |  SAP PP-PI  -  מסך ניווט מרכזי")
d1.font = f(18, True, RED); d1.alignment = Alignment(horizontal="right")
dash.merge_cells("B3:H3"); d2 = dash.cell(3, 2, "Project NEO  -  Production Planning for Process Industries  |  ECC 6.0 ➜ S/4HANA")
d2.font = f(10, True, "64748B"); d2.alignment = Alignment(horizontal="right")
dash.merge_cells("J2:M3"); neo = dash.cell(2, 10, "★ CBC\nProject NEO"); neo.font = f(12, True, "FFFFFF")
neo.fill = PatternFill("solid", fgColor=RED); neo.alignment = Alignment("center", "center", wrap_text=True)
CRC = f"'{COCKPIT}'"
TOT = f"COUNTA({CRC}!$C${DS}:$C${co_last})"; DON = f'COUNTIF({CRC}!$D${DS}:$D${co_last},"Done")'
TES = f'COUNTIF({CRC}!$D${DS}:$D${co_last},"Tested")'; OPN = f'COUNTIF({CRC}!$D${DS}:$D${co_last},"Not started")'
PRG = f'(COUNTIF({CRC}!$D${DS}:$D${co_last},"In analysis")+COUNTIF({CRC}!$D${DS}:$D${co_last},"In conversion"))'
dash.merge_cells("B5:H5"); kh = dash.cell(5, 2, "סיכום התקדמות מיגרציה (KPI - חי מה-Cockpit)")
kh.font = f(12, True, "FFFFFF"); kh.fill = PatternFill("solid", fgColor=RED); kh.alignment = Alignment(horizontal="right")
kp = [("סה\"כ אובייקטים", f"={TOT}", SLATE, "ECEFF1"), ("הושלם", f"={DON}", "1E5A44", "DCEFE0"),
      ("נבדק", f"={TES}", "2A4A57", "DCE6EC"), ("בתהליך", f"={PRG}", "37474F", "E1E6EA"),
      ("פתוח", f"={OPN}", "B01722", "FCE4E6"), ("% השלמה", f"=IFERROR(({DON})/({TOT}),0)", RED, "FBE3E4")]
for k, (lbl, fm, tx, fl) in enumerate(kp):
    n = dash.cell(6, 2 + k, fm); n.font = f(18, True, tx); n.alignment = Alignment("center", "center")
    n.fill = PatternFill("solid", fgColor=fl); n.border = BORDER
    if "%" in lbl: n.number_format = "0%"
    l = dash.cell(7, 2 + k, lbl); l.font = f(9, True, tx); l.alignment = Alignment("center", "center", wrap_text=True)
    l.fill = PatternFill("solid", fgColor=fl); l.border = BORDER
dash.row_dimensions[6].height = 30
dash.merge_cells("B9:H9"); nh = dash.cell(9, 2, "ניווט מהיר לגיליונות"); nh.font = f(12, True, "FFFFFF")
nh.fill = PatternFill("solid", fgColor=SLATE); nh.alignment = Alignment(horizontal="right")
nav = [(t["title"], sn, TAB[t["theme"]]) for t, sn in zip(TOPICS, sheet_names)]
nav += [("◆ " + COCKPIT, COCKPIT, RED), ("⇄ " + ER, ER, RED)]
cardc = [2, 4, 6, 8]; r = 10
for i, (title, sn, hdr) in enumerate(nav):
    if i % 4 == 0 and i: r += 3
    c0 = cardc[i % 4]; dash.merge_cells(start_row=r, start_column=c0, end_row=r + 1, end_column=c0)
    cell = dash.cell(r, c0, title); cell.hyperlink = f"#'{sn}'!A1"
    cell.font = f(10, True, SLATE if hdr == SILVER else "FFFFFF"); cell.fill = PatternFill("solid", fgColor=hdr)
    cell.alignment = Alignment("center", "center", wrap_text=True)
    for rr2 in (r, r + 1): dash.cell(rr2, c0).border = BORDER
r += 3
dash.merge_cells("B24:F24"); shh = dash.cell(24, 2, "🔍 חיפוש גלובלי (טבלה / שדה / טרנזקציה - עברית/אנגלית)")
shh.font = f(12, True, "FFFFFF"); shh.fill = PatternFill("solid", fgColor=SLATE); shh.alignment = Alignment(horizontal="right")
lb = dash.cell(25, 2, "הקלד כאן:"); lb.font = f(11, True); lb.alignment = Alignment(horizontal="right"); lb.fill = PatternFill("solid", fgColor=DASH_BG)
dash.merge_cells("C25:F25"); sc2 = dash.cell(25, 3, ""); sc2.fill = PatternFill("solid", fgColor="FCE4E6")
md = Side(style="medium", color=REDHOT); sc2.border = Border(md, md, md, md); dash.row_dimensions[25].height = 26
for j, h in enumerate(["קישור", "קוד", "עברית", "English", "גיליון", "תא"]):
    c = dash.cell(27, 2 + j, h); c.font = f(10, True, "FFFFFF"); c.fill = PatternFill("solid", fgColor=SLATE)
    c.alignment = Alignment(horizontal="center"); c.border = BORDER
C = lambda k: get_column_letter(IDX0 + k); rng = lambda k: f"${C(k)}$2:${C(k)}${N+1}"
formula = (f'=IF({SEARCH_CELL}="","הקלד מונח וקבל תוצאות מיידיות...",'
           f'IFERROR(_xlfn._xlws.FILTER(CHOOSE({{1,2,3,4,5,6}},{rng(6)},{rng(1)},{rng(2)},{rng(3)},{rng(4)},{rng(5)}),'
           f'(ISNUMBER(SEARCH({SEARCH_CELL},{rng(1)})))+(ISNUMBER(SEARCH({SEARCH_CELL},{rng(2)})))+'
           f'(ISNUMBER(SEARCH({SEARCH_CELL},{rng(3)})))+(ISNUMBER(SEARCH({SEARCH_CELL},{rng(0)})))>0),"לא נמצאו תוצאות"))')
fc = dash.cell(28, 2, formula); fc.font = f(10); fc.alignment = Alignment(horizontal="right")
# hidden index block (N..)
hdrs = ["Type", "Code", "He", "En", "Sheet", "Cell", "Link"]
for k, ttl in enumerate(hdrs):
    c = dash.cell(1, IDX0 + k, ttl); c.font = f(8, True, "FFFFFF"); c.fill = PatternFill("solid", fgColor=SLATE)
for ri, (typ, code, he, en, sn, cell) in enumerate(index_rows, start=2):
    dash.cell(ri, IDX0 + 0, typ); dash.cell(ri, IDX0 + 1, code).font = f(8, True, SLATE)
    dash.cell(ri, IDX0 + 2, he); dash.cell(ri, IDX0 + 3, en); dash.cell(ri, IDX0 + 4, sn); dash.cell(ri, IDX0 + 5, cell)
    lk = dash.cell(ri, IDX0 + 6); lk.value = f'=HYPERLINK("#\'"&{C(4)}{ri}&"\'!"&{C(5)}{ri},"➜ פתח")'; lk.font = f(8, False, "0563C1", u="single")
for k in range(7): dash.column_dimensions[get_column_letter(IDX0 + k)].hidden = True
dash.merge_cells("B56:H59"); note = dash.cell(56, 2,
   "פורטל זה הוא .xlsx נקי (ללא מאקרו/גרפים) - בטוח למייל ארגוני. עדכן סטטוס ב-Cockpit וה-KPI יתעדכן מיד. "
   "להוספת כפתור חיפוש-מאקרו: הדבק pppi_modPM.bas דרך Alt+F11 ושמור כ-.xlsm. גם זמין כפורטל ווב: index_pppi.html.")
note.font = f(9, c="475569"); note.alignment = Alignment("right", "top", wrap_text=True)
note.fill = PatternFill("solid", fgColor="F1F3F5"); note.border = BORDER; dash.freeze_panes = "A2"
dash.print_area = "A1:M60"; dash.oddFooter.center.text = '&"Segoe UI"&8&K6B7280CBC NEO  -  PP-PI Migration  -  Page &P of &N'

# universal Segoe UI / Consolas pass
for shx in wb.worksheets:
    for rowc in shx.iter_rows():
        for cell in rowc:
            ft = cell.font
            if ft is not None and ft.name not in (FONT, "Consolas"):
                cell.font = Font(name=FONT, size=ft.size, bold=ft.bold, italic=ft.italic, color=ft.color, underline=ft.underline)

OUT = "SAP_PPPI_ECC6_to_S4_Migration.xlsx"
wb.save(OUT)
with open("pppi_modPM.bas", "w", encoding="utf-8") as fh: fh.write(MACRO_SRC)
nt = sum(len(t["tables"]) for t in TOPICS); nf = sum(len(tb["fields"]) for t in TOPICS for tb in t["tables"])
print(f"OK -> {OUT}  | sheets:{len(wb.sheetnames)} tables:{nt} fields:{nf} joins:{len(joins)} index:{N}")
print("   VBA -> pppi_modPM.bas")
