# -*- coding: utf-8 -*-
"""
SAP PM ECC 6.0  ->  S/4HANA Migration Bridge - interactive enterprise workbook.

Features
--------
* Sheet 0 "מסך ניווט מרכזי" (Main Dashboard): styled card buttons (hyperlinks) to every
  topic sheet + a no-macro GLOBAL SEARCH (FILTER over an INDEX_DB sheet, returns clickable
  HYPERLINK rows). VBA alternative is provided separately in the chat answer.
* Every topic sheet: "חזרה למסך ראשי" jump button in cell A1.
* Expanded RTL grid - every Core Table expanded into its own section; each KEY FIELD on its
  own row with Technical name / English / Hebrew. Table-level T-codes, Functions/BAPIs (all),
  Programs (all) are vertically merged across the table's field block.
* S/4HANA Migration & Fiori matrix appended on the right of every table:
  Simplification status / Replacement table-field / Replacement T-code-program / Fiori app+ID.

Scope note (repo CLAUDE.md rule 3): indexed corpus is S/4HANA only; this is explicit opt-in
ECC 6.0 model knowledge with an S/4 mapping. Verify version-specific items (table/field
availability, Fiori App IDs) in SAP Help Portal / SAP Fiori Apps Reference Library.

Run:  python3 scripts/gen_pm_workbook_s4.py
Out:  SAP_PM_ECC6_to_S4_Migration.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.utils import get_column_letter

# ---- palette -------------------------------------------------------------
TH = {
 "master": {"h": "1F3864", "s": "2E5496", "b": "DCE3F2", "t": "1F3864"},  # Dark Blue
 "trans":  {"h": "1E5E34", "s": "2E8B53", "b": "DCEFE0", "t": "1E5E34"},  # Deep Green
 "config": {"h": "7B1E2B", "s": "A53444", "b": "F0DCE0", "t": "7B1E2B"},  # Maroon
 "cross":  {"h": "C55A11", "s": "E08020", "b": "FBE6D5", "t": "C55A11"},  # Orange
}
S4_HDR = "C55A11"   # Orange band for the S/4HANA section
S4_BAND = "FCEFE1"
DASH = "מסך ניווט מרכזי"

WHITE = "FFFFFF"
thin = Side(style="thin", color="C9C9C9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# 14 columns (RTL: col 1 = rightmost). (label, width)
COLS = [
 ("מס' (#)", 5),
 ("טרנזקציות ECC\n(Main & Secondary)", 18),
 ("טבלת בסיס\n(Core Table)", 13),
 ("שדה טכני\n(Field)", 14),
 ("Field (English)", 26),
 ("שדה (עברית)", 24),
 ("פונקציות / BAPIs", 26),
 ("תיאור הפונקציה", 30),
 ("תוכניות / דוחות", 16),
 ("תיאור התוכנית", 26),
 ("סטטוס S/4HANA", 20),
 ("טבלה/שדה חלופי ב-S/4", 22),
 ("טרנז'/תוכנית מחליפה ב-S/4", 22),
 ("אפליקציית Fiori (שם + ID)", 26),
]
NCOL = len(COLS)
S4_COL_START = 11  # columns 11..14 = S/4 matrix

# ============================================================================
# CONTENT  -  TOPICS -> tables -> fields + functions + programs + S/4 matrix
# table dict keys: name, he, en, tcodes, fields[(tech,en,he)], funcs[(n,he)],
#                  progs[(n,he)], s4_status, s4_repl, s4_tcode, fiori
# ============================================================================
TOPICS = [
# ---------------- 1. Org structure ----------------------------------------
{"title": "1. מבנה ארגוני ותשתית", "theme": "master", "tables": [
 {"name": "IFLOT", "he": "רשומת אב של מיקום פונקציונלי", "en": "Functional location master record",
  "tcodes": "IL01/IL02/IL03; IH01, IH06",
  "fields": [("TPLNR","Functional location number","מספר מיקום פונקציונלי"),
             ("STRNO","Formatted / structured number","מספר מבנה מעוצב"),
             ("FLTYP","Functional location category","קטגוריית מיקום"),
             ("TPLKZ","Structure indicator (edit mask)","מחוון מבנה / מסיכת קידוד"),
             ("TPLMA","Superior functional location","מיקום אב בהיררכיה"),
             ("OBJNR","Object number (status)","מפתח אובייקט לסטטוס")],
  "funcs": [("BAPI_FUNCLOC_CREATE","יצירת מיקום פונקציונלי דרך ממשק תקני"),
            ("BAPI_FUNCLOC_CHANGE","שינוי מיקום פונקציונלי"),
            ("BAPI_FUNCLOC_GETDETAIL","שליפת פרטי מיקום"),
            ("ISU_FUNCLOC_GETLIST","שליפת רשימת מיקומים לפי בחירה")],
  "progs": [("RIIFLO20","רשימת מיקומים פונקציונליים (IH06)"),
            ("RIFLET00","תחזוקת קטגוריות ומסיכות עריכה")],
  "s4_status": "ללא שינוי במודל הנתונים (תואם)", "s4_repl": "IFLOT (זהה)",
  "s4_tcode": "IL01-03 נתמכים; אסטרטגי: Fiori", "fiori": "Manage Technical Objects (F2079)"},
 {"name": "ILOA", "he": "נתוני מיקום וחיוב משותפים", "en": "PM object location & account assignment",
  "tcodes": "IL02; IE02 (ירושה)",
  "fields": [("ILOAN","Location & account assignment key","מפתח נתוני מיקום/חיוב"),
             ("KOSTL","Cost center","מרכז עלות"),
             ("SWERK","Maintenance plant","מפעל אחזקה"),
             ("INGRP","Planner group","קבוצת תכנון"),
             ("GEWRK","Main work center","מרכז עבודה ראשי"),
             ("TPLNR","Functional location","מיקום פונקציונלי")],
  "funcs": [("ILOA_READ","קריאת נתוני מיקום/חיוב"),
            ("ILOA_UPDATE","עדכון נתוני מיקום/חיוב"),
            ("ILOA_INHERIT_FROM_FUNCLOC","ירושת נתונים מהמיקום לציוד")],
  "progs": [("RILOAA00","התאמת נתוני מיקום/חיוב באובייקטים")],
  "s4_status": "ללא שינוי (תואם)", "s4_repl": "ILOA (זהה)",
  "s4_tcode": "ללא; מנוהל דרך אובייקט טכני", "fiori": "Manage Technical Objects (F2079)"},
 {"name": "CRHD", "he": "כותרת מרכז עבודה", "en": "Work center header",
  "tcodes": "IR01/IR02/IR03; CR05, CR06",
  "fields": [("OBJID","Internal work center object ID","מזהה אובייקט פנימי"),
             ("ARBPL","Work center code","קוד מרכז עבודה"),
             ("OBJTY","Object type ('A'=work center)","סוג אובייקט"),
             ("WERKS","Plant","מפעל"),
             ("VERWE","Work center usage","שימוש מרכז עבודה"),
             ("VERAN","Person responsible","אחראי")],
  "funcs": [("CR_WORKCENTER_READ","קריאת נתוני מרכז עבודה"),
            ("CRAP_WORKCENTER_GET_DETAIL","שליפת פרטי מרכז עבודה")],
  "progs": [("RCRACO00","דוח מרכזי עבודה וקיבולות")],
  "s4_status": "ללא שינוי (תואם)", "s4_repl": "CRHD (זהה)",
  "s4_tcode": "IR01-03 / CR0*; Fiori", "fiori": "Manage Work Centers (אמת ID)"},
 {"name": "CRTX", "he": "טקסטים של מרכז עבודה", "en": "Work center texts",
  "tcodes": "IR02; IR03",
  "fields": [("OBJID","Work center object ID","מזהה אובייקט"),
             ("SPRAS","Language key","שפה"),
             ("KTEXT","Work center description","תיאור מרכז עבודה")],
  "funcs": [("CR_TEXT_READ","קריאת טקסט מרכז עבודה")],
  "progs": [("RCRACO00","דוח מרכזי עבודה")],
  "s4_status": "ללא שינוי (תואם)", "s4_repl": "CRTX (זהה)",
  "s4_tcode": "IR0*; Fiori", "fiori": "Manage Work Centers (אמת ID)"},
 {"name": "T370T", "he": "טבלת טקסט להגדרות מבנה/קטגוריות PM (Customizing)",
  "en": "PM structure/category customizing text table",
  "tcodes": "SPRO; OIA1, OIMR",
  "fields": [("SPRAS","Language key","שפה"),
             ("FLTYP","Functional location category","קטגוריית מיקום"),
             ("FLTYPT","Category description","תיאור קטגוריה")],
  "funcs": [("STRUCTURE_INDICATOR_READ","קריאת מחוון מבנה והגדרותיו")],
  "progs": [("RIFLET00","תחזוקת קטגוריות ומסיכות עריכה")],
  "s4_status": "ללא שינוי (Customizing תואם)", "s4_repl": "T370T (זהה)",
  "s4_tcode": "SPRO (זהה)", "fiori": "אין Fiori ייעודי (Customizing)"},
]},
# ---------------- 2. Equipment --------------------------------------------
{"title": "2. ציוד ונתוני מאסטר", "theme": "master", "tables": [
 {"name": "EQUI", "he": "רשומת אב של ציוד", "en": "Equipment master record",
  "tcodes": "IE01/IE02/IE03; IH08, IE05",
  "fields": [("EQUNR","Equipment number","מספר ציוד"),
             ("EQTYP","Equipment category","קטגוריית ציוד"),
             ("EQART","Technical object type","סוג טכני"),
             ("HERST","Manufacturer","יצרן"),
             ("TYPBZ","Manufacturer model number","דגם היצרן"),
             ("OBJNR","Object number (status)","מפתח אובייקט לסטטוס")],
  "funcs": [("BAPI_EQUI_CREATE","יצירת ציוד דרך ממשק תקני"),
            ("BAPI_EQUI_CHANGE","שינוי ציוד"),
            ("BAPI_EQUI_GETDETAIL","שליפת פרטי ציוד"),
            ("BAPI_EQMT_INSTALL","התקנת ציוד במיקום/ציוד אב")],
  "progs": [("RIEQUI20","רשימת ציוד (IH08)"),
            ("RIEQUI00","ניתוח היסטוריית התקנות הציוד")],
  "s4_status": "ללא שינוי במודל הנתונים (תואם)", "s4_repl": "EQUI (זהה)",
  "s4_tcode": "IE01-03 נתמכים; אסטרטגי: Fiori", "fiori": "Manage Technical Objects (F2079)"},
 {"name": "EQKT", "he": "טקסטים (תיאורים) של ציוד", "en": "Equipment short texts",
  "tcodes": "IE02; IE03",
  "fields": [("EQUNR","Equipment number","מספר ציוד"),
             ("SPRAS","Language key","שפה"),
             ("EQKTX","Equipment description","תיאור הציוד")],
  "funcs": [("EQUIPMENT_TEXT_READ","קריאת טקסט ציוד")],
  "progs": [("RIEQUI20","רשימת ציוד עם תיאורים")],
  "s4_status": "ללא שינוי (תואם)", "s4_repl": "EQKT (זהה)",
  "s4_tcode": "IE0*; Fiori", "fiori": "Manage Technical Objects (F2079)"},
 {"name": "EQUZ", "he": "פלח זמן של ציוד (התקנות/שיוך)", "en": "Equipment time segment",
  "tcodes": "IE02; IE4N (התקנה/פירוק)",
  "fields": [("EQUNR","Equipment number","מספר ציוד"),
             ("DATBI","Valid-to date of segment","תקף עד"),
             ("HEQUI","Superior equipment","ציוד עליון"),
             ("ILOAN","Location/account key","מפתח מיקום/חיוב"),
             ("EQLFN","Time segment counter","מונה פלח זמן")],
  "funcs": [("EQUI_TIMESEGMENT_READ","קריאת פלחי הזמן של הציוד"),
            ("EQUIPMENT_DISMANTLE","פירוק ציוד מהמיקום/ציוד אב")],
  "progs": [("RIEQUI00","ניתוח שינויי התקנה והיסטוריה")],
  "s4_status": "ללא שינוי (תואם)", "s4_repl": "EQUZ (זהה)",
  "s4_tcode": "IE4N נתמך; Fiori", "fiori": "Manage Technical Objects (F2079)"},
 {"name": "OBJK", "he": "רשימת אובייקטים / מספרים סידוריים", "en": "Object list / serial assignment",
  "tcodes": "IQ01/IQ02/IQ03; IQ08, IQ09",
  "fields": [("OBKNR","Object list number","מספר רשומת אובייקט"),
             ("EQUNR","Equipment number","מספר ציוד"),
             ("SERNR","Serial number","מספר סידורי"),
             ("MATNR","Material number","חומר"),
             ("TASER","Serial number profile usage","שימוש פרופיל סריאלי")],
  "funcs": [("SERIAL_NUMBER_CREATE","יצירת מספר סידורי"),
            ("SERNR_ADD_TO_DOCUMENT","שיוך סריאל לתעודת מלאי"),
            ("BAPI_OBJCL_CREATE","שיוך סיווג לאובייקט")],
  "progs": [("RISERNR0","דוח סריאלים מול תנועות מלאי")],
  "s4_status": "ללא שינוי (תואם)", "s4_repl": "OBJK (זהה)",
  "s4_tcode": "IQ0*; Fiori", "fiori": "Manage Technical Objects (F2079)"},
]},
# ---------------- 3. BOM ---------------------------------------------------
{"title": "3. עצי מוצר של אחזקה (BOM)", "theme": "master", "tables": [
 {"name": "STKO", "he": "כותרת עץ מוצר", "en": "BOM header",
  "tcodes": "IB01/CS01; CS02/CS03",
  "fields": [("STLNR","Bill of material number","מספר עץ מוצר"),
             ("STLTY","BOM category (E/T/M)","קטגוריית BOM"),
             ("STLAL","Alternative BOM","חלופת עץ מוצר"),
             ("BMENG","Base quantity","כמות בסיס"),
             ("STLST","BOM status","סטטוס עץ מוצר")],
  "funcs": [("CSAP_MAT_BOM_CREATE","יצירת עץ מוצר"),
            ("CSAP_MAT_BOM_MAINTAIN","תחזוקת עץ מוצר"),
            ("CS_BT_BOM_HEADER_READ","קריאת כותרת עץ מוצר")],
  "progs": [("RCSBI010","פיצוץ עץ מוצר רב-רמתי")],
  "s4_status": "ללא שינוי (תואם)", "s4_repl": "STKO (זהה)",
  "s4_tcode": "CS01-03 / IB01; Fiori", "fiori": "Manage Bills of Material (אמת ID)"},
 {"name": "STPO", "he": "פריטי עץ מוצר (רכיבים)", "en": "BOM item (components)",
  "tcodes": "IB02/CS02; CS11, CS15",
  "fields": [("STLNR","BOM number","מספר עץ מוצר"),
             ("POSNR","BOM item number","מספר פריט"),
             ("IDNRK","Component (BOM item material)","רכיב / חומר"),
             ("MENGE","Component quantity","כמות רכיב"),
             ("POSTP","Item category","סוג פריט")],
  "funcs": [("CSAP_BOM_ITEM_MAINTAIN","תחזוקת פריטי עץ מוצר"),
            ("CS_BOM_EXPL_MAT_RC1","פיצוץ עץ מוצר לחומר")],
  "progs": [("RCSBI010","פיצוץ רב-רמתי"),("CS_BT","דוח Where-Used לרכיב")],
  "s4_status": "ללא שינוי (תואם)", "s4_repl": "STPO (זהה)",
  "s4_tcode": "CS0*; Fiori", "fiori": "Manage Bills of Material (אמת ID)"},
 {"name": "MAST", "he": "קישור חומר לעץ מוצר", "en": "Material-to-BOM link",
  "tcodes": "CS01; CS02/CS03",
  "fields": [("MATNR","Material number","חומר"),
             ("WERKS","Plant","מפעל"),
             ("STLAN","BOM usage (4=maintenance)","שימוש עץ מוצר"),
             ("STLNR","BOM number","מספר עץ מוצר"),
             ("STLAL","Alternative BOM","חלופה")],
  "funcs": [("CSAP_MAT_BOM_READ","קריאת קישור חומר-BOM")],
  "progs": [("RCSBI010","פיצוץ עץ מוצר לחומר")],
  "s4_status": "ללא שינוי (תואם)", "s4_repl": "MAST (זהה)",
  "s4_tcode": "CS0*; Fiori", "fiori": "Manage Bills of Material (אמת ID)"},
 {"name": "EQST", "he": "קישור ציוד לעץ מוצר", "en": "Equipment-to-BOM link",
  "tcodes": "IB01; IB02/IB03",
  "fields": [("EQUNR","Equipment number","מספר ציוד"),
             ("STLNR","BOM number","מספר עץ מוצר"),
             ("STLAL","Alternative BOM","חלופה"),
             ("STLAN","BOM usage","שימוש עץ מוצר")],
  "funcs": [("CS_BOM_EXPL_EQU_RC1","פיצוץ עץ מוצר של ציוד")],
  "progs": [("RCSBI010","פיצוץ עץ מוצר של ציוד")],
  "s4_status": "ללא שינוי (תואם)", "s4_repl": "EQST (זהה)",
  "s4_tcode": "IB01-03; Fiori", "fiori": "Manage Bills of Material (אמת ID)"},
 {"name": "TPST", "he": "קישור מיקום פונקציונלי לעץ מוצר", "en": "Functional location-to-BOM link",
  "tcodes": "IB11; IB12/IB13",
  "fields": [("TPLNR","Functional location","מיקום פונקציונלי"),
             ("STLNR","BOM number","מספר עץ מוצר"),
             ("STLAL","Alternative BOM","חלופה"),
             ("STLAN","BOM usage","שימוש עץ מוצר")],
  "funcs": [("CS_BOM_EXPL_FLO_RC1","פיצוץ עץ מוצר של מיקום")],
  "progs": [("RCSBI020","פיצוץ עצי מוצר של מיקומים")],
  "s4_status": "ללא שינוי (תואם)", "s4_repl": "TPST (זהה)",
  "s4_tcode": "IB11-13; Fiori", "fiori": "Manage Bills of Material (אמת ID)"},
]},
# ---------------- 4. Measuring points -------------------------------------
{"title": "4. נקודות מדידה ומונים", "theme": "master", "tables": [
 {"name": "IMPTT", "he": "רשומת אב של נקודת מדידה / מונה", "en": "Measuring point master",
  "tcodes": "IK01/IK02/IK03; IK07, IK08",
  "fields": [("POINT","Measuring point number","מספר נקודת מדידה"),
             ("MPOBJ","Measuring object (eq./FL)","אובייקט מדידה"),
             ("MPTYP","Measuring point category","סוג נקודת מדידה"),
             ("ATINN","Internal characteristic","מאפיין סיווג פנימי"),
             ("INDCT","Counter direction indicator","קוד כיוון מונה"),
             ("DECIM","Decimal places","ספרות עשרוניות")],
  "funcs": [("BAPI_MEASUREMENTPOINT_CREATE","יצירת נקודת מדידה"),
            ("BAPI_MEASUREMENTPOINT_GETLIST","שליפת רשימת נקודות מדידה"),
            ("MEASUREMENT_POINT_READ","קריאת פרטי נקודת מדידה")],
  "progs": [("RIMPOINT","רשימת נקודות מדידה ומונים")],
  "s4_status": "ללא שינוי (תואם)", "s4_repl": "IMPTT (זהה)",
  "s4_tcode": "IK01-03; Fiori", "fiori": "Manage Measurement Documents (אמת ID)"},
 {"name": "IMRG", "he": "מסמכי מדידה (קריאות מונה/ערך)", "en": "Measurement documents / readings",
  "tcodes": "IK11/IK12/IK13; IK21, IK41",
  "fields": [("MDOCM","Measurement document number","מספר מסמך מדידה"),
             ("POINT","Measuring point","נקודת מדידה"),
             ("RECDV","Measurement reading value","ערך נמדד"),
             ("READG","Counter reading","קריאת מונה"),
             ("CNTRR","Counter reading (total)","קריאה כוללת"),
             ("CDIFF","Counter difference","הפרש מונה")],
  "funcs": [("BAPI_MEASUREMENTDOCUM_CREATE","יצירת מסמך מדידה"),
            ("BAPI_MEASUREMENTDOCUM_CREATEM","יצירת מסמכי מדידה מרובים"),
            ("MEASUREM_DOCUM_RFC_SINGLE_001","קליטת קריאה בודדת RFC")],
  "progs": [("RISTRA20","ניטור מועדים מבוסס מונה (IP30)")],
  "s4_status": "ללא שינוי (תואם)", "s4_repl": "IMRG (זהה)",
  "s4_tcode": "IK11-13; Fiori", "fiori": "Enter Measurement Readings (אמת ID)"},
]},
# ---------------- 5. Catalogs ---------------------------------------------
{"title": "5. קטלוגים, קודים ופרופילים", "theme": "config", "tables": [
 {"name": "QPCD", "he": "קודים בתוך קבוצת קוד", "en": "Codes within a code group",
  "tcodes": "QS51; QS61",
  "fields": [("KATALOGART","Catalog type (B/C/D/2/5)","סוג קטלוג"),
             ("CODEGRUPPE","Code group","קבוצת קוד"),
             ("CODE","Code","קוד"),
             ("VERSION","Code version","גרסת קוד")],
  "funcs": [("QPK1_CODE_TEXT_READ","קריאת קוד וטקסט"),
            ("QPK1_CATALOG_READ","קריאת קטלוג")],
  "progs": [("RQEEAW01","סקירת מבנה קטלוגים")],
  "s4_status": "ללא שינוי (Customizing/Master תואם)", "s4_repl": "QPCD (זהה)",
  "s4_tcode": "QS41/QS51 (זהה)", "fiori": "אין Fiori ייעודי (Customizing)"},
 {"name": "QPGR", "he": "קבוצות קוד (במקום QPCR הלא-תקני)", "en": "Code groups (correct table)",
  "tcodes": "QS51; QS61",
  "fields": [("KATALOGART","Catalog type","סוג קטלוג"),
             ("CODEGRUPPE","Code group","קבוצת קוד"),
             ("STATUS","Code group status","סטטוס קבוצה"),
             ("VERWMERKM","Selection set usage","שימוש בקבוצת בחירה")],
  "funcs": [("QPK1_CODEGROUP_READ","קריאת קבוצת קוד")],
  "progs": [("RQEEAW02","דוח קבוצות קוד וקודים")],
  "s4_status": "ללא שינוי (תואם)", "s4_repl": "QPGR (זהה)",
  "s4_tcode": "QS51 (זהה)", "fiori": "אין Fiori ייעודי (Customizing)"},
 {"name": "T352B", "he": "שיוך פרופיל קטלוג לקבוצות בחירה", "en": "Catalog profile to selection-set link",
  "tcodes": "OIM1; OIN4",
  "fields": [("RBNR","Catalog profile","פרופיל קטלוג"),
             ("KATALOGART","Catalog type","סוג קטלוג"),
             ("CODEGRUPPE","Allowed code group","קבוצת קוד מותרת"),
             ("AUSWMENGEN","Selection set","קבוצת בחירה")],
  "funcs": [("CATALOG_PROFILE_READ","קריאת פרופיל קטלוג")],
  "progs": [("RIPROFCAT","ניתוח פרופילי קטלוג ושיוכם")],
  "s4_status": "ללא שינוי (Customizing תואם)", "s4_repl": "T352B (זהה)",
  "s4_tcode": "SPRO (זהה)", "fiori": "אין Fiori ייעודי (Customizing)"},
 {"name": "T352", "he": "פרופיל קטלוג (כותרת)", "en": "Catalog profile header",
  "tcodes": "OIM1; QCC0",
  "fields": [("RBNR","Catalog profile","פרופיל קטלוג"),
             ("HERKZ","Origin indicator","מחוון מקור"),
             ("RBNRX","Profile description","תיאור פרופיל")],
  "funcs": [("T352_READ","קריאת פרופיל קטלוג")],
  "progs": [("RIPROFCAT","ניתוח פרופילי קטלוג")],
  "s4_status": "ללא שינוי (Customizing תואם)", "s4_repl": "T352 (זהה)",
  "s4_tcode": "SPRO (זהה)", "fiori": "אין Fiori ייעודי (Customizing)"},
]},
# ---------------- 6. Notifications ----------------------------------------
{"title": "6. הודעות אחזקה (Notifications)", "theme": "trans", "tables": [
 {"name": "QMEL", "he": "כותרת הודעת אחזקה", "en": "Notification header",
  "tcodes": "IW21/IW22/IW23; IW24, IW28/IW29",
  "fields": [("QMNUM","Notification number","מספר הודעה"),
             ("QMART","Notification type (M1/M2/M3)","סוג הודעה"),
             ("QMTXT","Notification short text","תיאור קצר"),
             ("EQUNR","Reference equipment","ציוד מתייחס"),
             ("TPLNR","Reference functional location","מיקום מתייחס"),
             ("OBJNR","Object number (status)","מפתח אובייקט לסטטוס")],
  "funcs": [("BAPI_ALM_NOTIF_CREATE","יצירת הודעה דרך ממשק תקני"),
            ("BAPI_ALM_NOTIF_DATA_ADD","הוספת פריט/סיבה/פעילות"),
            ("BAPI_ALM_NOTIF_SAVE","שמירת ההודעה"),
            ("BAPI_ALM_NOTIF_CLOSE","סגירת ההודעה"),
            ("BAPI_ALM_NOTIF_GET_DETAIL","שליפת פרטי הודעה מלאים")],
  "progs": [("RIQMEL20","רשימת הודעות (IW28/29)"),
            ("RIQMEL00","ניתוח דפוסי כשל")],
  "s4_status": "ללא שינוי במודל הנתונים (תואם)", "s4_repl": "QMEL (זהה)",
  "s4_tcode": "IW21-23 נתמכים; אסטרטגי: Fiori", "fiori": "Report Malfunction (F2215)"},
 {"name": "QMFE", "he": "פריטי הודעה (ליקויים)", "en": "Notification items (damage)",
  "tcodes": "IW22; IW66",
  "fields": [("QMNUM","Notification number","מספר הודעה"),
             ("FENUM","Item number","מספר פריט"),
             ("FEGRP","Damage code group","קבוצת קוד ליקוי"),
             ("FECOD","Damage code","קוד ליקוי"),
             ("OTGRP","Object part code group","קבוצת אובייקט חלקי"),
             ("OTEIL","Object part code","קוד אובייקט חלקי")],
  "funcs": [("BAPI_ALM_NOTIF_DATA_ADD","הוספת פריט ליקוי"),
            ("NOTIF_ITEM_READ","קריאת פריטי הודעה")],
  "progs": [("RIQMEL00","ניתוח פריטי ליקוי ודפוסי כשל")],
  "s4_status": "ללא שינוי (תואם)", "s4_repl": "QMFE (זהה)",
  "s4_tcode": "IW22; Fiori", "fiori": "Find Maintenance Notification (אמת ID)"},
 {"name": "QMUR", "he": "סיבות לליקוי", "en": "Notification causes",
  "tcodes": "IW22; IW67",
  "fields": [("QMNUM","Notification number","מספר הודעה"),
             ("FENUM","Item number","מספר פריט"),
             ("URNUM","Cause number","מספר סיבה"),
             ("URGRP","Cause code group","קבוצת קוד סיבה"),
             ("URCOD","Cause code","קוד סיבה")],
  "funcs": [("NOTIF_CAUSE_READ","קריאת סיבות הליקוי")],
  "progs": [("RIQMUR00","דוח סיבות לניתוח שורש (RCA)")],
  "s4_status": "ללא שינוי (תואם)", "s4_repl": "QMUR (זהה)",
  "s4_tcode": "IW22; Fiori", "fiori": "Find Maintenance Notification (אמת ID)"},
 {"name": "QMMA", "he": "פעילויות שבוצעו (במקום QMAA הלא-תקני)", "en": "Notification activities (correct table)",
  "tcodes": "IW22; IW67",
  "fields": [("QMNUM","Notification number","מספר הודעה"),
             ("MANUM","Activity number","מספר פעילות"),
             ("MNGRP","Activity code group","קבוצת קוד פעילות"),
             ("MNCOD","Activity code","קוד פעילות")],
  "funcs": [("NOTIF_ACTIVITY_READ","קריאת פעילויות שבוצעו")],
  "progs": [("RIQMEL00","ניתוח פעילויות הודעה")],
  "s4_status": "ללא שינוי (תואם)", "s4_repl": "QMMA (זהה)",
  "s4_tcode": "IW22; Fiori", "fiori": "Report Malfunction (F2215)"},
 {"name": "QMSM", "he": "משימות לטיפול (במקום QMAS הלא-תקני)", "en": "Notification tasks (correct table)",
  "tcodes": "IW22; IW66",
  "fields": [("QMNUM","Notification number","מספר הודעה"),
             ("MANUM","Task number","מספר משימה"),
             ("MNGRP","Task code group","קבוצת קוד משימה"),
             ("MNCOD","Task code","קוד משימה"),
             ("PSTER","Task planned completion","מועד יעד למשימה")],
  "funcs": [("NOTIF_TASK_READ","קריאת משימות"),
            ("BAPI_ALM_NOTIF_TASK_ADD","הוספת משימה")],
  "progs": [("RIQMEL20","מעקב משימות פתוחות")],
  "s4_status": "ללא שינוי (תואם)", "s4_repl": "QMSM (זהה)",
  "s4_tcode": "IW22; Fiori", "fiori": "Find Maintenance Notification (אמת ID)"},
]},
# ---------------- 7. Orders -----------------------------------------------
{"title": "7. פקודות עבודה (פק\"ע)", "theme": "trans", "tables": [
 {"name": "AUFK", "he": "נתוני אב של פקודת עבודה", "en": "Order master data",
  "tcodes": "IW31/IW32/IW33; IW34, IW38/IW39",
  "fields": [("AUFNR","Order number","מספר פק\"ע"),
             ("AUART","Order type (PM01/02/03)","סוג פק\"ע"),
             ("KTEXT","Order short text","תיאור הפק\"ע"),
             ("OBJNR","Object number (status/CO)","מפתח אובייקט"),
             ("KOSTV","Responsible cost center","מרכז עלות אחראי")],
  "funcs": [("BAPI_ALM_ORDER_MAINTAIN","יצירה/שינוי פק\"ע (ממשק מרכזי)"),
            ("BAPI_ALM_ORDER_GET_DETAIL","שליפת פרטי פק\"ע"),
            ("BAPI_ALM_ORDERHEAD_GET_LIST","שליפת רשימת פק\"ע")],
  "progs": [("RIAUFK20","רשימת פק\"ע (IW38/39)")],
  "s4_status": "מותאם - מודל נתונים פשוט יותר לפקודות", "s4_repl": "AUFK (זהה) + סטטוס/עלות ב-ACDOCA",
  "s4_tcode": "IW31-33 נתמכים; אסטרטגי: Fiori", "fiori": "Find Maintenance Order (F2393)"},
 {"name": "AFKO", "he": "כותרת פקודה (נתוני אחזקה/ייצור)", "en": "Order header data",
  "tcodes": "IW32; IW37N",
  "fields": [("AUFNR","Order number","מספר פק\"ע"),
             ("AUFPL","Routing/operation plan number","מספר תוכנית פעולות"),
             ("GLTRP","Basic finish date","תאריך סיום בסיסי"),
             ("GSTRP","Basic start date","תאריך התחלה בסיסי"),
             ("PLNBEZ","BOM/material reference","הפניית עץ מוצר/חומר")],
  "funcs": [("CO_ZF_ORDER_HEADER_READ","קריאת כותרת הפקודה")],
  "progs": [("RIAUFK20","רשימת פק\"ע")],
  "s4_status": "מותאם (תואם עם פישוטי תזמון)", "s4_repl": "AFKO (זהה)",
  "s4_tcode": "IW32; Fiori", "fiori": "Create Maintenance Order (אמת ID)"},
 {"name": "AFPO", "he": "פריט פקודה", "en": "Order item",
  "tcodes": "IW32; IW33",
  "fields": [("AUFNR","Order number","מספר פק\"ע"),
             ("POSNR","Order item number","מספר פריט"),
             ("MATNR","Material (refurbishment)","חומר (שיפוץ)"),
             ("PSAMG","Order item quantity","כמות פריט")],
  "funcs": [("CO_ZF_ORDER_ITEM_READ","קריאת פריט הפקודה")],
  "progs": [("RIAFPO00","ניתוח פריטי פקודה")],
  "s4_status": "מותאם (תואם)", "s4_repl": "AFPO (זהה)",
  "s4_tcode": "IW32; Fiori", "fiori": "Find Maintenance Order (F2393)"},
 {"name": "AFIH", "he": "כותרת פקודה - נתוני אחזקה (PM)", "en": "Maintenance order header (PM)",
  "tcodes": "IW32; IW33",
  "fields": [("AUFNR","Order number","מספר פק\"ע"),
             ("EQUNR","Reference equipment","ציוד מתייחס"),
             ("TPLNR","Reference functional location","מיקום מתייחס"),
             ("ILART","Maintenance activity type","סוג פעילות אחזקה"),
             ("PRIOK","Priority","עדיפות"),
             ("GEWRK","Main work center","מרכז עבודה ראשי")],
  "funcs": [("BAPI_ALM_ORDER_GET_DETAIL","שליפת נתוני אחזקה של הפק\"ע")],
  "progs": [("RIAUFK20","רשימת פק\"ע אחזקה")],
  "s4_status": "מותאם (תואם)", "s4_repl": "AFIH (זהה)",
  "s4_tcode": "IW32; Fiori", "fiori": "Find Maintenance Order (F2393)"},
 {"name": "AFWI", "he": "תנועות מלאי המקושרות לדיווחי ביצוע", "en": "Goods movements for confirmations",
  "tcodes": "IW41; IW42",
  "fields": [("RUECK","Confirmation counter","מספר דיווח ביצוע"),
             ("RMZHL","Confirmation number","מונה דיווח"),
             ("MBLNR","Material document number","מסמך חומר"),
             ("MATNR","Material","חומר")],
  "funcs": [("BAPI_ALM_CONF_CREATE","יצירת דיווח ביצוע עם תנועות מלאי")],
  "progs": [("RIAFRU00","ניתוח דיווחי ביצוע ותנועות נלוות")],
  "s4_status": "מותאם (תנועת מלאי ב-MATDOC)", "s4_repl": "AFWI (זהה); מסמך חומר ב-MATDOC",
  "s4_tcode": "IW41/IW42; Fiori Confirmation", "fiori": "Confirm Maintenance Order (אמת ID)"},
]},
# ---------------- 8. Status -----------------------------------------------
{"title": "8. ניהול סטטוסים (Status Mgmt)", "theme": "config", "tables": [
 {"name": "JEST", "he": "סטטוס אובייקט (פעיל/לא פעיל)", "en": "Individual object status",
  "tcodes": "BS22/BS23; IW33, IE03",
  "fields": [("OBJNR","Object number","מפתח אובייקט"),
             ("STAT","Status code","קוד סטטוס"),
             ("INACT","Inactive indicator","מחוון לא-פעיל"),
             ("CHGNR","Change number","מספר שינוי")],
  "funcs": [("STATUS_READ","קריאת סטטוסי אובייקט"),
            ("STATUS_CHANGE_INTERN","שינוי סטטוס מערכת פנימי"),
            ("STATUS_CHANGE_EXTERN","שינוי סטטוס משתמש")],
  "progs": [("RISTAT00","ניתוח סטטוסים פעילים לאובייקטים")],
  "s4_status": "ללא שינוי (תואם)", "s4_repl": "JEST (זהה)",
  "s4_tcode": "BS22/BS23 (זהה)", "fiori": "מוטמע באפליקציות PM (אין ייעודי)"},
 {"name": "JSTO", "he": "כותרת אובייקט סטטוס + פרופיל", "en": "Status object header + profile",
  "tcodes": "BS02/BS03; OIBS",
  "fields": [("OBJNR","Object number","מפתח אובייקט"),
             ("STSMA","Status profile","פרופיל סטטוס"),
             ("OBTYP","Object type","סוג אובייקט"),
             ("CHGKZ","Change flag","מחוון שינוי")],
  "funcs": [("STATUS_OBJECT_CREATE","יצירת אובייקט סטטוס"),
            ("STATUS_PROFILE_READ","קריאת פרופיל סטטוס")],
  "progs": [("RIBS0200","סקירת פרופילי סטטוס וחסימות")],
  "s4_status": "ללא שינוי (תואם)", "s4_repl": "JSTO (זהה)",
  "s4_tcode": "BS02/BS03 (זהה)", "fiori": "מוטמע באפליקציות PM (אין ייעודי)"},
 {"name": "TJ02T", "he": "טקסטים של סטטוסי מערכת", "en": "System status texts",
  "tcodes": "BS23; BS22",
  "fields": [("ISTAT","Internal system status","סטטוס מערכת פנימי"),
             ("SPRAS","Language key","שפה"),
             ("TXT04","Status code (CRTD/REL/TECO/CLSD)","קוד סטטוס מקוצר"),
             ("TXT30","Status long text","תיאור ארוך")],
  "funcs": [("STATUS_TEXT_READ","קריאת טקסט סטטוס מערכת")],
  "progs": [("RISTAT00","פענוח קודי סטטוס בדוחות")],
  "s4_status": "ללא שינוי (תואם)", "s4_repl": "TJ02T (זהה)",
  "s4_tcode": "BS23 (זהה)", "fiori": "מוטמע (אין ייעודי)"},
 {"name": "TJ30T", "he": "טקסטים של סטטוסי משתמש", "en": "User status texts",
  "tcodes": "BS02; BS03",
  "fields": [("STSMA","Status profile","פרופיל סטטוס"),
             ("ESTAT","User status","סטטוס משתמש"),
             ("SPRAS","Language key","שפה"),
             ("TXT04","User status short code","קוד מקוצר"),
             ("TXT30","User status text","תיאור סטטוס משתמש")],
  "funcs": [("USER_STATUS_TEXT_READ","קריאת טקסט סטטוס משתמש")],
  "progs": [("RIBS0200","סקירת סטטוסי משתמש בפרופיל")],
  "s4_status": "ללא שינוי (תואם)", "s4_repl": "TJ30T (זהה)",
  "s4_tcode": "BS02/BS03 (זהה)", "fiori": "מוטמע (אין ייעודי)"},
]},
# ---------------- 9. MM ----------------------------------------------------
{"title": "9. אינטגרציית מלאי ורכש (PM-MM)", "theme": "cross", "tables": [
 {"name": "RESB", "he": "הזמנת רכיבים (Reservation)", "en": "Reservation / dependent requirements",
  "tcodes": "MB1A/MIGO; MB21, IW32",
  "fields": [("RSNUM","Reservation number","מספר הזמנה"),
             ("RSPOS","Reservation item","פריט הזמנה"),
             ("MATNR","Component material","חומר רכיב"),
             ("BDMNG","Requirement quantity","כמות נדרשת"),
             ("ENMNG","Quantity withdrawn","כמות שנפלטה"),
             ("AUFNR","Order (account assignment)","פק\"ע מחויבת")],
  "funcs": [("BAPI_RESERVATION_CREATE1","יצירת הזמנת מלאי"),
            ("BAPI_GOODSMVT_CREATE","רישום פליטת חומר (261)"),
            ("RESERVATION_READ","קריאת הזמנת רכיבים")],
  "progs": [("RM07RESL","דוח הזמנות מלאי פתוחות")],
  "s4_status": "ללא שינוי (תואם)", "s4_repl": "RESB (זהה)",
  "s4_tcode": "MIGO/MB1A נתמכים; Fiori", "fiori": "Post Goods Movement / MIGO (אמת ID)"},
 {"name": "EBAN", "he": "דרישת רכש (Purchase Requisition)", "en": "Purchase requisition",
  "tcodes": "ME51N; ME52N/ME53N, ME57",
  "fields": [("BANFN","Purchase requisition number","מספר דרישת רכש"),
             ("BNFPO","Requisition item","פריט דרישה"),
             ("MATNR","Material","חומר"),
             ("KNTTP","Account assignment cat. (F=order)","סוג חיוב"),
             ("MENGE","Requisition quantity","כמות")],
  "funcs": [("BAPI_PR_CREATE","יצירת דרישת רכש"),
            ("BAPI_REQUISITION_GETDETAIL","שליפת פרטי דרישת רכש")],
  "progs": [("RM06BANF","דוח דרישות רכש פתוחות")],
  "s4_status": "מותאם (Business Partner לספקים)", "s4_repl": "EBAN (זהה)",
  "s4_tcode": "ME51N נתמך; ספק דרך BP", "fiori": "Manage Purchase Requisitions (אמת ID)"},
 {"name": "EBKN", "he": "חיוב דרישת רכש (לפק\"ע)", "en": "Requisition account assignment",
  "tcodes": "ME51N; ME52N",
  "fields": [("BANFN","Purchase requisition number","מספר דרישת רכש"),
             ("BNFPO","Requisition item","פריט דרישה"),
             ("AUFNR","Order (account assignment)","פק\"ע מחויבת"),
             ("KOSTL","Cost center","מרכז עלות"),
             ("SAKTO","G/L account","חשבון ראשי")],
  "funcs": [("ACCOUNT_ASSIGNMENT_READ","קריאת נתוני חיוב לדרישה")],
  "progs": [("RM06BANF","דוח דרישות רכש מחויבות לפק\"ע")],
  "s4_status": "מותאם (חיוב ל-ACDOCA)", "s4_repl": "EBKN (זהה); עלות בפועל ב-ACDOCA",
  "s4_tcode": "ME51N; Fiori", "fiori": "Manage Purchase Requisitions (אמת ID)"},
 {"name": "MSEG", "he": "פריטי מסמך חומר (תנועות מלאי)", "en": "Material document items",
  "tcodes": "MIGO/MB1A; MB31, MB1C",
  "fields": [("MBLNR","Material document number","מספר מסמך חומר"),
             ("ZEILE","Material document item","פריט מסמך"),
             ("BWART","Movement type (261/101)","סוג תנועה"),
             ("MATNR","Material","חומר"),
             ("MENGE","Quantity","כמות"),
             ("AUFNR","Order (account assignment)","פק\"ע מחויבת")],
  "funcs": [("BAPI_GOODSMVT_CREATE","רישום תנועת מלאי"),
            ("BAPI_GOODSMVT_GETITEMS","שליפת פריטי תנועה")],
  "progs": [("RM07DOCS","דוח מסמכי חומר לפי פק\"ע/תנועה")],
  "s4_status": "הוחלף - מסמכי חומר מאוחדים ב-MATDOC", "s4_repl": "MATDOC (MSEG=View תאימות)",
  "s4_tcode": "MIGO נתמך; Fiori", "fiori": "Post Goods Movement (אמת ID)"},
 {"name": "MKPF", "he": "כותרת מסמך חומר", "en": "Material document header",
  "tcodes": "MIGO; MB1A",
  "fields": [("MBLNR","Material document number","מספר מסמך חומר"),
             ("MJAHR","Material document year","שנת מסמך"),
             ("BLDAT","Document date","תאריך מסמך"),
             ("BUDAT","Posting date","תאריך רישום"),
             ("USNAM","User name","משתמש")],
  "funcs": [("BAPI_GOODSMVT_GETDETAIL","שליפת פרטי מסמך חומר")],
  "progs": [("RM07DOCS","דוח מסמכי חומר")],
  "s4_status": "הוחלף - מאוחד ב-MATDOC", "s4_repl": "MATDOC (MKPF=View תאימות)",
  "s4_tcode": "MIGO נתמך; Fiori", "fiori": "Post Goods Movement (אמת ID)"},
]},
# ---------------- 10. CO ---------------------------------------------------
{"title": "10. עלויות והתחשבנות (PM-CO)", "theme": "cross", "tables": [
 {"name": "COSP", "he": "סך עלויות חיצוניות (Primary)", "en": "CO object cost totals - external postings",
  "tcodes": "KO88/KO8G; KOB1, S_ALR_87013611",
  "fields": [("OBJNR","CO object (order)","אובייקט CO"),
             ("KSTAR","Cost element","אלמנט עלות"),
             ("WRTTP","Value type (01 plan/04 actual)","סוג ערך"),
             ("GJAHR","Fiscal year","שנת כספים"),
             ("WTG001","Period value 001","ערך תקופה 001")],
  "funcs": [("K_COSTS_READ","קריאת עלויות אובייקט"),
            ("K_ORDER_SETTLEMENT","ביצוע התחשבנות פק\"ע")],
  "progs": [("RKO7KO88","תוכנית ההתחשבנות (KO88)")],
  "s4_status": "הוחלף - עלויות ב-Universal Journal", "s4_repl": "ACDOCA (COSP=View תאימות)",
  "s4_tcode": "KO88 נתמך; Fiori Settlement", "fiori": "Maintenance Order Actuals (אמת ID)"},
 {"name": "COSS", "he": "סך עלויות פנימיות (Secondary)", "en": "CO object cost totals - internal postings",
  "tcodes": "KO88; KOB1",
  "fields": [("OBJNR","CO object","אובייקט CO"),
             ("KSTAR","Secondary cost element","אלמנט עלות משני"),
             ("WRTTP","Value type","סוג ערך"),
             ("GJAHR","Fiscal year","שנת כספים"),
             ("PARGB","Partner / sender object","אובייקט שותף")],
  "funcs": [("K_COSTS_READ","קריאת עלויות פנימיות")],
  "progs": [("RKAEP000","דוח פריטי עלות בפועל")],
  "s4_status": "הוחלף - עלויות ב-Universal Journal", "s4_repl": "ACDOCA (COSS=View תאימות)",
  "s4_tcode": "KO88 נתמך; Fiori", "fiori": "Maintenance Order Actuals (אמת ID)"},
 {"name": "COBRA", "he": "כותרת חוק התחשבנות", "en": "Settlement rule header",
  "tcodes": "KO02; IW32",
  "fields": [("OBJNR","Source object (order)","אובייקט מקור"),
             ("BUREG","Settlement rule","חוק התחשבנות"),
             ("PERBZ","Period type","סוג תקופה"),
             ("ERLKZ","Completion indicator","מחוון השלמה")],
  "funcs": [("K_SETTLEMENT_RULE_READ","קריאת חוק התחשבנות")],
  "progs": [("RKACOR40","איתור חוקי התחשבנות חסרים")],
  "s4_status": "ללא שינוי (תואם)", "s4_repl": "COBRA (זהה)",
  "s4_tcode": "KO02/IW32 (זהה)", "fiori": "Manage Settlement Rules (אמת ID)"},
 {"name": "COBRB", "he": "פריטי חוק התחשבנות (חוקי חלוקה)", "en": "Settlement / distribution rules",
  "tcodes": "KO02; IW32",
  "fields": [("OBJNR","Source object","אובייקט מקור"),
             ("BUREG","Settlement rule","חוק התחשבנות"),
             ("KONTY","Receiver category","סוג מקבל"),
             ("EMPGE","Settlement receiver","המקבל בפועל"),
             ("PROZS","Settlement percentage","אחוז התחשבנות")],
  "funcs": [("K_SETTLEMENT_RULE_READ","קריאת פריטי חוק התחשבנות")],
  "progs": [("RKACOR40","בדיקת תקינות חוקי חלוקה")],
  "s4_status": "ללא שינוי (תואם)", "s4_repl": "COBRB (זהה)",
  "s4_tcode": "KO02/IW32 (זהה)", "fiori": "Manage Settlement Rules (אמת ID)"},
]},
# ---------------- 11. Preventive ------------------------------------------
{"title": "11. אחזקה מונעת ותוכניות", "theme": "master", "tables": [
 {"name": "PLKO", "he": "כותרת רשימת פעולות אחזקה", "en": "Task list header",
  "tcodes": "IA05/IA01/IA11; IA06, IA08",
  "fields": [("PLNNR","Task list group","קבוצת רשימת פעולות"),
             ("PLNAL","Group counter / alternative","חלופת רשימה"),
             ("PLNTY","Task list type (A/E/T)","סוג רשימה"),
             ("VERWE","Task list usage","שימוש רשימה"),
             ("STATU","Task list status","סטטוס רשימה")],
  "funcs": [("BAPI_TASKLIST_CREATE","יצירת רשימת פעולות"),
            ("CARO_TASKLIST_READ","קריאת רשימת פעולות")],
  "progs": [("RIPLKO00","דוח רשימות פעולות לפי קבוצה")],
  "s4_status": "ללא שינוי (תואם)", "s4_repl": "PLKO (זהה)",
  "s4_tcode": "IA05-IA12; Fiori", "fiori": "Manage Task Lists (אמת ID)"},
 {"name": "PLPO", "he": "פעולות ברשימת פעולות", "en": "Task list operations",
  "tcodes": "IA01/IA05; IA06",
  "fields": [("PLNTY","Task list type","סוג רשימה"),
             ("PLNNR","Task list group","קבוצת רשימה"),
             ("VORNR","Operation number","מספר פעולה"),
             ("ARBID","Work center (object id)","מרכז עבודה"),
             ("STEUS","Control key","מפתח בקרה"),
             ("ARBEI","Planned work","עבודה מתוכננת")],
  "funcs": [("CP_DI_OPERATION_READ","קריאת פעולות הרשימה")],
  "progs": [("RIPLKO00","דוח פעולות רשימה")],
  "s4_status": "ללא שינוי (תואם)", "s4_repl": "PLPO (זהה)",
  "s4_tcode": "IA0*; Fiori", "fiori": "Manage Task Lists (אמת ID)"},
 {"name": "MPLA", "he": "כותרת תכנית אחזקה ונתוני תזמון", "en": "Maintenance plan header",
  "tcodes": "IP01/IP02/IP03; IP41, IP42",
  "fields": [("WARPL","Maintenance plan number","מספר תכנית אחזקה"),
             ("MPTYP","Maintenance plan category","סוג תכנית"),
             ("STRAT","Maintenance strategy","אסטרטגיית אחזקה"),
             ("HORIZ","Call horizon","אופק שחרור"),
             ("ABRHO","Scheduling period","תקופת תזמון")],
  "funcs": [("BAPI_MAINTENANCEPLAN_CREATE","יצירת תכנית אחזקה"),
            ("MAINTENANCE_PLAN_SCHEDULE","תזמון תכנית אחזקה")],
  "progs": [("RIMPLA00","דוח תכניות אחזקה ופריטיהן")],
  "s4_status": "ללא שינוי (תואם)", "s4_repl": "MPLA (זהה)",
  "s4_tcode": "IP01-03; Fiori", "fiori": "Manage Maintenance Plans (אמת ID)"},
 {"name": "MPOS", "he": "פריט תכנית אחזקה", "en": "Maintenance item",
  "tcodes": "IP04; IP02",
  "fields": [("WAPOS","Maintenance item number","מספר פריט תכנית"),
             ("WARPL","Maintenance plan","תכנית אחזקה"),
             ("EQUNR","Reference equipment","ציוד מתייחס"),
             ("TPLNR","Reference functional location","מיקום מתייחס"),
             ("AUART","Order type to create","סוג פק\"ע ליצירה"),
             ("PLNNR","Task list group","קבוצת רשימת פעולות")],
  "funcs": [("MAINTENANCE_ITEM_READ","קריאת פריט תכנית אחזקה")],
  "progs": [("RIMPLA00","דוח פריטי תכנית אחזקה")],
  "s4_status": "ללא שינוי (תואם)", "s4_repl": "MPOS (זהה)",
  "s4_tcode": "IP04; Fiori", "fiori": "Manage Maintenance Plans (אמת ID)"},
 {"name": "MHIO", "he": "אובייקטי קריאת תכנית האחזקה", "en": "Maintenance plan call objects",
  "tcodes": "IP10; IP30, IP24",
  "fields": [("WARPL","Maintenance plan","תכנית אחזקה"),
             ("ABNUM","Maintenance call number","מספר קריאה"),
             ("NPLDA","Next planned date","תאריך מתוכנן הבא"),
             ("AUFNR","Order created","פק\"ע שנוצרה"),
             ("ABRUF","Call status","סטטוס קריאה")],
  "funcs": [("ISCHED_CALL_GENERATE","יצירת קריאת אחזקה ופק\"ע"),
            ("MAINTENANCE_PLAN_SCHEDULE","תזמון ושחרור קריאות")],
  "progs": [("RISTRA20","ניטור מועדים ברקע (IP30)")],
  "s4_status": "ללא שינוי (תואם)", "s4_repl": "MHIO (זהה)",
  "s4_tcode": "IP10/IP30; Fiori", "fiori": "Schedule Maintenance Plans (אמת ID)"},
 {"name": "MHIS", "he": "היסטוריית תזמון תכנית אחזקה", "en": "Maintenance plan scheduling history",
  "tcodes": "IP10; IP30",
  "fields": [("WARPL","Maintenance plan","תכנית אחזקה"),
             ("ABNUM","Call number","מספר קריאה"),
             ("NTERM","Planned date","מועד מתוכנן"),
             ("HORIZ","Call horizon at scheduling","אופק בעת תזמון")],
  "funcs": [("SCHEDULING_HISTORY_READ","קריאת היסטוריית תזמון")],
  "progs": [("RISTRA10","ניתוח אסטרטגיות וחבילות אחזקה")],
  "s4_status": "ללא שינוי (תואם)", "s4_repl": "MHIS (זהה)",
  "s4_tcode": "IP10/IP30; Fiori", "fiori": "Schedule Maintenance Plans (אמת ID)"},
]},
# ---------------- 12. History / Archiving ---------------------------------
{"title": "12. היסטוריה וארכיון", "theme": "trans", "tables": [
 {"name": "QMEL", "he": "הודעות (כולל היסטוריות) לניתוח אמינות", "en": "Notifications incl. historical",
  "tcodes": "IW64; IW66/IW67",
  "fields": [("QMNUM","Notification number","מספר הודעה"),
             ("QMART","Notification type","סוג הודעה"),
             ("AUSVN","Breakdown start","תחילת השבתה"),
             ("AUSBS","Breakdown end","סיום השבתה"),
             ("AUSZT","Breakdown duration","משך השבתה")],
  "funcs": [("BAPI_ALM_NOTIF_GET_DETAIL","שליפת פרטי הודעה היסטורית"),
            ("BAPI_ALM_NOTIF_LIST_FILTER","סינון הודעות לפי קריטריון")],
  "progs": [("RIQMEL20","דוח היסטוריית הודעות (MTBF/MTTR)")],
  "s4_status": "ללא שינוי; ניתוח דרך Analytics/CDS", "s4_repl": "QMEL (זהה) + CDS אנליטי",
  "s4_tcode": "IW64 נתמך; Fiori/Analytics", "fiori": "Maintenance Backlog (אמת ID)"},
 {"name": "AUFK", "he": "פקודות (כולל היסטוריות) לניתוח עלות", "en": "Orders incl. historical",
  "tcodes": "IW39; IW13",
  "fields": [("AUFNR","Order number","מספר פק\"ע"),
             ("AUART","Order type","סוג פק\"ע"),
             ("IDAT2","Technical completion date (TECO)","תאריך השלמה טכנית"),
             ("ERDAT","Creation date","תאריך יצירה"),
             ("OBJNR","Object number","מפתח אובייקט")],
  "funcs": [("BAPI_ALM_ORDERHEAD_GET_LIST","שליפת רשימת פק\"ע היסטוריות")],
  "progs": [("RIAUFK20","דוח פק\"ע סגורות וניתוח עלות מחזור חיים")],
  "s4_status": "מותאם; עלות ב-ACDOCA", "s4_repl": "AUFK (זהה); עלות ב-ACDOCA",
  "s4_tcode": "IW39 נתמך; Fiori", "fiori": "Find Maintenance Order (F2393)"},
 {"name": "ADMI_RUN", "he": "ריצות ארכוב נתונים", "en": "Archiving runs (ADK)",
  "tcodes": "SARA; AOBJ, DB15",
  "fields": [("RUNID","Archiving run ID","מזהה ריצת ארכוב"),
             ("OBJECT","Archiving object (PM_ORDER/PM_QMEL)","אובייקט ארכוב"),
             ("GENER_DATE","Generation date","תאריך יצירה"),
             ("STATUS","Run status","סטטוס ריצה")],
  "funcs": [("ARCHIVE_OPEN_FOR_WRITE","פתיחת ארכיון לכתיבה"),
            ("ARCHIVE_GET_NEXT_OBJECT","קריאת אובייקט הבא מהארכיון"),
            ("ARCHIVE_DELETE_FROM_DB","מחיקת רשומות מה-DB החי")],
  "progs": [("RIARCPM1","כתיבה/מחיקה של ארכוב פקודות אחזקה")],
  "s4_status": "ללא שינוי; ILM אופציונלי ב-S/4", "s4_repl": "ADMI_RUN (זהה) + SAP ILM",
  "s4_tcode": "SARA (זהה)", "fiori": "Data Archiving (אמת ID)"},
]},
]

# ============================================================================
# Build
# ============================================================================
wb = Workbook()
wb.remove(wb.active)

def safe(name):
    for ch in ':\\/?*[]':
        name = name.replace(ch, "")
    return name[:31]

# pre-compute safe sheet names for hyperlinks
sheet_names = [safe(t["title"]) for t in TOPICS]

# create dashboard placeholder first (tab order)
dash = wb.create_sheet(DASH)

index_rows = []  # (type, code, hebrew, english, sheet_name, cell)

def style_cell(c, *, bold=False, color="222222", fill=None, align_v="center",
               size=10, wrap=True):
    c.font = Font(bold=bold, size=size, color=color)
    c.alignment = Alignment(horizontal="right", vertical=align_v, wrap_text=wrap)
    c.border = BORDER
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)

# ---- topic sheets ----
for topic, sname in zip(TOPICS, sheet_names):
    th = TH[topic["theme"]]
    ws = wb.create_sheet(sname)
    ws.sheet_view.rightToLeft = True
    ws.sheet_view.showGridLines = True
    ws.sheet_properties.tabColor = th["t"]
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

    # Row 1: A1 back-to-main + title band
    back = ws.cell(1, 1, "▶ חזרה למסך ראשי")
    back.hyperlink = f"#'{DASH}'!A1"
    back.font = Font(bold=True, size=10, color="FFFFFF", underline="single")
    back.fill = PatternFill("solid", fgColor="404040")
    back.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=NCOL)
    t = ws.cell(1, 2, topic["title"] + "   |   ECC 6.0  ➜  S/4HANA")
    t.font = Font(bold=True, size=13, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=th["h"])
    t.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[1].height = 26

    # Row 2: column headers
    for j, (lbl, w) in enumerate(COLS, start=1):
        fill = S4_HDR if j >= S4_COL_START else th["s"]
        c = ws.cell(2, j, lbl)
        c.font = Font(bold=True, size=9, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
        c.fill = PatternFill("solid", fgColor=fill)
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[2].height = 40

    ws.freeze_panes = "A3"
    row = 3
    seq = 0
    for tbl in topic["tables"]:
        # table section header (full width)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOL)
        hc = ws.cell(row, 1, f"  ◆ טבלה {tbl['name']}  -  {tbl['he']}  ({tbl['en']})")
        hc.font = Font(bold=True, size=11, color="FFFFFF")
        hc.fill = PatternFill("solid", fgColor=th["h"])
        hc.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[row].height = 22
        index_rows.append(("טבלה", tbl["name"], tbl["he"], tbl["en"], sname, f"A{row}"))
        row += 1

        fstart = row
        funcs_txt = "\n".join(f"• {n}" for n, _ in tbl["funcs"])
        fdesc_txt = "\n".join(f"• {d}" for _, d in tbl["funcs"])
        prog_txt  = "\n".join(f"• {n}" for n, _ in tbl["progs"])
        pdesc_txt = "\n".join(f"• {d}" for _, d in tbl["progs"])

        for (tech, en, he) in tbl["fields"]:
            seq += 1
            band = th["b"] if (seq % 2 == 0) else None
            s4band = S4_BAND if (seq % 2 == 0) else None
            style_cell(ws.cell(row, 1, seq), bold=True, fill=band); ws.cell(row,1).alignment = Alignment(horizontal="center", vertical="center")
            # col2 table tcodes, col3 table name -> filled at fstart then merged
            style_cell(ws.cell(row, 2, ""), fill=band)
            style_cell(ws.cell(row, 3, ""), fill=band)
            style_cell(ws.cell(row, 4, tech), bold=True, color="1F3864", fill=band)
            style_cell(ws.cell(row, 5, en), fill=band)
            style_cell(ws.cell(row, 6, he), fill=band)
            for col in (7, 8, 9, 10):
                style_cell(ws.cell(row, col, ""), fill=band)
            # S/4 columns
            for col in (11, 12, 13, 14):
                style_cell(ws.cell(row, col, ""), fill=s4band)
            ws.row_dimensions[row].height = 30
            index_rows.append(("שדה", tech, he, en, sname, f"D{row}"))
            row += 1
        fend = row - 1

        # fill + vertical-merge table-level columns across [fstart..fend]
        def setmerge(col, value, *, bold=False, color="222222", s4=False):
            cell = ws.cell(fstart, col, value)
            cell.font = Font(bold=bold, size=9, color=color)
            cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
            cell.border = BORDER
            if fend > fstart:
                ws.merge_cells(start_row=fstart, start_column=col, end_row=fend, end_column=col)
        setmerge(2, tbl["tcodes"], bold=True, color="1F3864")
        setmerge(3, tbl["name"], bold=True, color="1F3864")
        setmerge(7, funcs_txt, bold=True, color="1F3864")
        setmerge(8, fdesc_txt)
        setmerge(9, prog_txt, bold=True, color="1F3864")
        setmerge(10, pdesc_txt)
        setmerge(11, tbl["s4_status"], color="7B3F00")
        setmerge(12, tbl["s4_repl"], color="7B3F00")
        setmerge(13, tbl["s4_tcode"], color="7B3F00")
        setmerge(14, tbl["fiori"], bold=True, color="C55A11")
        # index tcodes + functions + programs + fiori
        for code in [x.strip() for x in tbl["tcodes"].replace(";", ",").replace("/", ",").split(",") if x.strip()]:
            index_rows.append(("טרנזקציה", code, tbl["he"], tbl["en"], sname, f"A{fstart}"))
        for n, d in tbl["funcs"]:
            index_rows.append(("פונקציה", n, d, tbl["en"], sname, f"G{fstart}"))
        for n, d in tbl["progs"]:
            index_rows.append(("תוכנית", n, d, tbl["en"], sname, f"I{fstart}"))
        index_rows.append(("Fiori", tbl["fiori"], tbl["he"], tbl["en"], sname, f"N{fstart}"))

        row += 1  # spacer

# ============================================================================
# INDEX_DB sheet (searchable database)
# ============================================================================
idx = wb.create_sheet("INDEX_DB")
idx.sheet_view.rightToLeft = True
idx.sheet_properties.tabColor = "808080"
idx_headers = ["סוג", "קוד / שם טכני", "פירוש בעברית", "English", "גיליון", "תא", "קישור"]
for j, h in enumerate(idx_headers, start=1):
    c = idx.cell(1, j, h)
    c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="404040")
    c.alignment = Alignment(horizontal="center"); c.border = BORDER
widths = [12, 24, 34, 30, 22, 8, 14]
for j, w in enumerate(widths, start=1):
    idx.column_dimensions[get_column_letter(j)].width = w
ridx = 2
for typ, code, he, en, sn, cell in index_rows:
    idx.cell(ridx, 1, typ).alignment = Alignment(horizontal="right")
    idx.cell(ridx, 2, code).font = Font(bold=True, color="1F3864")
    idx.cell(ridx, 2).alignment = Alignment(horizontal="right")
    idx.cell(ridx, 3, he).alignment = Alignment(horizontal="right")
    idx.cell(ridx, 4, en).alignment = Alignment(horizontal="right")
    idx.cell(ridx, 5, sn).alignment = Alignment(horizontal="right")
    idx.cell(ridx, 6, cell).alignment = Alignment(horizontal="center")
    link = idx.cell(ridx, 7, '➜ פתח')
    link.value = f'=HYPERLINK("#\'"&E{ridx}&"\'!"&F{ridx},"➜ פתח")'
    link.font = Font(color="0563C1", underline="single", bold=True)
    link.alignment = Alignment(horizontal="center")
    for j in range(1, 8):
        idx.cell(ridx, j).border = BORDER
    ridx += 1
last = ridx - 1
idx.auto_filter.ref = f"A1:G{last}"
idx.freeze_panes = "A2"

# ============================================================================
# DASHBOARD
# ============================================================================
dash.sheet_view.rightToLeft = True
dash.sheet_view.showGridLines = False
dash.sheet_properties.tabColor = "111111"
for col, w in zip("ABCDEFGHIJKLM", [3,16,16,4,16,16,4,16,16,3,3,3,3]):
    dash.column_dimensions[col].width = w

dash.merge_cells("B2:I2")
d1 = dash.cell(2, 2, "SAP PM  -  מסך ניווט מרכזי  |  גשר מעבר ECC 6.0 ➜ S/4HANA")
d1.font = Font(bold=True, size=18, color="1F3864"); d1.alignment = Alignment(horizontal="right")
dash.merge_cells("B3:I3")
d2 = dash.cell(3, 2, "Interactive Migration-Ready Workbook  -  Plant Maintenance / EAM")
d2.font = Font(bold=True, size=11, color="7F7F7F"); d2.alignment = Alignment(horizontal="right")

# --- navigation cards (3 per row, 2 rows tall) ---
dash.merge_cells("B5:I5")
nh = dash.cell(5, 2, "ניווט מהיר לגיליונות  (לחץ על כרטיס כדי לעבור)")
nh.font = Font(bold=True, size=12, color="FFFFFF"); nh.fill = PatternFill("solid", fgColor="1F3864")
nh.alignment = Alignment(horizontal="right")

card_cols = [(2,3), (5,6), (8,9)]  # (startcol,endcol) for 3 cards per row
r = 6
for i, (topic, sname) in enumerate(zip(TOPICS, sheet_names)):
    th = TH[topic["theme"]]
    slot = i % 3
    if slot == 0 and i != 0:
        r += 3
    c0, c1 = card_cols[slot]
    dash.merge_cells(start_row=r, start_column=c0, end_row=r+1, end_column=c1)
    cell = dash.cell(r, c0, topic["title"])
    cell.hyperlink = f"#'{sname}'!A1"
    cell.font = Font(bold=True, size=11, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=th["h"])
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for rr in (r, r+1):
        for cc in range(c0, c1+1):
            dash.cell(rr, cc).border = BORDER
r += 3

# --- color legend ---
dash.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
lg = dash.cell(r, 2, "מקרא צבעים:  כחול=נתוני אב | ירוק=תנועתי | בורדו=קונפיגורציה | כתום=אינטגרציה ו-S/4HANA")
lg.font = Font(italic=True, size=9, color="404040"); lg.alignment = Alignment(horizontal="right")
r += 2

# --- global search ---
dash.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
sh = dash.cell(r, 2, "🔍 חיפוש גלובלי  (טבלה / טרנזקציה / שדה / פונקציה - עברית או אנגלית)")
sh.font = Font(bold=True, size=12, color="FFFFFF"); sh.fill = PatternFill("solid", fgColor=S4_HDR)
sh.alignment = Alignment(horizontal="right")
r += 1
lbl = dash.cell(r, 2, "הקלד כאן:")
lbl.font = Font(bold=True, size=11); lbl.alignment = Alignment(horizontal="right")
search_row = r
search_cell = f"C{r}"
dash.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
sc = dash.cell(r, 3, "")
sc.fill = PatternFill("solid", fgColor="FFF2CC")
sc.font = Font(bold=True, size=12, color="1F3864")
sc.alignment = Alignment(horizontal="right", vertical="center")
medium = Side(style="medium", color="C55A11")
sc.border = Border(left=medium, right=medium, top=medium, bottom=medium)
dash.row_dimensions[r].height = 24
r += 2

# results header
res_hdrs = ["קישור", "קוד / שם טכני", "פירוש בעברית", "English", "גיליון", "תא"]
for j, h in enumerate(res_hdrs):
    c = dash.cell(r, 2+j, h)
    c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="404040")
    c.alignment = Alignment(horizontal="center"); c.border = BORDER
res_row = r + 1
# spilled FILTER formula (Excel 365 / 2021 / Google Sheets)
rng = lambda col: f"INDEX_DB!${col}$2:${col}${last}"
formula = (
 f'=IF({search_cell}="","הקלד מונח לחיפוש למעלה כדי לראות תוצאות לחיצות...",'
 f'IFERROR(FILTER(CHOOSE({{1,2,3,4,5,6}},{rng("G")},{rng("B")},{rng("C")},{rng("D")},{rng("E")},{rng("F")}),'
 f'(ISNUMBER(SEARCH({search_cell},{rng("B")})))+(ISNUMBER(SEARCH({search_cell},{rng("C")})))+'
 f'(ISNUMBER(SEARCH({search_cell},{rng("D")})))+(ISNUMBER(SEARCH({search_cell},{rng("A")})))>0),'
 f'"לא נמצאו תוצאות - נסה מונח אחר"))'
)
fcell = dash.cell(res_row, 2, formula)
fcell.font = Font(size=10); fcell.alignment = Alignment(horizontal="right")
dash["B2"].value  # noop

# note about compatibility / VBA
note_r = res_row + 22
dash.merge_cells(start_row=note_r, start_column=2, end_row=note_r+3, end_column=9)
note = dash.cell(note_r, 2,
  "הערות:  1) החיפוש הדינמי עובד ב-Excel 365 / 2021 וב-Google Sheets (פונקציית FILTER). "
  "במהדורות ישנות יותר - השתמש בכרטיסיית INDEX_DB עם מסנן ה-AutoFilter המובנה.  "
  "2) להוספת מאקרו VBA אמיתי לחיפוש - ראה ההוראות והקוד בתשובת הצ'אט (שמור כקובץ .xlsm).  "
  "3) הקורפוס המאונדקס הוא S/4HANA בלבד; זהו ידע-מודל על ECC 6.0 עם מיפוי S/4 - אמת ID של אפליקציות Fiori "
  "ב-SAP Fiori Apps Reference Library ופריטים תלויי-גרסה ב-SAP Help Portal.")
note.font = Font(italic=True, size=9, color="7B1E2B")
note.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)

dash.cell(search_row, 3).value = ""  # ensure empty search

out = "SAP_PM_ECC6_to_S4_Migration.xlsx"
wb.save(out)
print(f"OK -> {out}")
print(f"   topic sheets: {len(TOPICS)} | index rows: {len(index_rows)} | dashboard + INDEX_DB")
