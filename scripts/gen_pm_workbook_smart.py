# -*- coding: utf-8 -*-
"""
SAP PM (Plant Maintenance) ECC 6.0 - INTERACTIVE master blueprint & syllabus.

Smart UX: every transaction / table / field / function / program is decoded onto its
own row (Hebrew + English). Items are grouped under collapsible category headers using
Excel's native OUTLINE grouping (+/- buttons in the left margin). Click a category
("Main T-codes", "Key Fields", ...) to expand its decoded list. No macros required -
opens the same in Excel desktop, Excel web and Google Sheets.

Scope note (repo CLAUDE.md rule 3): indexed corpus is S/4HANA only. This is explicit,
opt-in ECC 6.0 model knowledge. Verify version-specific items (table/field availability,
config transactions, list-report program names) against SAP Help Portal before production.

Run:  python3 scripts/gen_pm_workbook_smart.py
Out:  SAP_PM_ECC6_Interactive.xlsx  (repo root)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.properties import Outline, PageSetupProperties
from openpyxl.utils import get_column_letter

THEMES = {
    "master": {"hdr": "1F3864", "font": "FFFFFF", "sub": "2E5496", "band": "DCE3F2", "tab": "1F3864"},
    "trans":  {"hdr": "1E5E34", "font": "FFFFFF", "sub": "2E8B53", "band": "DCEFE0", "tab": "1E5E34"},
    "config": {"hdr": "7B1E2B", "font": "FFFFFF", "sub": "A53444", "band": "F0DCE0", "tab": "7B1E2B"},
    "cross":  {"hdr": "C55A11", "font": "FFFFFF", "sub": "E08020", "band": "FBE6D5", "tab": "C55A11"},
}

# Category (group) labels
G_MAIN = "טרנזקציות מרכזיות (Main T-codes)"
G_SEC  = "טרנזקציות משניות (Secondary T-codes)"
G_TAB  = "טבלאות ליבה (Core Tables)"
G_FLD  = "שדות מרכזיים וחשובים (Key Fields)"
G_FUN  = "פונקציות / BAPIs (Key Functions)"
G_PRG  = "תוכניות ודוחות רקע (Programs / Reports)"
GROUP_ORDER = [G_MAIN, G_SEC, G_TAB, G_FLD, G_FUN, G_PRG]

# Each item: (code, hebrew_meaning, english_meaning)
# Each topic: (topic_title, {group: [items...]})
# Each sheet: {"title","theme","topics":[...]}

DATA = [
# ============================ SHEET 1 ============================
{"title": "1. מבנה ארגוני ותשתית", "theme": "master", "topics": [
 ("מיקומים פונקציונליים (Functional Locations)", {
   G_MAIN: [
     ("IL01", "יצירת מיקום פונקציונלי", "Create Functional Location"),
     ("IL02", "שינוי מיקום פונקציונלי", "Change Functional Location"),
     ("IL03", "תצוגת מיקום פונקציונלי", "Display Functional Location"),
   ],
   G_SEC: [
     ("IL04", "יצירת מיקום מתוך תבנית / רשימה", "Create FL using reference / list"),
     ("IH01", "תצוגת מבנה היררכי של מיקומים", "Functional Location Structure (hierarchy)"),
     ("IH06", "רשימת מיקומים פונקציונליים", "Display Functional Location list"),
   ],
   G_TAB: [
     ("IFLOT", "טבלת אב של מיקומים פונקציונליים", "Functional location master record"),
     ("IFLOTX", "טקסטים (תיאורים) של מיקומים", "Functional location short texts"),
     ("ILOA", "נתוני מיקום וחיוב (מרכז עלות/מפעל)", "PM object location & account assignment"),
     ("T370F", "קטגוריות מיקום פונקציונלי", "Functional location categories"),
     ("T370S", "מחווני מבנה (Structure Indicator)", "Structure indicators for FL coding"),
   ],
   G_FLD: [
     ("TPLNR", "מספר המיקום הפונקציונלי", "Functional location number"),
     ("STRNO", "מספר מבנה (תצוגה לפי מחוון)", "Structure / formatted FL number"),
     ("TPLKZ", "מחוון מבנה הקובע מסיכת קידוד", "Structure indicator (edit mask)"),
     ("FLTYP", "קטגוריית המיקום הפונקציונלי", "Functional location category"),
     ("TPLMA", "מיקום אב בהיררכיה", "Superior functional location"),
     ("OBJNR", "מפתח אובייקט פנימי (לסטטוס)", "Internal object number (status mgmt)"),
   ],
   G_FUN: [
     ("BAPI_FUNCLOC_CREATE", "יצירת מיקום דרך ממשק תקני", "Create functional location via BAPI"),
     ("BAPI_FUNCLOC_CHANGE", "שינוי מיקום דרך ממשק תקני", "Change functional location via BAPI"),
     ("BAPI_FUNCLOC_GETDETAIL", "שליפת פרטי מיקום", "Read functional location detail"),
   ],
   G_PRG: [
     ("RIIFLO20", "תוכנית הרשימה מאחורי IH06", "Functional location list (logical DB)"),
   ],
 }),
 ("מרכזי עבודה (Work Centers)", {
   G_MAIN: [
     ("IR01", "יצירת מרכז עבודה", "Create Work Center"),
     ("IR02", "שינוי מרכז עבודה", "Change Work Center"),
     ("IR03", "תצוגת מרכז עבודה", "Display Work Center"),
   ],
   G_SEC: [
     ("CR05", "רשימת מרכזי עבודה וקיבולת", "Work center list / capacity"),
     ("CR06", "שיוך מרכז עבודה למרכז עלות", "Work center cost center assignment"),
   ],
   G_TAB: [
     ("CRHD", "כותרת מרכז עבודה", "Work center header"),
     ("CRTX", "טקסטים של מרכז עבודה", "Work center texts"),
     ("CRCA", "הקצאת קיבולת למרכז", "Work center capacity allocation"),
     ("CRCO", "שיוך מרכז עבודה לבקרה (CO)", "Work center cost center link"),
     ("KAKO", "כותרת קיבולת", "Capacity header"),
   ],
   G_FLD: [
     ("ARBPL", "קוד מרכז העבודה", "Work center code"),
     ("OBJID", "מזהה אובייקט פנימי של המרכז", "Internal work center object ID"),
     ("OBJTY", "סוג אובייקט ('A'=מרכז עבודה)", "Object type ('A'=work center)"),
     ("WERKS", "מפעל", "Plant"),
     ("VERWE", "שימוש (אילו רשימות/פק\"ע)", "Work center usage"),
     ("VERAN", "אחראי על מרכז העבודה", "Person responsible"),
   ],
   G_FUN: [
     ("CR_WORKCENTER_READ", "קריאת נתוני מרכז עבודה", "Read work center data"),
   ],
   G_PRG: [
     ("RCRACO00", "דוח מרכזי עבודה וקיבולות", "Work center / capacity report"),
   ],
 }),
 ("אישורי עבודה ומחוון מבנה (Permits & Structure Indicator)", {
   G_MAIN: [
     ("SPRO", "כניסה ל-Customizing להגדרות PM", "IMG - PM customizing entry"),
   ],
   G_SEC: [
     ("OIA1", "הגדרת מסיכות עריכה למיקומים", "Define FL edit masks"),
     ("OIBR", "הגדרת אישורי עבודה (Permits)", "Define permits"),
   ],
   G_TAB: [
     ("T370S", "מחווני מבנה (Structure Indicator)", "Structure indicators"),
     ("T357", "אישורי עבודה", "Permits"),
     ("T357G", "קבוצות אישורי עבודה", "Permit groups / categories"),
   ],
   G_FLD: [
     ("STRIN", "מחוון המבנה", "Structure indicator"),
     ("EDITM", "מסיכת עריכה (פורמט קידוד)", "Edit mask (coding format)"),
     ("HIERA", "רמות היררכיה במחוון", "Hierarchy levels"),
     ("PMGEW", "קבוצת אישורי עבודה", "Permit group"),
     ("GEWRK", "קוד אישור העבודה", "Permit code"),
   ],
   G_FUN: [
     ("PERMIT_CHECK", "בדיקת אישורי עבודה חסרים", "Check required permits"),
   ],
   G_PRG: [
     ("RIFLET00", "תחזוקת קטגוריות / מסיכות מיקום", "FL categories & mask maintenance"),
   ],
 }),
]},
# ============================ SHEET 2 ============================
{"title": "2. ציוד ונתוני מאסטר", "theme": "master", "topics": [
 ("ציוד (Equipment Master)", {
   G_MAIN: [
     ("IE01", "יצירת ציוד", "Create Equipment"),
     ("IE02", "שינוי ציוד", "Change Equipment"),
     ("IE03", "תצוגת ציוד", "Display Equipment"),
   ],
   G_SEC: [
     ("IE05", "שינוי ציוד ברשימה (List Editing)", "Change equipment (list editing)"),
     ("IE07", "רשימת ציוד רב-רמתית", "Equipment list (multi-level)"),
     ("IH08", "רשימת ציוד", "Display equipment list"),
     ("IE4N", "התקנה / פירוק ציוד", "Install / dismantle equipment"),
   ],
   G_TAB: [
     ("EQUI", "טבלת אב של ציוד", "Equipment master record"),
     ("EQKT", "טקסטים (תיאורים) של ציוד", "Equipment short texts"),
     ("EQUZ", "פלחי זמן של ציוד (התקנות)", "Equipment time segment"),
     ("ILOA", "נתוני מיקום וחיוב משותפים", "Location & account assignment"),
   ],
   G_FLD: [
     ("EQUNR", "מספר הציוד", "Equipment number"),
     ("EQTYP", "סוג / קטגוריית ציוד", "Equipment category"),
     ("EQART", "סוג טכני של הציוד", "Technical object type"),
     ("HERST", "יצרן", "Manufacturer"),
     ("TYPBZ", "דגם / סוג היצרן", "Manufacturer model"),
     ("OBJNR", "מפתח אובייקט פנימי (לסטטוס)", "Internal object number"),
   ],
   G_FUN: [
     ("BAPI_EQUI_CREATE", "יצירת ציוד דרך ממשק תקני", "Create equipment via BAPI"),
     ("BAPI_EQUI_CHANGE", "שינוי ציוד דרך ממשק תקני", "Change equipment via BAPI"),
     ("BAPI_EQUI_GETDETAIL", "שליפת פרטי ציוד", "Read equipment detail"),
   ],
   G_PRG: [
     ("RIEQUI20", "תוכנית הרשימה מאחורי IH08", "Equipment list (logical DB)"),
   ],
 }),
 ("מספרים סידוריים (Serialization)", {
   G_MAIN: [
     ("IQ01", "יצירת מספר סידורי לחומר", "Create material serial number"),
     ("IQ02", "שינוי מספר סידורי", "Change serial number"),
     ("IQ03", "תצוגת מספר סידורי", "Display serial number"),
   ],
   G_SEC: [
     ("IQ04", "יצירת מספרים סידוריים מרובים", "Create serial numbers (multiple)"),
     ("IQ08", "רשימת מספרים סידוריים", "Serial number list"),
     ("IQ09", "תצוגת מספר סידורי (רשימה)", "Display serial number list"),
   ],
   G_TAB: [
     ("OBJK", "רשימת אובייקטים / סריאלים", "Object list / serial assignment"),
     ("SER01", "כותרת מסמך סריאלי (מסירה)", "Serial numbers in delivery doc"),
     ("SER03", "סריאלים בתעודת מלאי", "Serial numbers in material doc"),
     ("EQUI", "ציוד הנוצר עבור הסריאל", "Equipment created for serial"),
   ],
   G_FLD: [
     ("SERNR", "המספר הסידורי", "Serial number"),
     ("MATNR", "מספר החומר המשויך", "Material number"),
     ("SERNP", "פרופיל סריאליזציה", "Serial number profile"),
     ("OBKNR", "מספר רשומת אובייקט", "Object list number"),
     ("LFBNR", "מסמך התנועה המקושר", "Reference material document"),
   ],
   G_FUN: [
     ("SERIAL_NUMBER_CREATE", "יצירת מספר סידורי", "Create serial number"),
     ("BAPI_OBJCL_CREATE", "שיוך סיווג לאובייקט הסריאלי", "Assign classification to object"),
   ],
   G_PRG: [
     ("RISERNR0", "דוח סריאלים מול תנועות מלאי", "Serial numbers vs. stock movements"),
   ],
 }),
 ("ירושת נתונים מהמיקום לציוד (Inheritance Logic)", {
   G_SEC: [
     ("IE02", "שינוי סעיף מיקום/חיוב בציוד", "Change equipment location/account tab"),
     ("IL02", "מקור הירושה - המיקום הפונקציונלי", "Source FL of inherited data"),
   ],
   G_TAB: [
     ("ILOA", "מרכזת נתוני מיקום/חיוב לירושה", "Holds inheritable location/account data"),
     ("EQUZ", "פלח הזמן הקובע מאיזה מיקום יורש", "Time segment linking eq. to FL"),
   ],
   G_FLD: [
     ("KOSTL", "מרכז עלות (יונק/לא יונק)", "Cost center (inherited / overridden)"),
     ("SWERK", "מפעל אחזקה (יונק)", "Maintenance plant (inherited)"),
     ("INGRP", "קבוצת תכנון (יונק)", "Planner group (inherited)"),
     ("GEWRK", "מרכז עבודה אחראי (יונק)", "Responsible work center (inherited)"),
     ("ANLNR", "נכס קבוע משויך", "Asset number"),
   ],
   G_FUN: [
     ("ILOA_READ", "קריאת נתוני מיקום/חיוב", "Read location & account data"),
   ],
   G_PRG: [
     ("RIEQUI20", "אימות עקביות ירושה בין מיקום לציוד", "Verify inheritance consistency"),
   ],
 }),
]},
# ============================ SHEET 3 ============================
{"title": "3. עצי מוצר של אחזקה (BOM)", "theme": "master", "topics": [
 ("עץ מוצר לציוד (Equipment BOM)", {
   G_MAIN: [
     ("IB01", "יצירת עץ מוצר לציוד", "Create equipment BOM"),
     ("IB02", "שינוי עץ מוצר לציוד", "Change equipment BOM"),
     ("IB03", "תצוגת עץ מוצר לציוד", "Display equipment BOM"),
   ],
   G_SEC: [
     ("IB18", "רשימת עצי מוצר של ציוד", "Equipment BOM list"),
   ],
   G_TAB: [
     ("EQST", "קישור ציוד לעץ מוצר", "Equipment-to-BOM link"),
     ("STKO", "כותרת עץ מוצר", "BOM header"),
     ("STPO", "פריטי עץ מוצר (רכיבים)", "BOM item (components)"),
   ],
   G_FLD: [
     ("STLNR", "מספר עץ מוצר", "BOM number"),
     ("STLTY", "קטגוריית BOM ('E'=ציוד)", "BOM category ('E'=equipment)"),
     ("STLAL", "חלופת עץ מוצר", "Alternative BOM"),
     ("POSNR", "מספר פריט בעץ", "BOM item number"),
     ("IDNRK", "מספר הרכיב (חומר)", "Component material"),
   ],
   G_FUN: [
     ("CSAP_MAT_BOM_MAINTAIN", "תחזוקת עץ מוצר דרך ממשק", "Maintain BOM via interface"),
     ("CS_BOM_EXPL_EQU_RC1", "פיצוץ עץ מוצר של ציוד", "Explode equipment BOM"),
   ],
   G_PRG: [
     ("RCSBI010", "פיצוץ עץ מוצר רב-רמתי", "Multi-level BOM explosion"),
   ],
 }),
 ("עץ מוצר למיקום ולחומר (FL & Material BOM)", {
   G_MAIN: [
     ("IB11", "יצירת עץ מוצר למיקום פונקציונלי", "Create functional location BOM"),
     ("CS01", "יצירת עץ מוצר לחומר", "Create material BOM"),
     ("CS02", "שינוי עץ מוצר לחומר", "Change material BOM"),
   ],
   G_SEC: [
     ("IB13", "תצוגת עץ מוצר למיקום", "Display functional location BOM"),
     ("CS11", "פיצוץ עץ מוצר רב-רמתי", "Display BOM - multilevel"),
     ("CS15", "היכן בשימוש (Where-Used)", "Where-used list"),
   ],
   G_TAB: [
     ("TPST", "קישור מיקום לעץ מוצר", "Functional location-to-BOM link"),
     ("MAST", "קישור חומר לעץ מוצר", "Material-to-BOM link"),
     ("STAS", "ניהול חלופות / שינויים בעץ", "BOM item selection / variants"),
   ],
   G_FLD: [
     ("TPLNR", "מיקום פונקציונלי", "Functional location"),
     ("MATNR", "חומר", "Material"),
     ("STLAN", "שימוש עץ מוצר (4=אחזקה)", "BOM usage (4=maintenance)"),
     ("MENGE", "כמות הרכיב", "Component quantity"),
     ("POSTP", "סוג פריט בעץ", "Item category"),
   ],
   G_FUN: [
     ("CSAP_MAT_BOM_CREATE", "יצירת עץ מוצר לחומר", "Create material BOM"),
     ("CS_BOM_EXPL_MAT_RC1", "פיצוץ עץ מוצר לחומר", "Explode material BOM"),
   ],
   G_PRG: [
     ("RCSBI020", "פיצוץ עצי מוצר של מיקומים", "FL BOM explosion"),
   ],
 }),
]},
# ============================ SHEET 4 ============================
{"title": "4. נקודות מדידה ומונים", "theme": "master", "topics": [
 ("נקודות מדידה (Measuring Points)", {
   G_MAIN: [
     ("IK01", "יצירת נקודת מדידה / מונה", "Create measuring point / counter"),
     ("IK02", "שינוי נקודת מדידה", "Change measuring point"),
     ("IK03", "תצוגת נקודת מדידה", "Display measuring point"),
   ],
   G_SEC: [
     ("IK07", "רשימת נקודות מדידה", "Measuring point list"),
     ("IK08", "שינוי נקודות מדידה ברשימה", "Change measuring points (list)"),
   ],
   G_TAB: [
     ("IMPTT", "טבלת אב של נקודות מדידה", "Measuring point master"),
   ],
   G_FLD: [
     ("POINT", "מספר נקודת המדידה", "Measuring point number"),
     ("MPOBJ", "אובייקט המדידה (ציוד/מיקום)", "Measuring object"),
     ("MPTYP", "סוג נקודת מדידה", "Measuring point category"),
     ("ATINN", "מאפיין סיווג נמדד", "Internal characteristic"),
     ("INDCT", "קוד כיוון מונה (עולה/יורד)", "Counter direction indicator"),
   ],
   G_FUN: [
     ("BAPI_MEASUREMENTPOINT_CREATE", "יצירת נקודת מדידה", "Create measuring point"),
     ("BAPI_MEASUREMENTPOINT_GETLIST", "שליפת רשימת נקודות מדידה", "Read measuring point list"),
   ],
   G_PRG: [
     ("RIMPOINT", "רשימת נקודות מדידה ומונים", "Measuring point / counter list"),
   ],
 }),
 ("מסמכי מדידה וקריאות מונה (Measurement Documents)", {
   G_MAIN: [
     ("IK11", "יצירת מסמך מדידה (קריאה)", "Create measurement document"),
     ("IK12", "שינוי מסמך מדידה", "Change measurement document"),
     ("IK13", "תצוגת מסמכי מדידה", "Display measurement documents"),
   ],
   G_SEC: [
     ("IK21", "קליטה מרוכזת של קריאות", "Collective entry of measurements"),
     ("IK22", "שינוי קליטה מרוכזת", "Collective entry (change)"),
     ("IK41", "רשימת מסמכי מדידה", "Measurement document list"),
   ],
   G_TAB: [
     ("IMRG", "מסמכי מדידה (קריאות)", "Measurement documents / readings"),
     ("IMRC", "מונה - קריאות מצטברות", "Counter cumulative readings"),
   ],
   G_FLD: [
     ("MDOCM", "מספר מסמך המדידה", "Measurement document number"),
     ("READG", "ערך הקריאה שנקלט", "Counter reading entered"),
     ("RECDV", "ערך נמדד (Reading)", "Measurement reading value"),
     ("CDIFF", "הפרש מונה מהקריאה הקודמת", "Counter difference"),
     ("CNTRC", "קריאה כוללת מצטברת (גלישה)", "Total cumulative counter reading"),
   ],
   G_FUN: [
     ("BAPI_MEASUREMENTDOCUM_CREATE", "יצירת מסמך מדידה", "Create measurement document"),
     ("BAPI_MEASUREMENTDOCUM_CREATEM", "יצירת מסמכי מדידה מרובים", "Create multiple meas. documents"),
   ],
   G_PRG: [
     ("RISTRA20", "ניטור מועדים מבוסס מונה", "Counter-based deadline monitoring"),
   ],
 }),
]},
# ============================ SHEET 5 ============================
{"title": "5. קטלוגים, קודים ופרופילים", "theme": "config", "topics": [
 ("קטלוגים, קבוצות קוד וקודים (Catalogs & Codes)", {
   G_MAIN: [
     ("QS41", "תחזוקת סוגי קטלוג", "Maintain catalog (types)"),
     ("QS51", "תחזוקת קבוצות קוד וקודים", "Maintain code groups & codes"),
   ],
   G_SEC: [
     ("QS61", "תצוגת קבוצות קוד וקודים", "Display code groups & codes"),
     ("QS42", "תחזוקת סוגי קוד", "Maintain code group catalog"),
   ],
   G_TAB: [
     ("TQ15", "סוגי קטלוג", "Catalog types"),
     ("QPGR", "קבוצות קוד", "Code groups"),
     ("QPCD", "קודים בתוך קבוצה", "Codes within group"),
     ("QPCT", "טקסטים של קודים", "Code texts"),
   ],
   G_FLD: [
     ("KATALOGART", "סוג קטלוג (B/C/D/2/5)", "Catalog type"),
     ("CODEGRUPPE", "קבוצת קוד", "Code group"),
     ("CODE", "קוד בודד", "Code"),
     ("KURZTEXT", "טקסט קצר של הקוד", "Short text"),
   ],
   G_FUN: [
     ("QPK1_CODEGROUP_READ", "קריאת קבוצת קוד", "Read code group"),
     ("QPK1_CATALOG_READ", "קריאת קטלוג", "Read catalog"),
   ],
   G_PRG: [
     ("RQEEAW01", "סקירת מבנה קטלוגים וסוגים", "Catalog structure overview"),
   ],
 }),
 ("פרופיל קטלוגים ושיוכו (Catalog Profile)", {
   G_MAIN: [
     ("OIM1", "הגדרת פרופיל קטלוגים", "Define catalog profile"),
   ],
   G_SEC: [
     ("OIN4", "שיוך פרופיל לסוג הודעה/ציוד", "Assign profile to notif./equipment"),
     ("IE02", "שיוך פרופיל קטלוג לציוד", "Assign catalog profile to equipment"),
   ],
   G_TAB: [
     ("T352", "פרופיל קטלוג", "Catalog profile"),
     ("T352B", "שיוך פרופיל לקבוצות בחירה", "Profile-to-selection-set link"),
   ],
   G_FLD: [
     ("RBNR", "פרופיל קטלוג (Report/Catalog Profile)", "Catalog profile"),
     ("AUSWAHLMGE", "קבוצת בחירה (Selection Set)", "Selection set"),
     ("QMART", "סוג הודעה המשויך", "Notification type"),
     ("EQART", "סוג טכני של ציוד", "Technical object type"),
   ],
   G_FUN: [
     ("CATALOG_PROFILE_READ", "קריאת פרופיל קטלוג", "Read catalog profile"),
   ],
   G_PRG: [
     ("RIPROFCAT", "ניתוח פרופילי קטלוג ושיוכם", "Catalog profile assignment analysis"),
   ],
 }),
]},
# ============================ SHEET 6 ============================
{"title": "6. הודעות אחזקה (Notifications)", "theme": "trans", "topics": [
 ("הודעת אחזקה - כותרת (Notification Header)", {
   G_MAIN: [
     ("IW21", "יצירת הודעת אחזקה", "Create maintenance notification"),
     ("IW22", "שינוי הודעת אחזקה", "Change maintenance notification"),
     ("IW23", "תצוגת הודעת אחזקה", "Display maintenance notification"),
   ],
   G_SEC: [
     ("IW24", "יצירת הודעת תקלה (M2)", "Create malfunction report (M2)"),
     ("IW25", "יצירת הודעת פעילות (M3)", "Create activity report (M3)"),
     ("IW28", "שינוי הודעות ברשימה", "Change notifications (list)"),
     ("IW29", "תצוגת הודעות ברשימה", "Display notifications (list)"),
   ],
   G_TAB: [
     ("QMEL", "כותרת הודעת האחזקה", "Notification header"),
   ],
   G_FLD: [
     ("QMNUM", "מספר ההודעה", "Notification number"),
     ("QMART", "סוג ההודעה (M1/M2/M3)", "Notification type"),
     ("QMTXT", "תיאור קצר של ההודעה", "Notification short text"),
     ("EQUNR", "ציוד מתייחס", "Reference equipment"),
     ("TPLNR", "מיקום פונקציונלי מתייחס", "Reference functional location"),
     ("OBJNR", "מפתח אובייקט (לסטטוס)", "Object number (status)"),
   ],
   G_FUN: [
     ("BAPI_ALM_NOTIF_CREATE", "יצירת הודעה דרך ממשק תקני", "Create notification via BAPI"),
     ("BAPI_ALM_NOTIF_SAVE", "שמירת ההודעה", "Save notification"),
     ("BAPI_ALM_NOTIF_CLOSE", "סגירת ההודעה", "Close notification"),
   ],
   G_PRG: [
     ("RIQMEL20", "תוכנית הרשימה מאחורי IW28/29", "Notification list (logical DB)"),
   ],
 }),
 ("פריטים, סיבות ופעילויות (Items/Causes/Activities)", {
   G_SEC: [
     ("IW66", "ניתוח פריטי הודעה", "Notification item analysis"),
     ("IW67", "ניתוח פעילויות הודעה", "Notification activity analysis"),
   ],
   G_TAB: [
     ("QMFE", "פריטי הודעה (ליקויים)", "Notification items (damage)"),
     ("QMUR", "סיבות לליקוי", "Notification causes"),
     ("QMMA", "פעילויות שבוצעו", "Notification activities"),
     ("QMSM", "משימות לטיפול", "Notification tasks"),
   ],
   G_FLD: [
     ("FEGRP", "קבוצת קוד ליקוי", "Damage code group"),
     ("FECOD", "קוד הליקוי", "Damage code"),
     ("OTGRP", "קבוצת אובייקט חלקי", "Object part code group"),
     ("URCOD", "קוד הסיבה", "Cause code"),
     ("MNCOD", "קוד הפעילות", "Activity code"),
   ],
   G_FUN: [
     ("BAPI_ALM_NOTIF_DATA_ADD", "הוספת פריט/סיבה/פעילות", "Add item/cause/activity"),
     ("BAPI_ALM_NOTIF_GET_DETAIL", "שליפת פרטי הודעה מלאים", "Read full notification detail"),
   ],
   G_PRG: [
     ("RIQMEL00", "ניתוח פריטי ליקוי ודפוסי כשל", "Failure / item pattern analysis"),
   ],
 }),
 ("קונפיגורציה ועדיפויות (Customizing & Priorities)", {
   G_MAIN: [
     ("OIAL", "הגדרת פריסת מסך לסוג הודעה", "Notification screen layout config"),
   ],
   G_SEC: [
     ("QCC0", "תפריט Customizing של איכות/הודעות", "QM/notification customizing menu"),
   ],
   G_TAB: [
     ("TQ80", "סוגי הודעה (Notification Types)", "Notification types"),
     ("T356", "עדיפויות", "Priorities"),
   ],
   G_FLD: [
     ("PRIOK", "עדיפות ההודעה", "Priority"),
     ("ARTPR", "סוג עדיפות", "Priority type"),
     ("AUART", "סוג פק\"ע ברירת מחדל מההודעה", "Default order type from notification"),
     ("RBNR", "פרופיל קטלוג משויך", "Catalog profile"),
   ],
   G_FUN: [
     ("PRIORITY_DETERMINE", "קביעת עדיפות ומועדי יעד", "Determine priority & dates"),
   ],
   G_PRG: [
     ("RISCREEN_NOTIF", "בקרת פריסות מסך ושדות חובה", "Screen layout / required-field check"),
   ],
 }),
]},
# ============================ SHEET 7 ============================
{"title": "7. פקודות עבודה (פק\"ע)", "theme": "trans", "topics": [
 ("פק\"ע - כותרת ומחזור חיים (Order Header)", {
   G_MAIN: [
     ("IW31", "יצירת פקודת עבודה", "Create maintenance order"),
     ("IW32", "שינוי פקודת עבודה", "Change maintenance order"),
     ("IW33", "תצוגת פקודת עבודה", "Display maintenance order"),
   ],
   G_SEC: [
     ("IW34", "יצירת פק\"ע מתוך הודעה", "Create order from notification"),
     ("IW36", "יצירת תת-פק\"ע", "Create sub-order"),
     ("IW38", "שינוי פק\"ע ברשימה", "Change orders (list)"),
     ("IW39", "תצוגת פק\"ע ברשימה", "Display orders (list)"),
   ],
   G_TAB: [
     ("AUFK", "כותרת פק\"ע (נתוני אב)", "Order master data"),
     ("AFKO", "כותרת פק\"ע (נתוני ייצור/אחזקה)", "Order header (production/PM)"),
     ("AFPO", "פריט פק\"ע", "Order item"),
   ],
   G_FLD: [
     ("AUFNR", "מספר פקודת העבודה", "Order number"),
     ("AUART", "סוג פק\"ע (PM01/PM02/PM03)", "Order type"),
     ("AUFPL", "מספר תוכנית הפעולות", "Routing/operation plan number"),
     ("GSTRP", "תאריך התחלה בסיסי", "Basic start date"),
     ("OBJNR", "מפתח אובייקט (לסטטוס)", "Object number (status)"),
   ],
   G_FUN: [
     ("BAPI_ALM_ORDER_MAINTAIN", "יצירה/שינוי פק\"ע (ממשק מרכזי)", "Create/change order (central BAPI)"),
     ("BAPI_ALM_ORDER_GET_DETAIL", "שליפת פרטי פק\"ע", "Read order detail"),
   ],
   G_PRG: [
     ("RIAUFK20", "תוכנית הרשימה מאחורי IW38/39", "Order list (logical DB)"),
   ],
 }),
 ("פעולות, רכיבים ודיווח (Operations/Components/Confirm)", {
   G_SEC: [
     ("IW37N", "פעולות ורכיבים בתצוגה משולבת", "Operations & components (list)"),
     ("IW41", "דיווח זמן על פעולה", "Time confirmation"),
     ("IW42", "דיווח השלמה כולל", "Overall completion confirmation"),
     ("CM01", "עומס קיבולת", "Capacity load evaluation"),
   ],
   G_TAB: [
     ("AFVC", "פעולות הפק\"ע", "Order operations"),
     ("AFVV", "כמויות וזמני פעולה", "Operation quantities/dates"),
     ("RESB", "הזמנת רכיבים (Reservation)", "Component reservation"),
     ("AFRU", "דיווחי ביצוע (Confirmations)", "Order confirmations"),
   ],
   G_FLD: [
     ("VORNR", "מספר הפעולה", "Operation number"),
     ("STEUS", "מפתח בקרה (פנימי/חיצוני)", "Control key (in-house/external)"),
     ("ARBEI", "עבודה מתוכננת", "Planned work"),
     ("RSNUM", "מספר הזמנת רכיבים", "Reservation number"),
     ("RUECK", "מספר דיווח ביצוע", "Confirmation counter"),
   ],
   G_FUN: [
     ("BAPI_ALM_CONF_CREATE", "יצירת דיווח ביצוע", "Create confirmation"),
     ("BAPI_ALM_ORDER_MAINTAIN", "הוספת פעולה/רכיב לפק\"ע", "Add operation/component"),
   ],
   G_PRG: [
     ("RIAFRU00", "ניתוח דיווחי ביצוע מול תכנון", "Confirmation vs. plan analysis"),
   ],
 }),
]},
# ============================ SHEET 8 ============================
{"title": "8. ניהול סטטוסים (Status Mgmt)", "theme": "config", "topics": [
 ("סטטוס מערכת (System Status)", {
   G_MAIN: [
     ("BS22", "תחזוקת סטטוסי מערכת", "Maintain system statuses"),
     ("BS23", "תצוגת סטטוסי מערכת", "Display system statuses"),
   ],
   G_SEC: [
     ("IW33", "צפייה בסטטוס פק\"ע", "Display order status"),
     ("IE03", "צפייה בסטטוס ציוד", "Display equipment status"),
   ],
   G_TAB: [
     ("JEST", "סטטוס אובייקט (פעיל/לא פעיל)", "Object status (active/inactive)"),
     ("JSTO", "כותרת אובייקט סטטוס + פרופיל", "Status object header + profile"),
     ("TJ02T", "טקסטים של סטטוסי מערכת", "System status texts"),
   ],
   G_FLD: [
     ("OBJNR", "מפתח אובייקט", "Object number"),
     ("STAT", "קוד הסטטוס", "Status code"),
     ("INACT", "מחוון לא-פעיל", "Inactive indicator"),
     ("ISTAT", "סטטוס מערכת פנימי (CRTD/REL/TECO/CLSD)", "Internal system status"),
   ],
   G_FUN: [
     ("STATUS_READ", "קריאת סטטוסי אובייקט", "Read object statuses"),
     ("STATUS_CHANGE_INTERN", "שינוי סטטוס מערכת פנימי", "Change system status (internal)"),
   ],
   G_PRG: [
     ("RISTAT00", "ניתוח סטטוסים פעילים לאובייקטים", "Active object status analysis"),
   ],
 }),
 ("סטטוס משתמש ופרופילים (User Status & Profiles)", {
   G_MAIN: [
     ("BS02", "תחזוקת פרופיל סטטוס משתמש", "Maintain status profile"),
     ("BS03", "תצוגת פרופיל סטטוס", "Display status profile"),
   ],
   G_SEC: [
     ("OIBS", "שיוך פרופיל סטטוס לסוגי PM", "Assign status profile to PM types"),
   ],
   G_TAB: [
     ("TJ30", "סטטוסי משתמש בפרופיל", "User statuses in profile"),
     ("TJ30T", "טקסטים של סטטוסי משתמש", "User status texts"),
     ("TJ20", "פרופילי סטטוס", "Status profiles"),
   ],
   G_FLD: [
     ("STSMA", "פרופיל הסטטוס", "Status profile"),
     ("ESTAT", "סטטוס משתמש", "User status"),
     ("VRGNG", "קוד תהליך עסקי מותר/חסום", "Business transaction code"),
     ("ANWND", "יישום (PM)", "Application (PM)"),
   ],
   G_FUN: [
     ("STATUS_CHANGE_EXTERN", "שינוי סטטוס משתמש", "Change user status"),
   ],
   G_PRG: [
     ("RIBS0200", "סקירת פרופילי סטטוס וחסימות", "Status profile / blocking overview"),
   ],
 }),
]},
# ============================ SHEET 9 ============================
{"title": "9. אינטגרציית מלאי ורכש (PM-MM)", "theme": "cross", "topics": [
 ("פליטת מלאי לפק\"ע (Goods Issue)", {
   G_MAIN: [
     ("MB1A", "פליטת חומר לפק\"ע (261)", "Goods issue to order (261)"),
     ("MIGO", "תנועת מלאי כללית", "Goods movement (general)"),
   ],
   G_SEC: [
     ("MB21", "יצירת הזמנת מלאי ידנית", "Create reservation manually"),
     ("MB31", "קליטת סחורה לפק\"ע", "Goods receipt for order"),
   ],
   G_TAB: [
     ("RESB", "הזמנת רכיבים", "Reservation / dependent requirements"),
     ("MSEG", "פריטי מסמך חומר", "Material document items"),
     ("MKPF", "כותרת מסמך חומר", "Material document header"),
   ],
   G_FLD: [
     ("BWART", "סוג תנועה (261/262)", "Movement type"),
     ("AUFNR", "פק\"ע מחויבת", "Order (account assignment)"),
     ("MBLNR", "מספר מסמך החומר", "Material document number"),
     ("MENGE", "כמות", "Quantity"),
   ],
   G_FUN: [
     ("BAPI_GOODSMVT_CREATE", "רישום תנועת מלאי", "Post goods movement"),
     ("BAPI_RESERVATION_CREATE1", "יצירת הזמנת מלאי", "Create reservation"),
   ],
   G_PRG: [
     ("RM07RESL", "דוח הזמנות מלאי פתוחות", "Reservation list report"),
   ],
 }),
 ("דרישות והזמנות רכש (PR & PO)", {
   G_MAIN: [
     ("ME51N", "יצירת דרישת רכש", "Create purchase requisition"),
     ("ME21N", "יצירת הזמנת רכש", "Create purchase order"),
   ],
   G_SEC: [
     ("ME57", "הקצאת מקור לדרישת רכש", "Assign source to requisition"),
     ("ME2N", "רשימת הזמנות רכש", "Purchase order list"),
   ],
   G_TAB: [
     ("EBAN", "כותרת/פריטי דרישת רכש", "Purchase requisition"),
     ("EBKN", "חיוב דרישת רכש (לפק\"ע)", "Requisition account assignment"),
     ("EKKO", "כותרת הזמנת רכש", "Purchase order header"),
     ("EKPO", "פריטי הזמנת רכש", "Purchase order items"),
   ],
   G_FLD: [
     ("BANFN", "מספר דרישת רכש", "Purchase requisition number"),
     ("EBELN", "מספר הזמנת רכש", "Purchase order number"),
     ("KNTTP", "סוג חיוב (F=פק\"ע)", "Account assignment cat. (F=order)"),
     ("LIFNR", "ספק", "Vendor"),
   ],
   G_FUN: [
     ("BAPI_PR_CREATE", "יצירת דרישת רכש", "Create purchase requisition"),
     ("BAPI_PO_CREATE1", "יצירת הזמנת רכש", "Create purchase order"),
   ],
   G_PRG: [
     ("RM06BANF", "דוח דרישות רכש פתוחות", "Open purchase requisition list"),
   ],
 }),
]},
# ============================ SHEET 10 ============================
{"title": "10. עלויות והתחשבנות (PM-CO)", "theme": "cross", "topics": [
 ("צבירת עלויות והתחשבנות (Costing & Settlement)", {
   G_MAIN: [
     ("KO88", "התחשבנות פק\"ע בודדת", "Settle order (individual)"),
   ],
   G_SEC: [
     ("KO8G", "התחשבנות מרוכזת", "Collective settlement"),
     ("KOB1", "פריטי עלות בפועל של פק\"ע", "Order actual line items"),
   ],
   G_TAB: [
     ("COSP", "סך עלויות חיצוניות (Primary)", "Cost totals - external postings"),
     ("COSS", "סך עלויות פנימיות (Secondary)", "Cost totals - internal postings"),
     ("COEP", "פריטי עלות בפועל", "CO actual line items"),
   ],
   G_FLD: [
     ("OBJNR", "אובייקט CO (הפק\"ע)", "CO object (order)"),
     ("KSTAR", "אלמנט עלות", "Cost element"),
     ("WRTTP", "סוג ערך (01=מתוכנן/04=בפועל)", "Value type (plan/actual)"),
     ("GJAHR", "שנת כספים", "Fiscal year"),
   ],
   G_FUN: [
     ("K_ORDER_SETTLEMENT", "ביצוע התחשבנות פק\"ע", "Perform order settlement"),
     ("BAPI_INTERNALORDER_GETDETAIL", "שליפת פרטי פקודה פנימית", "Read internal order detail"),
   ],
   G_PRG: [
     ("RKO7KO88", "תוכנית ההתחשבנות מאחורי KO88", "Settlement program (KO88)"),
   ],
 }),
 ("חוקי התחשבנות ויעדים (Settlement Rules & Receivers)", {
   G_MAIN: [
     ("KO02", "שינוי פקודה / חוק התחשבנות", "Change order / settlement rule"),
   ],
   G_SEC: [
     ("KSB1", "פריטי עלות במרכז עלות", "Cost center line items"),
     ("CJ20N", "בונה פרויקטים (WBS)", "Project builder (WBS)"),
   ],
   G_TAB: [
     ("COBRA", "כותרת חוק התחשבנות", "Settlement rule header"),
     ("COBRB", "פריטי חוק (חוקי חלוקה)", "Settlement / distribution rules"),
     ("PRPS", "אלמנט WBS", "WBS element master"),
   ],
   G_FLD: [
     ("BUREG", "חוק ההתחשבנות", "Settlement rule"),
     ("KONTY", "סוג מקבל (מרכז/נכס/WBS)", "Receiver category"),
     ("EMPGE", "המקבל בפועל", "Settlement receiver"),
     ("PROZS", "אחוז ההתחשבנות", "Settlement percentage"),
   ],
   G_FUN: [
     ("K_SETTLEMENT_RULE_READ", "קריאת חוק התחשבנות", "Read settlement rule"),
   ],
   G_PRG: [
     ("RKACOR40", "איתור חוקי התחשבנות חסרים", "Missing settlement rule check"),
   ],
 }),
]},
# ============================ SHEET 11 ============================
{"title": "11. אחזקה מונעת ותוכניות", "theme": "master", "topics": [
 ("רשימות פעולות אחזקה (Task Lists)", {
   G_MAIN: [
     ("IA05", "יצירת רשימת פעולות כללית", "Create general task list"),
     ("IA01", "יצירת רשימת פעולות לציוד", "Create equipment task list"),
     ("IA11", "יצירת רשימת פעולות למיקום", "Create functional location task list"),
   ],
   G_SEC: [
     ("IA06", "שינוי רשימת פעולות כללית", "Change general task list"),
     ("IA08", "שינוי רשימות פעולות (סקירה)", "Change task lists (overview)"),
   ],
   G_TAB: [
     ("PLKO", "כותרת רשימת פעולות", "Task list header"),
     ("PLPO", "פעולות ברשימה", "Task list operations"),
     ("PLAS", "שיוך פעולות לרשימה", "Task list - operation assignment"),
   ],
   G_FLD: [
     ("PLNNR", "קבוצת רשימת פעולות", "Task list group"),
     ("PLNAL", "חלופת רשימה", "Group counter / alternative"),
     ("PLNTY", "סוג רשימה (A/E/T)", "Task list type"),
     ("VORNR", "מספר פעולה", "Operation number"),
   ],
   G_FUN: [
     ("BAPI_TASKLIST_CREATE", "יצירת רשימת פעולות", "Create task list"),
     ("CARO_TASKLIST_READ", "קריאת רשימת פעולות", "Read task list"),
   ],
   G_PRG: [
     ("RIPLKO00", "דוח רשימות פעולות לפי קבוצה", "Task list report"),
   ],
 }),
 ("תוכניות אחזקה ותזמון (Maintenance Plans & Scheduling)", {
   G_MAIN: [
     ("IP01", "יצירת תכנית אחזקה", "Create maintenance plan"),
     ("IP41", "יצירת תכנית מחזור יחיד (זמן)", "Create single-cycle plan (time)"),
     ("IP42", "יצירת תכנית אסטרטגיה", "Create strategy plan"),
   ],
   G_SEC: [
     ("IP10", "תזמון תכנית אחזקה", "Schedule maintenance plan"),
     ("IP30", "ניטור מועדים ברקע", "Deadline monitoring (batch)"),
     ("IP43", "תכנית מבוססת מונים", "Multiple counter plan (performance)"),
   ],
   G_TAB: [
     ("MPLA", "כותרת תכנית אחזקה ותזמון", "Maintenance plan header"),
     ("MPOS", "פריט תכנית אחזקה", "Maintenance item"),
     ("MHIO", "אובייקטי קריאת התכנית", "Maintenance plan call objects"),
     ("MHIS", "היסטוריית תזמון", "Scheduling history"),
   ],
   G_FLD: [
     ("WARPL", "מספר תכנית האחזקה", "Maintenance plan number"),
     ("STRAT", "אסטרטגיית אחזקה", "Maintenance strategy"),
     ("ABNUM", "מספר קריאת אחזקה", "Maintenance call number"),
     ("NPLDA", "תאריך מתוכנן הבא", "Next planned date"),
   ],
   G_FUN: [
     ("ISCHED_CALL_GENERATE", "יצירת קריאת אחזקה ופק\"ע", "Generate maintenance call / order"),
     ("MAINTENANCE_PLAN_SCHEDULE", "תזמון תכנית אחזקה", "Schedule maintenance plan"),
   ],
   G_PRG: [
     ("RISTRA20", "תוכנית ניטור המועדים מאחורי IP30", "Deadline monitoring (IP30)"),
   ],
 }),
]},
# ============================ SHEET 12 ============================
{"title": "12. היסטוריה וארכיון", "theme": "trans", "topics": [
 ("היסטוריית הודעות ופק\"ע (History)", {
   G_MAIN: [
     ("IW64", "תצוגת הודעות (היסטוריה)", "Display notifications (history)"),
     ("IW39", "תצוגת פק\"ע גמורות", "Display completed orders"),
   ],
   G_SEC: [
     ("IW13", "פק\"ע לפי חומר (היכן בשימוש)", "Material where-used in orders"),
     ("IW66", "ניתוח פריטי הודעה היסטוריים", "Historical item analysis"),
   ],
   G_TAB: [
     ("QMEL", "הודעות (כולל היסטוריות)", "Notifications (incl. historical)"),
     ("AUFK", "פקודות (כולל היסטוריות)", "Orders (incl. historical)"),
     ("RIHQMEL", "מבנה לוגי לשליפת הודעות", "Logical DB structure - notifications"),
     ("RIHAUFK", "מבנה לוגי לשליפת פק\"ע", "Logical DB structure - orders"),
   ],
   G_FLD: [
     ("QMNUM", "מספר הודעה", "Notification number"),
     ("AUFNR", "מספר פק\"ע", "Order number"),
     ("AUSVN", "תחילת השבתה (Breakdown)", "Breakdown start"),
     ("MAUEH", "יחידת זמן השבתה", "Breakdown duration unit"),
   ],
   G_FUN: [
     ("BAPI_ALM_ORDERHEAD_GET_LIST", "שליפת רשימת פק\"ע", "Read order header list"),
     ("BAPI_ALM_NOTIF_GET_DETAIL", "שליפת פרטי הודעה", "Read notification detail"),
   ],
   G_PRG: [
     ("RIAUFK20", "דוח פק\"ע סגורות וניתוח עלות", "Closed order / cost analysis"),
   ],
 }),
 ("ארכוב נתונים (Archiving)", {
   G_MAIN: [
     ("SARA", "ניהול ארכוב מרכזי", "Archive administration"),
   ],
   G_SEC: [
     ("AOBJ", "הגדרת אובייקטי ארכוב", "Archiving object customizing"),
     ("DB15", "תלויות אובייקט ארכוב לטבלאות", "Archiving object / table relations"),
   ],
   G_TAB: [
     ("ADMI_RUN", "ריצות ארכוב", "Archiving runs"),
     ("ADMI_FILES", "קבצי ארכיון שנוצרו", "Generated archive files"),
   ],
   G_FLD: [
     ("OBJECT", "אובייקט ארכוב (PM_ORDER/PM_QMEL)", "Archiving object"),
     ("ARCHIV_KEY", "מפתח קובץ הארכיון", "Archive file key"),
     ("RUNID", "מזהה ריצת הארכוב", "Archiving run ID"),
   ],
   G_FUN: [
     ("ARCHIVE_OPEN_FOR_WRITE", "פתיחת ארכיון לכתיבה", "Open archive for write"),
     ("ARCHIVE_DELETE_FROM_DB", "מחיקת רשומות מה-DB החי", "Delete records from live DB"),
   ],
   G_PRG: [
     ("RIARCPM1", "כתיבה/מחיקה של ארכוב פקודות", "PM order archiving write/delete"),
   ],
 }),
]},
]

# ----------------------------------------------------------------------------
thin = Side(style="thin", color="C9C9C9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

COLS = [
    ("מס' (#)", 6),
    ("קוד / שם טכני (Code)", 26),
    ("פירוש בעברית (Hebrew)", 46),
    ("Description / Purpose (English)", 52),
]

wb = Workbook()
wb.remove(wb.active)

# ---------- Cover ----------
cov = wb.create_sheet("מפתח ראשי (Index)")
cov.sheet_view.rightToLeft = True
cov.sheet_view.showGridLines = False
cov.sheet_properties.tabColor = "404040"
cov["A1"] = "SAP PM (Plant Maintenance) ECC 6.0 - חוברת אינטראקטיבית למיישם זוטר"
cov["A1"].font = Font(bold=True, size=15, color="1F3864")
cov["A2"] = "Interactive Master Blueprint & Learning Syllabus"
cov["A2"].font = Font(bold=True, size=10, color="7F7F7F")
cov.merge_cells("A1:D1"); cov.merge_cells("A2:D2")

how = ("איך משתמשים: בכל גיליון, לחץ על כפתורי ה-(+) בשוליים השמאליים כדי לפתוח כל קטגוריה "
       "(טרנזקציות, טבלאות, שדות...). כל פריט מופיע בשורה נפרדת עם פירוש בעברית ובאנגלית. "
       "לחץ (-) כדי לכווץ. אפשר גם ללחוץ על המספרים 1/2/3 בראש השוליים כדי לפתוח/לכווץ הכל בבת אחת.")
cov["A4"] = how
cov["A4"].font = Font(size=10, color="1F3864")
cov["A4"].alignment = Alignment(wrap_text=True, horizontal="right", vertical="top")
cov.merge_cells("A4:D6")

scope = ("הערת היקף: הקורפוס המאונדקס במאגר הוא S/4HANA בלבד. חוברת זו היא ידע-מודל מפורש על ECC 6.0 "
         "(לפי בקשתך). אמת פריטים תלויי-גרסה - זמינות טבלאות/שדות, טרנזקציות קונפיגורציה, ושמות תוכניות "
         "דוח - מול SAP Help Portal לפני שימוש בייצור.")
cov["A7"] = scope
cov["A7"].font = Font(italic=True, size=9, color="7B1E2B")
cov["A7"].alignment = Alignment(wrap_text=True, horizontal="right", vertical="top")
cov.merge_cells("A7:D9")

cov.column_dimensions["A"].width = 30
for c in "BCD":
    cov.column_dimensions[c].width = 26

CAT = {"master": "נתוני אב (Master Data)", "trans": "תנועתי (Transactional)",
       "config": "קונפיגורציה (Customizing)", "cross": "אינטגרציה חוצת-מודולים (Cross-Module)"}
hr = 11
cov.cell(hr, 1, "גיליון").font = Font(bold=True, color="FFFFFF")
cov.cell(hr, 2, "קטגוריה").font = Font(bold=True, color="FFFFFF")
cov.cell(hr, 3, "מספר נושאים").font = Font(bold=True, color="FFFFFF")
for col in (1, 2, 3):
    cc = cov.cell(hr, col)
    cc.fill = PatternFill("solid", fgColor="404040")
    cc.alignment = Alignment(horizontal="right"); cc.border = border
r = hr + 1
for sh in DATA:
    th = THEMES[sh["theme"]]
    cov.cell(r, 1, sh["title"]).alignment = Alignment(horizontal="right")
    c2 = cov.cell(r, 2, CAT[sh["theme"]]); c2.alignment = Alignment(horizontal="right")
    c2.fill = PatternFill("solid", fgColor=th["band"])
    cov.cell(r, 3, len(sh["topics"])).alignment = Alignment(horizontal="center")
    for col in (1, 2, 3):
        cov.cell(r, col).border = border
    r += 1

# ---------- Sheets ----------
def safe(name):
    for ch in ':\\/?*[]':
        name = name.replace(ch, "")
    return name[:31]

for sh in DATA:
    th = THEMES[sh["theme"]]
    ws = wb.create_sheet(safe(sh["title"]))
    ws.sheet_view.rightToLeft = True
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = th["tab"]
    # outline: summary (header) ABOVE its detail rows
    ws.sheet_properties.outlinePr = Outline(summaryBelow=False, summaryRight=False, applyStyles=False)
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

    ncol = len(COLS)
    # Title band
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    t = ws.cell(1, 1, sh["title"])
    t.font = Font(bold=True, size=14, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=th["hdr"])
    t.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[1].height = 26
    # Column header
    for j, (lbl, w) in enumerate(COLS, start=1):
        c = ws.cell(2, j, lbl)
        c.font = Font(bold=True, size=10, color=th["font"])
        c.fill = PatternFill("solid", fgColor=th["sub"])
        c.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        c.border = border
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[2].height = 22

    row = 3
    for topic_title, groups in sh["topics"]:
        # Topic header (outline level 0)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncol)
        tc = ws.cell(row, 1, "  " + topic_title)
        tc.font = Font(bold=True, size=12, color="FFFFFF")
        tc.fill = PatternFill("solid", fgColor=th["hdr"])
        tc.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[row].height = 24
        ws.row_dimensions[row].outline_level = 0
        row += 1

        for gname in GROUP_ORDER:
            items = groups.get(gname)
            if not items:
                continue
            # Category subheader (outline level 1)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncol)
            sc = ws.cell(row, 1, "    " + gname)
            sc.font = Font(bold=True, size=10, color="FFFFFF")
            sc.fill = PatternFill("solid", fgColor=th["sub"])
            sc.alignment = Alignment(horizontal="right", vertical="center")
            ws.row_dimensions[row].height = 20
            ws.row_dimensions[row].outline_level = 1
            row += 1
            # Items (outline level 2, hidden -> collapsed by default)
            for idx, (code, he, en) in enumerate(items, start=1):
                vals = [idx, code, he, en]
                for j, v in enumerate(vals, start=1):
                    c = ws.cell(row, j, v)
                    bold = (j == 2)
                    c.font = Font(bold=bold, size=10,
                                  color="1F3864" if bold else "222222")
                    c.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
                    c.border = border
                    if idx % 2 == 0:
                        c.fill = PatternFill("solid", fgColor=th["band"])
                ws.row_dimensions[row].outline_level = 2
                ws.row_dimensions[row].hidden = True  # collapsed by default -> click + to open
                row += 1

    ws.freeze_panes = "A3"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

out = "SAP_PM_ECC6_Interactive.xlsx"
wb.save(out)
n_topics = sum(len(s["topics"]) for s in DATA)
print(f"OK -> {out}  ({len(DATA)} sheets, {n_topics} topics)")
